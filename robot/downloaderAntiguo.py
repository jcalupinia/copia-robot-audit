from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
from pathlib import Path
from typing import Optional
from urllib.parse import urlencode
import pandas as pd
import csv, re, json, os, time, unicodedata, html
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill
import xml.etree.ElementTree as ET

from robot.parser import construir_reporte

from typing import Callable

from robot.captcha_solver import (
    CaptchaSolverError,
    MAX_ATTEMPTS as CAPTCHA_MAX_ATTEMPTS,
    is_enabled as captcha_solver_enabled,
    solve_image as solve_captcha_image,
)

USER_NOTIFICATION_CALLBACK: Optional[Callable[[str], None]] = None

def set_user_notifier(callback: Optional[Callable[[str], None]]):
    global USER_NOTIFICATION_CALLBACK
    USER_NOTIFICATION_CALLBACK = callback

# ====== Configuracion global ======
if os.name == "nt":
    local_app = os.getenv("LOCALAPPDATA")
    if local_app:
        pw_path = os.path.join(local_app, "ms-playwright")
        os.environ.setdefault("PLAYWRIGHT_BROWSERS_PATH", pw_path)
        os.environ.setdefault("PYPPETEER_HOME", pw_path)
else:
    os.environ.setdefault("PLAYWRIGHT_BROWSERS_PATH", "/root/.cache/ms-playwright")
    os.environ.setdefault("PYPPETEER_HOME", "/root/.cache/ms-playwright")

HEADLESS_ENV = os.getenv("PLAYWRIGHT_HEADLESS", "0").strip().lower()
HEADLESS = HEADLESS_ENV not in {"0", "false", "no", "off"}
try:
    SLOW_MO = int(os.getenv("PLAYWRIGHT_SLOWMO", "0"))
except ValueError:
    SLOW_MO = 0
try:
    DOWNLOAD_TIMEOUT = int(os.getenv("SRI_DOWNLOAD_TIMEOUT_MS", "120000"))
except (TypeError, ValueError):
    DOWNLOAD_TIMEOUT = 120000
PAUSE_AT_LOGIN_ENV = os.getenv("PAUSE_BEFORE_INGRESAR", "0").strip().lower()
PAUSE_AT_LOGIN = PAUSE_AT_LOGIN_ENV in {"1", "true", "yes", "on"}
PAUSE_PROMPT = os.getenv(
    "PAUSE_BEFORE_INGRESAR_PROMPT",
    "Pausa antes de hacer clic en 'Ingresar'. Realiza los cambios necesarios y presiona Enter para continuar.",
).strip()

URLS = {
    "Recibidos": "https://srienlinea.sri.gob.ec/comprobantes-electronicos-internet/pages/consultas/recibidos/comprobantesRecibidos.jsf",
    "Emitidos":  "https://srienlinea.sri.gob.ec/comprobantes-electronicos-internet/pages/consultas/emitidos/comprobantesEmitidos.jsf",
}
RECIBIDOS_DIRECT_URL = "https://srienlinea.sri.gob.ec/comprobantes-electronicos-internet/pages/consultas/recibidos/comprobantesRecibidos.jsf"
RECUPERAR_COMPROBANTES_URL = "https://srienlinea.sri.gob.ec/comprobantes-electronicos-internet/pages/consultas/recuperarComprobantes.jsf"
MENU_URL = "https://srienlinea.sri.gob.ec/comprobantes-electronicos-internet/pages/consultas/menu.jsf"
MENU_URL_ALT = (
    "https://srienlinea.sri.gob.ec/comprobantes-electronicos-internet/pages/consultas/menu.jsf?&contextoMPT=https://srienlinea.sri.gob.ec/tuportal-internet&pathMPT=Facturacion%20Electronica%20%2F%20Produccion&actualMPT=Consultas%20&linkMPT=%2Fcomprobantes-electronicos-internet%2Fpages%2Fconsultas%2Fmenu.jsf%3F&esFavorito=S#"
)
MENU_EMITIDOS_TRIGGER_URL = "https://srienlinea.sri.gob.ec/comprobantes-electronicos-internet/pages/consultas/menu.jsf?&contextoMPT=https://srienlinea.sri.gob.ec/tuportal-internet"
PORTAL_HOME = "https://srienlinea.sri.gob.ec/auth/realms/Internet/protocol/openid-connect/auth?client_id=app-sri-claves-angular&redirect_uri=https%3A%2F%2Fsrienlinea.sri.gob.ec%2Fsri-en-linea%2F%2Fcontribuyente%2Fperfil&state=34e5716b-3474-46e7-8c52-ddfe62a2404c&nonce=46d2f0a2-ce75-4cec-856d-987329a6f17e&response_mode=fragment&response_type=code&scope=openid"
MENU_TOGGLE_SELECTOR = "#sri-menu, button#sri-menu, button.menu-button, button[aria-label*='menu']"
FACTURACION_MENU_SELECTOR = "xpath=//a[.//span[contains(@class,'ui-menuitem-text') and normalize-space()='FACTURACIÓN ELECTRÓNICA']]"
MODULO_PRODUCCION_SELECTOR = "xpath=//span[contains(@class,'ui-menuitem-text') and normalize-space()='Producción']/ancestor::a[1]"
CONSULTAS_SELECTOR = "xpath=//span[contains(@class,'ui-menuitem-text') and normalize-space()='Consultas']/ancestor::a[1]"
OVERLAY_SELECTORS = ["#disablingDiv", "#disablingOverlay"]
AUTORIZACION_COMPROBANTES_SOAP_URL = "https://cel.sri.gob.ec/comprobantes-electronicos-ws/AutorizacionComprobantesOffline"
PORTAL_INDISPONIBLE_MENSAJE = (
    "El portal del SRI reporta indisponibilidad temporal. Intenta nuevamente en unos minutos."
)


TIPOS_MAP = {
    "Facturas": "Factura",
    "Retenciones": "Comprobante de Retencion",
    "Notas de credito": "Notas de Credito",
    "Notas de credito": "Notas de Credito",
    "Notas de debito": "Notas de Debito",
    "Notas de debito": "Notas de Debito",
    "Liquidacion de compra": "Liquidacion de compra de bienes y prestacion de servicios",
    "Liquidacion de compra": "Liquidacion de compra de bienes y prestacion de servicios",
    "Guia de Remision": "Guia de Remision",
    "Guia de Remision": "Guia de Remision",
    "Guias de Remision": "Guia de Remision",
    "Guias de Remision": "Guia de Remision",
    "Guia de remision": "Guia de Remision",
    "Guia de remision": "Guia de Remision",
    "Guias de remision": "Guia de Remision",
    "Guias de remision": "Guia de Remision",
}




ESTADOS_EMITIDOS_MAP = {
    "Autorizados": "Autorizados",
    "Autorizado": "Autorizados",
    "No Autorizados": "No Autorizados",
    "No autorizados": "No Autorizados",
    "Por Procesar": "Por Procesar",
    "Por procesar": "Por Procesar",
}

DOC_LABELS = {
    "01": "Factura",
    "03": "Liquidacion_de_Compra",
    "04": "NotaCredito",
    "05": "NotaDebito",
    "06": "GuiaRemision",
    "07": "Retencion",
}

EMITIDOS_FECHA_SELECTORS = [
    "input[id$='fecha_input']",
    "input[name$='fecha_input']",
    "input[id*='fechaEmision']",
    "input[name*='fechaEmision']",
    "input[id*='calFecha']",
]

EMITIDOS_ESTABLECIMIENTO_SELECTORS = [
    "input[id$='establecimiento']",
    "input[name$='establecimiento']",
    "select[id$='establecimiento']",
    "select[name$='establecimiento']",
]

EMITIDOS_PUNTO_SELECTORS = [
    "input[id$='ptoEmision']",
    "input[name$='ptoEmision']",
    "input[id$='punto']",
    "input[name$='punto']",
]



# ====== Funciones auxiliares ======
def _mes_a_texto(mes: int) -> str:
    return ["Enero","Febrero","Marzo","Abril","Mayo","Junio",
            "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"][mes-1]

def _es_clave(valor: str) -> bool:
    return bool(re.fullmatch(r"\d{49}", (valor or "").strip()))

def _detectar_delimitador(sample: str) -> str:
    counts = { ';': sample.count(';'), ',': sample.count(','), '\t': sample.count('\t') }
    return max(counts, key=counts.get) if any(counts.values()) else ';'

def _extraer_claves_desde_txt(txt_path: Path):
    claves = []
    sample = txt_path.read_text(encoding="utf-8", errors="ignore")[:4096]
    sep = _detectar_delimitador(sample)
    with open(txt_path, "r", encoding="utf-8", errors="ignore") as f:
        reader = csv.reader(f, delimiter=sep)
        for row in reader:
            if not row: 
                continue
            clave = next((c.strip() for c in row if _es_clave(c)), None)
            if not clave:
                continue
            tipo = next((c.strip() for c in row if c.lower().startswith(("factura","comprobante","nota","liquidacion"))), "")
            fecha = next((c.strip() for c in row if re.fullmatch(r"\d{2}/\d{2}/\d{4}", c.strip())), "")
            claves.append({"clave": clave, "tipo": tipo, "fecha": fecha})
    return claves

def _sanear_nombre_archivo(texto: str, sufijo: str = "") -> str:
    base = unicodedata.normalize("NFKD", texto or "").encode("ascii", "ignore").decode("ascii")
    base = re.sub(r"[^A-Za-z0-9._-]+", "_", base).strip("_")
    if not base:
        base = "documento"
    if sufijo:
        base = f"{base}_{sufijo}"
    return base


def _nombre_carpeta_tipo(tipo: str) -> str:
    base = unicodedata.normalize("NFKD", (tipo or "")).encode("ascii", "ignore").decode("ascii")
    base = re.sub(r"[^A-Za-z0-9]+", "_", base).strip("_")
    return base or "Otros"


def _slug_tipo(tipo: str) -> str:
    return _nombre_carpeta_tipo(tipo).lower()


TIPO_LABEL_MAP = {
    "factura": (1, "Factura"),
    "facturas": (1, "Factura"),
    "liquidacion_de_compra": (2, "Liquidacion_de_Compra"),
    "liquidacion_de_compra_de_bienes_y_prestacion_de_servicios": (2, "Liquidacion_de_Compra"),
    "comprobante_de_retencion": (6, "Retencion"),
    "comprobantes_de_retencion": (6, "Retencion"),
    "retencion": (6, "Retencion"),
    "retenciones": (6, "Retencion"),
    "nota_de_credito": (3, "NotaCredito"),
    "notas_de_credito": (3, "NotaCredito"),
    "nota_credito": (3, "NotaCredito"),
    "notas_credito": (3, "NotaCredito"),
    "nota_de_debito": (4, "NotaDebito"),
    "notas_de_debito": (4, "NotaDebito"),
    "nota_debito": (4, "NotaDebito"),
    "notas_debito": (4, "NotaDebito"),
    "guia_de_remision": (5, "GuiaRemision"),
    "guias_de_remision": (5, "GuiaRemision"),
    "guia_remision": (5, "GuiaRemision"),
    "comprobante_de_retencion_venta": (6, "Retencion"),
}


def _normalizar_tipo_clave(texto: str) -> str:
    base = unicodedata.normalize("NFKD", (texto or "")).encode("ascii", "ignore").decode("ascii")
    base = re.sub(r"[^A-Za-z0-9]+", "_", base).strip("_").lower()
    return base


def _formatear_label(texto: str) -> str:
    if not texto:
        return "Documentos"
    partes = [fragment.capitalize() for fragment in texto.split("_") if fragment]
    return "_".join(partes) if partes else (texto or "Documentos")


def _resolver_tipo_label(tipo_texto: str) -> tuple[int, str]:
    clave = _normalizar_tipo_clave(tipo_texto)
    if clave in TIPO_LABEL_MAP:
        return TIPO_LABEL_MAP[clave]
    if clave.endswith("s"):
        clave_singular = clave.rstrip("s")
        if clave_singular in TIPO_LABEL_MAP:
            return TIPO_LABEL_MAP[clave_singular]
    label_sanitizado = _nombre_carpeta_tipo(tipo_texto)
    label = _formatear_label(label_sanitizado)
    return 99, label


def _prefijo_tipo(tipo_texto: str) -> tuple[int, str, str]:
    orden, etiqueta = _resolver_tipo_label(tipo_texto)
    prefijo = f"{orden:02d}_{etiqueta}"
    return orden, etiqueta, prefijo

def _portal_indisponible(page) -> bool:
    try:
        contenido = page.content()
    except Exception:
        return False
    if not contenido:
        return False
    texto = unicodedata.normalize("NFKD", contenido).lower()
    return "ha ocurrido un error" in texto and "indisponibil" in texto


def _asegurar_portal_disponible(page):
    if _portal_indisponible(page):
        raise RuntimeError(PORTAL_INDISPONIBLE_MENSAJE)


def _obtener_view_state(page) -> str:
    try:
        view_input = page.locator("input[name='javax.faces.ViewState']")
        if view_input.count():
            return (view_input.first.input_value() or "").strip()
    except Exception:
        pass
    return ""


def _actualizar_view_state_input(page, nuevo_view_state: str):
    if not nuevo_view_state:
        return
    try:
        page.locator("input[name='javax.faces.ViewState']").evaluate(
            "(el, val) => { el.value = val; }",
            nuevo_view_state,
        )
    except Exception:
        pass


def _notificar_usuario_captcha(tipo: str, contexto: str):
    mensaje = (
        f"[ACCION] Se detecto un {tipo} en '{contexto}'. "
        'Resuelvelo manualmente en la ventana del navegador y luego continua.'
    )
    print(mensaje)
    if USER_NOTIFICATION_CALLBACK:
        try:
            USER_NOTIFICATION_CALLBACK(mensaje)
        except Exception as err:
            print(f"[WARN] No se pudo enviar notificacion al UI: {err}")


def _obtener_form_base_emitidos(page):
    try:
        datos = page.evaluate(
            """() => {
                const form = document.querySelector('form#frmPrincipal');
                if (!form) { return {}; }
                const payload = {};
                const elementos = form.querySelectorAll('input, select, textarea');
                for (const el of elementos) {
                    if (!el.name) { continue; }
                    if ((el.type === 'checkbox' || el.type === 'radio') && !el.checked) {
                        continue;
                    }
                    payload[el.name] = el.value ?? '';
                }
                return payload;
            }"""
        ) or {}
    except Exception:
        datos = {}
    datos.pop("javax.faces.ViewState", None)
    return datos


def _extraer_autorizacion_desde_partial(respuesta: str):
    if not respuesta:
        return None, None
    autorizacion_bruta = None
    cdata_matches = re.findall(r"<!\[CDATA\[(.*?)\]\]>", respuesta, flags=re.DOTALL)
    for chunk in cdata_matches:
        texto = chunk.strip()
        if not texto:
            continue
        if "&lt;autorizacion" in texto and "<autorizacion" not in texto:
            texto = html.unescape(texto)
        if "<autorizacion" in texto:
            autorizacion_bruta = texto
            break
    if autorizacion_bruta is None:
        desen = html.unescape(respuesta)
        match = re.search(r"(<autorizacion.*?</autorizacion>)", desen, flags=re.DOTALL | re.IGNORECASE)
        if match:
            autorizacion_bruta = match.group(1)
    view_state_match = re.search(
        r'<update id="javax\.faces\.ViewState"><!\[CDATA\[(.*?)\]\]></update>',
        respuesta,
        flags=re.DOTALL,
    )
    nuevo_view_state = view_state_match.group(1).strip() if view_state_match else None
    return autorizacion_bruta, nuevo_view_state


def _strip_xml_namespaces(element: ET.Element):
    if element is None:
        return
    for node in element.iter():
        if '}' in node.tag:
            node.tag = node.tag.split('}', 1)[1]
        if ':' in node.tag:
            node.tag = node.tag.split(':', 1)[1]
        if node.attrib:
            node.attrib = {
                key.split('}', 1)[-1].split(':', 1)[-1]: val
                for key, val in node.attrib.items()
            }


def _limpiar_cdata(texto: str) -> str:
    if not texto:
        return ""
    contenido = texto.strip()
    if contenido.startswith("<![CDATA[") and contenido.endswith("]]>"):
        contenido = contenido[9:-3]
    return contenido.strip()


def _buscar_autorizacion_en_json(payload):
    if isinstance(payload, dict):
        for val in payload.values():
            encontrado = _buscar_autorizacion_en_json(val)
            if encontrado:
                return encontrado
    elif isinstance(payload, list):
        for item in payload:
            encontrado = _buscar_autorizacion_en_json(item)
            if encontrado:
                return encontrado
    elif isinstance(payload, str) and "<autorizacion" in payload:
        return payload
    return None


def _es_url_autorizacion(url: str) -> bool:
    if not url:
        return False
    url_lower = url.lower()
    if "sri.gob.ec" not in url_lower:
        return False
    return "autoriz" in url_lower


def _extraer_comprobante_desde_autorizacion(payload: str):
    if not payload:
        raise ValueError("Respuesta de autorizacion vacia.")
    bruto = payload.strip()
    if bruto.startswith("{"):
        try:
            data = json.loads(bruto)
            posible = _buscar_autorizacion_en_json(data)
            if isinstance(posible, str) and "<autorizacion" in posible:
                bruto = posible
        except json.JSONDecodeError:
            pass
    bruto = html.unescape(bruto)
    try:
        root = ET.fromstring(bruto)
    except ET.ParseError as err:
        match = re.search(r"(<autorizaciones?.*?</autorizaciones?>)", bruto, re.DOTALL | re.IGNORECASE)
        if match:
            try:
                root = ET.fromstring(match.group(1))
            except ET.ParseError as inner_err:
                raise ValueError("No se pudo interpretar la respuesta de autorizacion.") from inner_err
        else:
            contenido_match = re.search(r"<comprobante[^>]*>(.*?)</comprobante>", bruto, re.DOTALL | re.IGNORECASE)
            if contenido_match:
                return _limpiar_cdata(contenido_match.group(1)), {}
            raise ValueError("No se pudo interpretar la respuesta de autorizacion.") from err
    _strip_xml_namespaces(root)
    candidatos = []
    if root.tag.lower().endswith("autorizacion"):
        candidatos = [root]
    else:
        candidatos = list(root.findall(".//autorizacion"))
    for autorizacion in candidatos:
        comprobante_texto = autorizacion.findtext("comprobante", "")
        if not comprobante_texto:
            continue
        meta = {
            "estado": (autorizacion.findtext("estado") or "").strip(),
            "numero_autorizacion": (autorizacion.findtext("numeroAutorizacion") or "").strip(),
            "fecha_autorizacion": (autorizacion.findtext("fechaAutorizacion") or "").strip(),
            "ambiente": (autorizacion.findtext("ambiente") or "").strip(),
        }
        return _limpiar_cdata(comprobante_texto), meta
    raise ValueError("La respuesta de autorizacion no contiene un comprobante valido.")


def _primer_texto(node: ET.Element, nombres):
    if node is None:
        return ""
    for nombre in nombres:
        valor = node.findtext(nombre)
        if valor and valor.strip():
            return valor.strip()
    return ""


def _fecha_slug(fecha: str) -> str:
    if not fecha:
        return ""
    texto = fecha.strip()
    if not texto:
        return ""
    m = re.search(r"(\d{2})/(\d{2})/(\d{4})", texto)
    if m:
        return f"{m.group(3)}{m.group(2)}{m.group(1)}"
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", texto)
    if m:
        return f"{m.group(1)}{m.group(2)}{m.group(3)}"
    m = re.search(r"(\d{4})(\d{2})(\d{2})", texto)
    if m:
        return f"{m.group(1)}{m.group(2)}{m.group(3)}"
    solo_digitos = re.sub(r"[^\d]", "", texto)
    return solo_digitos[:8]


def _total_slug(valor) -> str:
    if isinstance(valor, (int, float)):
        numero = float(valor)
    else:
        numero = _parse_decimal(valor) if isinstance(valor, str) else None
        if numero is None:
            return ""
    return f"{numero:.2f}".replace(".", "_")


def _parse_emitido_comprobante(xml_texto: str, meta_autorizacion: Optional[dict] = None):
    contenido = _limpiar_cdata(xml_texto)
    if not contenido:
        raise ValueError("El comprobante emitido llego vacio.")
    try:
        root = ET.fromstring(contenido)
    except ET.ParseError as err:
        raise ValueError(f"XML de comprobante emitido invalido: {err}") from err
    _strip_xml_namespaces(root)
    info_trib = root.find("infoTributaria")
    if info_trib is None:
        raise ValueError("El comprobante emitido no contiene infoTributaria.")
    cod_doc = (info_trib.findtext("codDoc") or "").strip()
    estab = (info_trib.findtext("estab") or "").strip()
    pto_emi = (info_trib.findtext("ptoEmi") or "").strip()
    secuencial = (info_trib.findtext("secuencial") or "").strip()
    clave = (info_trib.findtext("claveAcceso") or "").strip()

    info_node = None
    for candidato in (
        "infoFactura",
        "infoLiquidacionCompra",
        "infoNotaCredito",
        "infoNotaDebito",
        "infoCompRetencion",
        "infoGuiaRemision",
    ):
        info_node = root.find(candidato)
        if info_node is not None:
            break

    fecha_emision = ""
    identificacion = ""
    razon_receptor = ""
    importe_total = ""
    if info_node is not None:
        fecha_emision = (info_node.findtext("fechaEmision") or "").strip()
        identificacion = _primer_texto(
            info_node,
            [
                "identificacionComprador",
                "identificacionSujetoRetenido",
                "identificacionBeneficiario",
                "identificacionCliente",
                "identificacionDestinatario",
            ],
        )
        razon_receptor = _primer_texto(
            info_node,
            [
                "razonSocialComprador",
                "razonSocialSujetoRetenido",
                "razonSocialBeneficiario",
                "razonSocialCliente",
                "razonSocialDestinatario",
                "nombreComercialComprador",
                "nombreComprador",
            ],
        )
        importe_total = _primer_texto(
            info_node,
            [
                "importeTotal",
                "valorTotal",
                "total",
                "totalPagar",
                "valorModificacion",
                "valorRetenido",
            ],
        )
        if not importe_total:
            importe_total = _primer_texto(info_node, ["totalSinImpuestos"])

    meta = {
        "cod_doc": cod_doc,
        "doc_label": DOC_LABELS.get(cod_doc, cod_doc or "Documento"),
        "estab": estab,
        "pto_emi": pto_emi,
        "secuencial": secuencial,
        "clave_acceso": clave,
        "identificacion_receptor": identificacion,
        "razon_social_receptor": razon_receptor,
        "fecha_emision": fecha_emision,
        "importe_total": importe_total,
        "xml_contenido": contenido,
    }
    if meta_autorizacion:
        meta["estado_autorizacion"] = meta_autorizacion.get("estado", "")
        meta["numero_autorizacion"] = meta_autorizacion.get("numero_autorizacion", "")
        meta["fecha_autorizacion"] = meta_autorizacion.get("fecha_autorizacion", "")
        meta["ambiente"] = meta_autorizacion.get("ambiente", "")
    return meta


def _construir_nombre_xml_emitido(meta: dict, fallback: str) -> str:
    serie = "-".join(
        parte for parte in (meta.get("estab"), meta.get("pto_emi"), meta.get("secuencial")) if parte
    )
    partes = []
    doc_label = meta.get("doc_label")
    if doc_label:
        partes.append(doc_label)
    if serie:
        partes.append(serie)
    identificacion = meta.get("identificacion_receptor")
    if identificacion:
        partes.append(identificacion)
    fecha_token = _fecha_slug(meta.get("fecha_emision", ""))
    if fecha_token:
        partes.append(fecha_token)
    total_token = _total_slug(meta.get("importe_total", ""))
    if total_token:
        partes.append(f"total-{total_token}")
    if not partes:
        fallback_nombre = fallback or meta.get("clave_acceso") or meta.get("secuencial") or "emitido"
        partes.append(fallback_nombre)
    nombre = _sanear_nombre_archivo("_".join(partes))
    if len(nombre) > 180:
        nombre = nombre[:180].rstrip("._-")
    return nombre or "emitido"


def _capturar_xml_emitido(
    page,
    request_context,
    source_id: str,
    xml_dir: Path,
    fallback_nombre: str,
    claves_guardadas: set[str],
    payload_base: dict,
    view_state: str,
    headers: dict,
):
    if not source_id:
        return None, view_state
    dialog_error = None
    try:
        resultado_dialogo, view_state_dialogo = _capturar_xml_emitido_por_dialogo(
            page,
            request_context,
            source_id,
            xml_dir,
            fallback_nombre,
            claves_guardadas,
            headers,
        )
        return resultado_dialogo, view_state_dialogo
    except Exception as err:
        dialog_error = err
    payload = dict(payload_base or {})
    payload.update(
        {
            "javax.faces.partial.ajax": "true",
            "javax.faces.source": source_id,
            "javax.faces.partial.execute": source_id,
            "javax.faces.partial.render": "form-detalle-factura:panel-detalle-factura",
            source_id: source_id,
            "javax.faces.ViewState": view_state or "",
        }
    )
    payload.setdefault("frmPrincipal", "frmPrincipal")

    respuesta = request_context.post(
        RECUPERAR_COMPROBANTES_URL,
        data=payload,
        headers=headers,
    )
    if respuesta.status != 200:
        raise RuntimeError(f"Error HTTP {respuesta.status} al recuperar comprobante.")
    cuerpo = respuesta.text()
    autorizacion_bruta, nuevo_view_state = _extraer_autorizacion_desde_partial(cuerpo)
    if not autorizacion_bruta:
        if dialog_error:
            raise RuntimeError("No se pudo extraer la autorizacion del comprobante.") from dialog_error
        raise RuntimeError("No se pudo extraer la autorizacion del comprobante.")

    comprobante_xml, meta_aut = _extraer_comprobante_desde_autorizacion(autorizacion_bruta)
    meta = _parse_emitido_comprobante(comprobante_xml, meta_aut)
    clave = meta.get("clave_acceso")
    if clave and clave in claves_guardadas:
        return None, nuevo_view_state or view_state

    nombre_archivo = _construir_nombre_xml_emitido(meta, fallback_nombre)
    destino = xml_dir / f"{nombre_archivo}.xml"
    sufijo = 1
    while destino.exists():
        destino = xml_dir / f"{nombre_archivo}_{sufijo}.xml"
        sufijo += 1

    contenido = meta.get("xml_contenido", "").strip()
    if not contenido:
        raise ValueError("El comprobante retornado esta vacio.")
    if not contenido.startswith("<?xml"):
        contenido = '<?xml version="1.0" encoding="UTF-8"?>\n' + contenido
    destino.write_text(contenido, encoding="utf-8")
    if clave:
        claves_guardadas.add(clave)

    if nuevo_view_state:
        _actualizar_view_state_input(page, nuevo_view_state)
    return destino, nuevo_view_state or view_state




def _capturar_xml_emitido_por_dialogo(
    page,
    request_context,
    source_id: str,
    xml_dir: Path,
    fallback_nombre: str,
    claves_guardadas: set[str],
    base_headers: Optional[dict] = None,
) -> tuple[Optional[Path], str]:
    if not source_id:
        return None, _obtener_view_state(page)

    dialog_actual = None
    resultado_path: Optional[Path] = None

    try:
        disparador = None
        candidatos = [
            f"[id='{source_id}']",
            f"a[id='{source_id}']",
            f"button[id='{source_id}']",
        ]
        sufijo = source_id.split(":")[-1] if ":" in source_id else source_id
        candidatos.extend(
            [
                f"a[onclick*='{sufijo}']",
                f"button[onclick*='{sufijo}']",
            ]
        )
        for selector in candidatos:
            try:
                loc = page.locator(selector)
            except Exception:
                continue
            if loc.count():
                disparador = loc.first
                break
        if disparador is None:
            return None, _obtener_view_state(page)

        try:
            disparador.scroll_into_view_if_needed(timeout=500)
        except Exception:
            pass
        try:
            disparador.click(timeout=1000)
        except Exception as err:
            raise RuntimeError(f"No se pudo abrir el detalle emitido ({source_id}).") from err

        dialog_locator = page.locator("div.ui-dialog:has(span.ui-dialog-title:has-text('XML'))")
        if not dialog_locator.count():
            dialog_locator = page.locator("div.ui-dialog:has(span.ui-dialog-title)")
        try:
            dialog_locator.evaluate_all(
                "(els) => els.forEach(el => { el.classList.remove('ui-overlay-hidden'); el.style.visibility = 'visible'; })"
            )
        except Exception:
            pass
        try:
            dialog_locator.first.wait_for(state="visible", timeout=DOWNLOAD_TIMEOUT)
        except PlaywrightTimeoutError as err:
            print(f"[WARN] No aparecio el dialogo XML para {source_id}: {err}")
            raise RuntimeError("No aparecio el dialogo de descarga de XML.") from err
        dialog_actual = dialog_locator.first

        form_payload = page.evaluate(
            """() => {
                const form = document.querySelector('form#j_idt913');
                if (!form) { return null; }
                const data = new FormData(form);
                const payload = {};
                for (const [key, value] of data.entries()) {
                    payload[key] = value;
                }
                return payload;
            }"""
        )
        if not form_payload:
            print(f"[WARN] No se encontro el formulario j_idt913 en el dialogo para {source_id}.")
            raise RuntimeError("No se pudo obtener el formulario de descarga de XML.")

        encoded_payload = urlencode(form_payload, doseq=True)
        request_headers = dict(base_headers or {})
        request_headers.update(
            {
                "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
                "Faces-Request": "partial/ajax",
                "X-Requested-With": "XMLHttpRequest",
                "Accept": "application/xml, text/xml, */*; q=0.01",
                "Origin": "https://srienlinea.sri.gob.ec",
                "Referer": page.url,
            }
        )

        try:
            respuesta = request_context.post(
            RECUPERAR_COMPROBANTES_URL,
            data=encoded_payload,
            headers=request_headers,
        )
        except Exception as err:
            print(f"[WARN] Fallo la solicitud POST de XML para {source_id}: {err}")
            raise
        if respuesta.status != 200:
            print(f"[WARN] HTTP {respuesta.status} al solicitar XML de emitidos para {source_id}.")
            raise RuntimeError(f"Error HTTP {respuesta.status} al descargar XML de emitidos.")
        try:
            cuerpo_bytes = respuesta.body()
        except Exception as err:
            print(f"[WARN] No se pudo leer cuerpo de respuesta XML para {source_id}: {err}")
            raise RuntimeError("No se pudo leer la respuesta del XML de emitidos.") from err
        if not cuerpo_bytes:
            print(f"[WARN] Respuesta vacia al descargar XML para {source_id}.")
            raise RuntimeError("La respuesta del XML de emitidos llego vacia.")
        try:
            contenido = cuerpo_bytes.decode("utf-8")
        except UnicodeDecodeError:
            contenido = cuerpo_bytes.decode("utf-8", "ignore")
        contenido = contenido.strip()
        if not contenido:
            raise ValueError("El XML descargado esta vacio.")


        try:
            meta = _parse_emitido_comprobante(contenido, None)
        except Exception as err:
            print(f"[WARN] No se pudo interpretar el XML de emitidos descargado: {err}")
            meta = {"xml_contenido": contenido}
        if not isinstance(meta, dict):
            meta = {"xml_contenido": contenido}
        meta.setdefault("xml_contenido", contenido)

        clave = meta.get("clave_acceso")
        if clave and clave in claves_guardadas:
            resultado_path = None
        else:
            contenido_xml = meta.get("xml_contenido") or contenido
            if not contenido_xml.lstrip().startswith("<?xml"):
                contenido_xml = '<?xml version="1.0" encoding="UTF-8"?>\n' + contenido_xml.lstrip()

            nombre_archivo = _construir_nombre_xml_emitido(meta, fallback_nombre)
            destino_base = xml_dir / nombre_archivo
            destino_final = _resolver_destino_unico(destino_base, ".xml")
            destino_final.write_text(contenido_xml, encoding="utf-8")
            resultado_path = destino_final

            if clave:
                claves_guardadas.add(clave)

    finally:
        if dialog_actual:
            try:
                cierre = dialog_actual.locator("a.ui-dialog-titlebar-close")
                if cierre.count():
                    cierre.first.click()
                else:
                    cierre_btn = dialog_actual.locator("button[title*='Cerrar' i], button:has-text('Cerrar')")
                    if cierre_btn.count():
                        cierre_btn.first.click()
            except Exception:
                try:
                    page.keyboard.press("Escape")
                except Exception:
                    pass

    nuevo_view_state = _obtener_view_state(page)
    return resultado_path, nuevo_view_state


SOAP_ENVELOPE_TEMPLATE = """<?xml version="1.0" encoding="UTF-8"?>
<soapenv:Envelope xmlns:soapenv="http://schemas.xmlsoap.org/soap/envelope/"
                  xmlns:ec="http://ec.gob.sri.ws.autorizacion">
  <soapenv:Header/>
  <soapenv:Body>
    <ec:autorizacionComprobante>
      <claveAccesoComprobante>{clave}</claveAccesoComprobante>
    </ec:autorizacionComprobante>
  </soapenv:Body>
</soapenv:Envelope>
"""


def _descargar_xml_emitido_por_clave(
    request_context,
    clave_acceso: str,
    xml_dir: Path,
    fallback_nombre: str,
    claves_guardadas: set[str],
) -> Optional[Path]:
    clave = (clave_acceso or "").strip()
    if not clave:
        raise ValueError("Clave de acceso vacia para la descarga SOAP.")

    envelope = SOAP_ENVELOPE_TEMPLATE.format(clave=clave)
    headers = {
        "Content-Type": "text/xml; charset=UTF-8",
        "SOAPAction": "",
        "Accept": "text/xml",
    }
    respuesta = request_context.post(
        AUTORIZACION_COMPROBANTES_SOAP_URL,
        data=envelope.encode("utf-8"),
        headers=headers,
    )
    if respuesta.status != 200:
        raise RuntimeError(f"Servicio SOAP respondio HTTP {respuesta.status}.")
    cuerpo = respuesta.text()
    if not cuerpo:
        raise RuntimeError("El servicio SOAP devolvio respuesta vacia.")

    match = re.search(r"(<autorizacion[\s\S]*?</autorizacion>)", cuerpo, flags=re.IGNORECASE)
    if not match:
        raise RuntimeError("El servicio SOAP no devolvio un bloque <autorizacion>.")
    autorizacion_xml = html.unescape(match.group(1))

    comprobante_xml, meta_aut = _extraer_comprobante_desde_autorizacion(autorizacion_xml)
    if not comprobante_xml or not comprobante_xml.strip():
        estado = (meta_aut or {}).get("estado", "desconocido") if meta_aut else "desconocido"
        raise RuntimeError(f"El servicio SOAP retorno estado '{estado}' sin comprobante.")

    meta = _parse_emitido_comprobante(comprobante_xml, meta_aut)
    meta.setdefault("xml_contenido", comprobante_xml)

    clave_meta = (meta.get("clave_acceso") or clave).strip()
    if clave_meta and clave_meta in claves_guardadas:
        return None

    contenido_xml = meta.get("xml_contenido") or comprobante_xml
    if not contenido_xml.lstrip().startswith("<?xml"):
        contenido_xml = '<?xml version="1.0" encoding="UTF-8"?>\n' + contenido_xml.lstrip()

    nombre_archivo = _construir_nombre_xml_emitido(meta, fallback_nombre)
    destino_base = xml_dir / nombre_archivo
    destino_final = _resolver_destino_unico(destino_base, ".xml")
    destino_final.write_text(contenido_xml, encoding="utf-8")

    if clave_meta:
        claves_guardadas.add(clave_meta)
    return destino_final


def _parse_decimal(texto: str) -> Optional[float]:
    bruto = (texto or "").strip().replace("\xa0", "").replace(" ", "")
    if not bruto:
        return None
    candidatos = [
        bruto,
        bruto.replace(".", "").replace(",", "."),
        bruto.replace(",", ""),
    ]
    for candidato in candidatos:
        try:
            return float(candidato)
        except ValueError:
            continue
    return None


def _parse_datetime_local(texto: str) -> Optional[datetime]:
    bruto = (texto or "").strip()
    if not bruto:
        return None
    bruto = re.sub(r"\s+", " ", bruto)
    formatos = ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y")
    for fmt in formatos:
        try:
            return datetime.strptime(bruto, fmt)
        except ValueError:
            continue
    return None


def _click_texto(page, texto: str) -> bool:
    for metodo in [
        lambda: page.get_by_role("button", name=texto, exact=False),
        lambda: page.get_by_text(texto, exact=False),
        lambda: page.locator(f"//button[contains(., '{texto}') or @title[contains(.,'{texto}')]]")
    ]:
        try:
            metodo().first.click(timeout=1000)
            return True
        except Exception:
            continue
    return False

def _resolver_destino_unico(base_path: Path, extension: str) -> Path:
    """
    Ajusta el nombre final asegurando que exista la extension indicada
    y evitando colisiones con archivos existentes.
    """
    extension = extension if extension.startswith(".") else f".{extension}"
    destino = base_path.with_suffix(extension)
    contador = 1
    while destino.exists():
        destino = base_path.with_name(f"{base_path.stem}_{contador}{extension}")
        contador += 1
    return destino

def _es_respuesta_pdf(response) -> bool:
    try:
        headers = response.headers or {}
    except Exception:
        headers = {}
    content_type = headers.get("content-type", "").lower()
    content_disposition = headers.get("content-disposition", "").lower()
    if "application/pdf" in content_type:
        return True
    if "pdf" in content_disposition:
        return True
    return False

def _guardar_pdf_desde_enlace(page, link_locator, base_destino: Path) -> Optional[Path]:
    """
    Intenta descargar un PDF a partir de un locator. Primero espera un evento de download.
    Si no llega, intenta capturar la respuesta HTTP del PDF y guardarla manualmente.
    Devuelve la ruta final del archivo o None si salio mal.
    """
    def _click_objetivo():
        try:
            link_locator.scroll_into_view_if_needed(timeout=500)
        except Exception:
            pass
        link_locator.click(no_wait_after=True)

    errores = []

    # Primer intento: evento de descarga tradicional
    try:
        with page.expect_download(timeout=DOWNLOAD_TIMEOUT) as descarga_info:
            _click_objetivo()
        descarga = descarga_info.value
        sugerido = Path(descarga.suggested_filename or base_destino.name)
        extension = sugerido.suffix or ".pdf"
        destino_final = _resolver_destino_unico(base_destino, extension)
        descarga.save_as(str(destino_final))
        return destino_final
    except PlaywrightTimeoutError as err:
        errores.append(f"Falla al esperar descarga directa de PDF (timeout): {err}")
    except Exception as err:
        errores.append(f"Falla al esperar descarga directa de PDF: {err}")

    # Segundo intento: capturar la respuesta HTTP del PDF
    try:
        with page.expect_response(
            lambda response: _es_respuesta_pdf(response), timeout=DOWNLOAD_TIMEOUT
        ) as respuesta_info:
            _click_objetivo()
        respuesta = respuesta_info.value
        if not _es_respuesta_pdf(respuesta):
            return None
        try:
            cuerpo = respuesta.body()
        except Exception:
            return None
        if not cuerpo:
            return None
        headers = respuesta.headers or {}
        sugerido = headers.get("content-disposition", "")
        extension = ".pdf"
        if "filename=" in sugerido:
            nombre = sugerido.split("filename=")[-1].strip().strip('"').strip("'")
            if "." in nombre:
                extension = Path(nombre).suffix or extension
        destino_final = _resolver_destino_unico(base_destino, extension)
        Path(destino_final).write_bytes(cuerpo)
        return destino_final
    except PlaywrightTimeoutError as err:
        errores.append(f"Falla al capturar respuesta PDF (timeout): {err}")
    except Exception as err:
        errores.append(f"Falla al capturar respuesta PDF: {err}")

    for mensaje in errores:
        print(f"[WARN] {mensaje}")
    return None

def _es_respuesta_xml(response) -> bool:
    try:
        headers = response.headers or {}
    except Exception:
        headers = {}
    content_type = headers.get("content-type", "").lower()
    content_disposition = headers.get("content-disposition", "").lower()
    if "xml" in content_type:
        return True
    if "xml" in content_disposition:
        return True
    return False

def _guardar_xml_desde_enlace(page, link_locator, base_destino: Path) -> Optional[Path]:
    """
    Descarga un XML ya sea por evento de download o capturando la respuesta HTTP asociada.
    """
    def _click_objetivo():
        try:
            link_locator.scroll_into_view_if_needed(timeout=500)
        except Exception:
            pass
        link_locator.click(no_wait_after=True)

    errores = []

    try:
        with page.expect_download(timeout=DOWNLOAD_TIMEOUT) as descarga_info:
            _click_objetivo()
        descarga = descarga_info.value
        sugerido = Path(descarga.suggested_filename or f"{base_destino.name}.xml")
        extension = sugerido.suffix or ".xml"
        destino_final = _resolver_destino_unico(base_destino, extension)
        descarga.save_as(str(destino_final))
        return destino_final
    except PlaywrightTimeoutError as err:
        errores.append(f"Falla al esperar descarga directa de XML (timeout): {err}")
    except Exception as err:
        errores.append(f"Falla al esperar descarga directa de XML: {err}")

    try:
        with page.expect_response(
            lambda response: _es_respuesta_xml(response), timeout=DOWNLOAD_TIMEOUT
        ) as respuesta_info:
            _click_objetivo()
        respuesta = respuesta_info.value
        if not _es_respuesta_xml(respuesta):
            return None
        try:
            cuerpo = respuesta.body()
        except Exception:
            cuerpo = b""
        if not cuerpo:
            return None
        headers = respuesta.headers or {}
        sugerido = headers.get("content-disposition", "")
        extension = ".xml"
        if "filename=" in sugerido:
            nombre = sugerido.split("filename=")[-1].strip().strip('"').strip("'")
            if "." in nombre:
                extension = Path(nombre).suffix or extension
        destino_final = _resolver_destino_unico(base_destino, extension)
        destino_final.write_bytes(cuerpo)
        return destino_final
    except PlaywrightTimeoutError as err:
        errores.append(f"Falla al capturar respuesta XML (timeout): {err}")
    except Exception as err:
        errores.append(f"Falla al capturar respuesta XML: {err}")

    for mensaje in errores:
        print(f"[WARN] {mensaje}")
    return None

def _seleccionar(page, etiqueta: str, valor_visible: str):
    try:
        sel = page.get_by_label(etiqueta, exact=False).locator("select")
        sel.select_option(label=valor_visible)
        return True
    except Exception:
        pass
    try:
        page.locator(f"text={etiqueta}").locator("xpath=..").locator("select").select_option(label=valor_visible)
        return True
    except Exception:
        pass
    return False


def _normalizar_token(texto: str) -> str:
    base = unicodedata.normalize("NFKD", (texto or "")).encode("ascii", "ignore").decode("ascii").lower()
    return re.sub(r"[^a-z0-9]+", "", base)


def _rellenar_input_por_label(page, etiquetas, valor: str, selectores_extra=None) -> bool:
    if valor is None:
        return False
    texto = (valor or "").strip()
    if not texto:
        return False
    if isinstance(etiquetas, str):
        etiquetas = [etiquetas]
    if selectores_extra is None:
        selectores_extra = []
    try:
        selectores_extra = list(selectores_extra)
    except TypeError:
        selectores_extra = [selectores_extra]

    for etiqueta in etiquetas:
        try:
            loc = page.get_by_label(etiqueta, exact=False)
            if loc.count():
                loc.first.fill("")
                loc.first.fill(texto)
                return True
        except Exception:
            continue

    tokens = [_normalizar_token(e) for e in etiquetas if e]
    tokens = [t for t in tokens if t]
    candidatos = list(selectores_extra)
    for token in tokens:
        candidatos.extend([
            f"input[name*='{token}']",
            f"input[id*='{token}']",
            f"input[data-testid*='{token}']",
        ])

    for selector in candidatos:
        try:
            loc = page.locator(selector)
            if loc.count():
                loc.first.fill("")
                loc.first.fill(texto)
                return True
        except Exception:
            continue

    return False


def _seleccionar_en_select(page, selector: str, *valores) -> bool:
    try:
        locator = page.locator(selector)
    except Exception:
        return False
    if not locator or not locator.count():
        return False
    if _seleccionar_por_label(locator, *valores):
        return True
    for valor in valores:
        if valor is None:
            continue
        try:
            locator.select_option(value=valor)
            return True
        except Exception:
            continue
    try:
        page.evaluate(
            """({selector, valores}) => {
                const el = document.querySelector(selector);
                if (!el) { return false; }
                const candidatos = valores
                    .map(v => (v == null ? "" : String(v)))
                    .filter(v => v.trim().length > 0);
                if (!candidatos.length) { return false; }
                const comparar = (texto) => (texto || "").trim().toLowerCase();
                for (const opcion of Array.from(el.options)) {
                    const label = comparar(opcion.label);
                    const value = comparar(opcion.value);
                    for (const objetivo of candidatos) {
                        const norm = comparar(objetivo);
                        if (norm && (norm === label || norm === value)) {
                            if (el.value !== opcion.value) {
                                el.value = opcion.value;
                                el.dispatchEvent(new Event('change', { bubbles: true }));
                            }
                            return true;
                        }
                    }
                }
                el.value = candidatos[0];
                el.dispatchEvent(new Event('change', { bubbles: true }));
                return true;
            }""",
            {"selector": selector, "valores": list(valores)},
        )
        return True
    except Exception:
        return False

def _espera_captcha(page, timeout: int = 1000):
    try:
        loc = page.locator("img[alt='captcha']")
        if loc.is_visible(timeout=1000):
            page.wait_for_selector("img[alt='captcha']", state="detached", timeout=timeout)
    except Exception:
        pass


def _captcha_visible(page, timeout: int = 0) -> bool:
    try:
        loc = page.locator("img[alt='captcha']")
        if timeout:
            return loc.is_visible(timeout=timeout)
        return loc.is_visible()
    except Exception:
        return False


def _recaptcha_presente(page) -> bool:
    try:
        if page.locator("iframe[src*='recaptcha']").count():
            return True
    except Exception:
        pass
    try:
        if page.locator("[data-sitekey]").count():
            return True
    except Exception:
        pass
    return False


def _esperar_recaptcha_resuelto(page, timeout: int = 1000):
    try:
        page.wait_for_function(
            """() => {
                const challenge = document.querySelector("iframe[src*='recaptcha/api2/bframe']");
                if (!challenge) { return true; }
                const rect = challenge.getBoundingClientRect();
                const oculto = rect.width === 0 || rect.height === 0 || window.getComputedStyle(challenge).visibility === 'hidden';
                return oculto;
            }""",
            timeout=timeout,
        )
    except Exception:
        pass


def _recaptcha_challenge_activo(page) -> bool:
    try:
        frame = page.locator("iframe[src*='recaptcha/api2/bframe']")
        if frame.count():
            try:
                return frame.first.is_visible()
            except Exception:
                pass
    except Exception:
        pass
    return False


CAPTCHA_INPUT_SELECTORS = [
    "input[name*='captcha' i]",
    "input[id*='captcha' i]",
    "input[name='captcha']",
    "input[id='captcha']",
    "input#captchaIngresar",
    "input#captchaTxt",
]
CAPTCHA_INPUT_QUERY = ",".join(CAPTCHA_INPUT_SELECTORS)


def _esperar_captcha_manual_input(page, timeout: int = 120000) -> bool:
    """
    Espera hasta que el usuario ingrese un valor de captcha de forma manual.
    Se considera resuelto cuando cualquiera de los inputs registrados tiene texto.
    """
    try:
        page.wait_for_function(
            """(selectorCadena) => {
                const inputs = document.querySelectorAll(selectorCadena);
                for (const input of inputs) {
                    const valor = (input.value || "").trim();
                    if (valor.length >= 4) {
                        return true;
                    }
                }
                return false;
            }""",
            arg=CAPTCHA_INPUT_QUERY,
            timeout=timeout,
        )
        return True
    except Exception:
        return False


def _localizar_input_captcha(page):
    for selector in CAPTCHA_INPUT_SELECTORS:
        try:
            locator = page.locator(selector)
            if locator.count():
                return locator.first
        except Exception:
            continue
    return None


def _resolver_captcha(page, contexto: str) -> bool:
    recaptcha_detectado = False
    try:
        recaptcha_detectado = _recaptcha_presente(page)
    except Exception:
        recaptcha_detectado = False

    if recaptcha_detectado:
        _notificar_usuario_captcha("reCAPTCHA", contexto)
        _esperar_recaptcha_resuelto(page, timeout=120000)
        return True

    try:
        if not _captcha_visible(page, timeout=1000):
            return False
    except Exception:
        return False

    if not captcha_solver_enabled():
        _notificar_usuario_captcha("captcha de imagen", contexto)
        _espera_captcha(page)
        return False

    for intento in range(1, CAPTCHA_MAX_ATTEMPTS + 1):
        try:
            if not _captcha_visible(page, timeout=1000):
                return False
        except Exception:
            return False

        input_captcha = _localizar_input_captcha(page)
        if input_captcha is None:
            print(f"[WARN] Campo de texto para captcha no encontrado ({contexto}); esperando resolucion manual.")
            _notificar_usuario_captcha("captcha de imagen", contexto)
            _espera_captcha(page)
            return False

        try:
            imagen = page.locator("img[alt='captcha']").screenshot(type="png")
        except Exception as err:
            print(f"[WARN] No se pudo capturar la imagen del captcha (intento {intento}/{CAPTCHA_MAX_ATTEMPTS}): {err}")
            break

        try:
            codigo = solve_captcha_image(imagen)
        except CaptchaSolverError as err:
            print(f"[WARN] Fallo al resolver captcha con 2Captcha (intento {intento}/{CAPTCHA_MAX_ATTEMPTS}): {err}")
            continue

        try:
            input_captcha.fill("")
            input_captcha.fill(codigo)
            return True
        except Exception as err:
            print(f"[WARN] No se pudo escribir el captcha resuelto (intento {intento}/{CAPTCHA_MAX_ATTEMPTS}): {err}")

    print("[WARN] Se agotaron los intentos automaticos de captcha; esperando resolucion manual.")
    _notificar_usuario_captcha("captcha de imagen", contexto)
    _espera_captcha(page)
    return False

def _resolver_autenticacion_persistente(page) -> bool:
    """Gestiona la pantalla opcional de 'autenticacion persistente' mostrada por Keycloak."""
    try:
        page.wait_for_load_state("domcontentloaded", timeout=1000)
    except Exception:
        pass

    objetivos = [page]
    try:
        objetivos.extend([frame for frame in page.frames if frame not in objetivos])
    except Exception:
        pass

    palabras_clave = [
        "autenticacion persistente",
        "autenticacion persistente",
        "mantener la sesion",
        "mantener la sesion",
        "permanecer conectado",
        "persistente",
        "recordar este dispositivo",
    ]

    selectores_prioridad = [
        "button#cancel",
        "input#cancel",
        "button[name='cancel']",
        "input[name='cancel']",
        "button[value='cancel']",
        "input[value='cancel']",
        "button[data-kc-button='cancel']",
        "button#kc-cancel",
        "input#kc-cancel",
        "a#cancel",
        "a[name='cancel']",
        "button:has-text('No')",
        "button:has-text('NO')",
        "button:has-text('Cancelar')",
        "button:has-text('Cancelar autenticacion')",
        "input[value='No']",
        "input[value='NO']",
        "input[value='Cancelar']",
        "button:has-text('Continuar sin recordar')",
        "button:has-text('No mantener')",
    ]

    secundarios = [
        "button:has-text('Si')",
        "button:has-text('Si')",
        "input[value='Si']",
        "button[type='submit']",
        "input[type='submit']",
    ]

    def _clic_en_objetivo(objetivo, selector):
        try:
            loc = objetivo.locator(selector)
            if not loc.count():
                return False
            loc = loc.first
            loc.wait_for(state="visible", timeout=1000)
            loc.click()
            try:
                page.wait_for_load_state("networkidle", timeout=1000)
            except Exception:
                pass
            time.sleep(0.3)
            return True
        except Exception:
            return False

    for objetivo in objetivos:
        try:
            cuerpo = objetivo.locator("body")
            texto_cuerpo = (
                cuerpo.inner_text(timeout=1000).lower()
                if cuerpo.count()
                else ""
            )
        except Exception:
            texto_cuerpo = ""

        if texto_cuerpo and any(palabra in texto_cuerpo for palabra in palabras_clave):
            for selector in selectores_prioridad + secundarios:
                if _clic_en_objetivo(objetivo, selector):
                    return True

        # Si no hubo coincidencia de texto, igual intentar con selectores
        for selector in selectores_prioridad:
            if _clic_en_objetivo(objetivo, selector):
                return True

    # Como ultimo recurso, intentar con coincidencia textual global.
    for etiqueta in ["No", "No, gracias", "No mantener", "Cancelar", "Continuar"]:
        if _click_texto(page, etiqueta):
            try:
                page.wait_for_load_state("networkidle", timeout=1000)
            except Exception:
                pass
            time.sleep(0.3)
            return True

    try:
        Path("debug_login_persistente.html").write_text(
            page.content(),
            encoding="utf-8",
            errors="ignore",
        )
    except Exception:
        pass
    return False

def _cerrar_modal_encuesta(page) -> bool:
    """Intenta cerrar la alerta de encuesta mostrada tras el inicio de sesion."""
    try:
        page.wait_for_timeout(200)
    except Exception:
        pass

    selectores_prioridad = [
        "[role='dialog'] button[aria-label*='Cerrar']",
        "[role='dialog'] button[title*='Cerrar']",
        "[role='dialog'] button[aria-label*='Close']",
        "[role='dialog'] button[title*='Close']",
        "[role='dialog'] button.close",
        "[role='dialog'] button.mat-icon-button",
        "[role='dialog'] button:has-text('×')",
        "[role='dialog'] button:has-text('X')",
        "[role='dialog'] button:has-text('?')",
        "button[aria-label='Cerrar']",
        "button[aria-label='Close']",
        "button[title='Cerrar']",
        "button[title='Close']",
        "button[aria-label*='Cerrar encuesta']",
    ]

    for selector in selectores_prioridad:
        try:
            boton = page.locator(selector)
        except Exception:
            continue
        if not boton.count():
            continue
        boton = boton.first
        try:
            boton.wait_for(state="visible", timeout=500)
        except Exception:
            pass
        try:
            if not boton.is_enabled():
                continue
        except Exception:
            pass
        try:
            boton.click()
            try:
                page.wait_for_timeout(300)
            except Exception:
                pass
            return True
        except Exception:
            continue

    try:
        cerrado = page.evaluate(
            """
            () => {
                const normalize = (texto) => {
                    if (!texto) {
                        return "";
                    }
                    if (typeof texto.normalize === "function") {
                        return texto.normalize("NFD").replace(/[\\u0300-\\u036f]/g, "").toLowerCase();
                    }
                    return String(texto).toLowerCase();
                };
                const keywords = [
                    "tu opinion nos permite mejorar",
                    "encuesta de satisfaccion de servicios",
                    "quiero responder"
                ];
                const candidatos = Array.from(document.querySelectorAll("[role='dialog'], .mat-dialog-container, .modal, .cdk-overlay-pane"))
                    .map(el => el.closest("[role='dialog']") || el)
                    .filter(Boolean);
                const dialog = candidatos.find(el => {
                    const texto = normalize(el.innerText || el.textContent || "");
                    return keywords.some(keyword => texto.includes(keyword));
                });
                if (!dialog) {
                    return false;
                }
                const posibles = Array.from(dialog.querySelectorAll("button, a, span, mat-icon, i, div"));
                for (const elemento of posibles) {
                    const raw = (elemento.textContent || elemento.innerText || "").trim();
                    const texto = normalize(raw);
                    const aria = normalize(elemento.getAttribute && elemento.getAttribute("aria-label"));
                    const title = normalize(elemento.getAttribute && elemento.getAttribute("title"));
                    const clases = normalize(elemento.className || "");
                    if (["x", "×", "?"].includes(raw) ||
                        texto === "cerrar" ||
                        texto === "close" ||
                        aria.includes("cerrar") ||
                        title.includes("cerrar") ||
                        clases.includes("close") ||
                        clases.includes("cerrar")) {
                        if (typeof elemento.click === "function") {
                            elemento.click();
                            return true;
                        }
                    }
                }
                return false;
            }
            """
        )
        if cerrado:
            try:
                page.wait_for_timeout(300)
            except Exception:
                pass
            return True
    except Exception:
        pass

    for texto in ["×", "X", "Cerrar", "CERRAR", "Close", "CLOSE"]:
        if _click_texto(page, texto):
            try:
                page.wait_for_timeout(300)
            except Exception:
                pass
            return True
    return False

def _abrir_modulo_consultas(page, origen: str):
    """Cierra popups, abre el panel izquierdo y navega al formulario correspondiente."""
    _cerrar_modal_encuesta(page)
    destino_url = RECUPERAR_COMPROBANTES_URL if origen == "Emitidos" else RECIBIDOS_DIRECT_URL

    def _click_locator(selector: str, descripcion: str, timeout: int = 1200):
        try:
            locator = page.locator(selector)
            if not locator.count():
                return False
            locator.first.wait_for(state="visible", timeout=timeout)
            locator.first.click(timeout=timeout)
            return True
        except Exception as err:
            print(f"[WARN] No se pudo interactuar con {descripcion}: {err}")
            return False

    def _wait_overlays():
        for overlay in OVERLAY_SELECTORS:
            try:
                page.wait_for_selector(overlay, state="hidden", timeout=3000)
            except Exception:
                pass

    def _goto_form():
        ultimo_error = None
        for intento in range(3):
            try:
                page.goto(destino_url, wait_until="domcontentloaded", timeout=8000)
                page.wait_for_load_state("domcontentloaded", timeout=2000)
                page.wait_for_load_state("networkidle", timeout=2000)
                return
            except Exception as err:
                ultimo_error = err
                print(f"[WARN] Reintentando acceso directo al formulario ({intento + 1}/3): {err}")
        raise RuntimeError(f"No se pudo abrir el formulario de {origen.lower()}: {ultimo_error}")

    def _ensure_menu_visible():
        try:
            if page.locator(FACTURACION_MENU_SELECTOR).first.is_visible(timeout=500):
                return True
        except Exception:
            pass
        return _click_locator(MENU_TOGGLE_SELECTOR, "el botón de menú", timeout=1500)

    def _expand_panel(selector: str, descripcion: str):
        header = page.locator(selector)
        if not header.count():
            print(f"[WARN] No se encontró {descripcion}.")
            return False
        try:
            expanded = (header.first.get_attribute("aria-expanded") or "").lower()
        except Exception:
            expanded = ""
        if expanded == "true":
            return True
        try:
            header.first.click(timeout=1500)
            page.wait_for_timeout(200)
            return True
        except Exception as err:
            print(f"[WARN] No se pudo expandir {descripcion}: {err}")
            return False

    _wait_overlays()
    current_url = ""
    try:
        current_url = page.url or ""
    except Exception:
        current_url = ""
    if "consultas/menu.jsf" in current_url:
        _goto_form()
        return page

    _ensure_menu_visible()
    _expand_panel(FACTURACION_MENU_SELECTOR, "el panel de Facturación Electrónica")
    _expand_panel(MODULO_PRODUCCION_SELECTOR, "el panel Producción")

    consultas_locator = page.locator(CONSULTAS_SELECTOR)
    if consultas_locator.count():
        try:
            consultas_locator.first.click(timeout=1500)
        except Exception as err:
            print(f"[WARN] No se pudo hacer clic en 'Consultas': {err}")
    else:
        print("[WARN] No se encontró el botón de Consultas; intentando acceso directo.")

    _goto_form()

    return page


def _esperar_ajax(page, timeout: int = 1000):
    """Espera a que la cola AJAX de PrimeFaces quede vacia para evitar sobrescrituras."""
    try:
        page.wait_for_function(
            "() => window.PrimeFaces && PrimeFaces.ajax && PrimeFaces.ajax.Queue && PrimeFaces.ajax.Queue.isEmpty()",
            timeout=timeout,
        )
    except PlaywrightTimeoutError:
        try:
            page.wait_for_timeout(300)
        except Exception:
            pass


def _cerrar_sesion(pagina) -> bool:
    """Intenta cerrar la sesion activa del SRI en la pagina dada."""
    if not pagina:
        return False

    selectores = [
        "a[href*='salir.jspa']",
        "a[title*='Cerrar sesion']",
        "a[tooltip*='Cerrar sesion']",
        "a:has-text('Cerrar sesion')",
        "button:has-text('Cerrar sesion')",
    ]
    for selector in selectores:
        try:
            objetivo = pagina.locator(selector)
            if not objetivo.count():
                continue
            try:
                objetivo.first.scroll_into_view_if_needed(timeout=1000)
            except Exception:
                pass
            with pagina.expect_navigation(wait_until="load", timeout=1000):
                objetivo.first.click()
            try:
                pagina.wait_for_load_state("networkidle", timeout=1000)
            except Exception:
                pass
            return True
        except PlaywrightTimeoutError:
            continue
        except Exception:
            continue

    salir_urls = [
        "https://srienlinea.sri.gob.ec/tuportal-internet/salir.jspa",
        "/tuportal-internet/salir.jspa",
    ]
    for url in salir_urls:
        try:
            pagina.goto(url, wait_until="load", timeout=1000)
            try:
                pagina.wait_for_load_state("networkidle", timeout=1000)
            except Exception:
                pass
            return True
        except Exception:
            continue

    try:
        pagina.evaluate(
            """() => {
                if (window.location && typeof window.location.replace === "function") {
                    window.location.replace("https://srienlinea.sri.gob.ec/tuportal-internet/salir.jspa");
                    return true;
                }
                return false;
            }"""
        )
    except Exception:
        return False

    try:
        pagina.wait_for_load_state("load", timeout=1000)
    except Exception:
        pass
    return True

def _login(
    context,
    page,
    ruc,
    clave,
    cookies_path: Path,
    destino_url: str,
    ci_adicional: Optional[str] = None,
):
    """Realiza la autenticacion y gestiona la reutilizacion de cookies almacenadas."""

    def _navegar():
        try:
            page.goto(destino_url, timeout=1000)
        except Exception:
            pass
        try:
            page.wait_for_load_state("domcontentloaded", timeout=1000)
        except Exception:
            pass
        _asegurar_portal_disponible(page)

    if cookies_path.exists():
        try:
            context.add_cookies(json.loads(cookies_path.read_text()))
            _navegar()
            if "auth/realms" not in page.url:
                return
        except Exception:
            pass

    _navegar()

    if "auth/realms" in page.url:
        page.wait_for_selector(
            "input[name='usuario']:visible, input[name='username']:visible",
            timeout=1000,
        )

        usuario_selectores = [
            "input[name='usuario']:visible",
            "input[name='username']:visible",
            "input#usuario:visible",
            "input#username:visible",
        ]

        captcha_retry = 0
        while True:
            for selector in usuario_selectores:
                locator = page.locator(selector)
                if locator.count() and locator.first.is_enabled():
                    try:
                        locator.first.fill(ruc)
                        break
                    except Exception:
                        continue
            else:
                raise RuntimeError("No se encontro el campo de usuario en el formulario de login del SRI.")

            ci_locator = page.locator("input[name='ciAdicional']:visible")
            if ci_locator.count() and ci_locator.first.is_enabled():
                valor_ci = (ci_adicional or "").strip()
                try:
                    ci_locator.first.fill(valor_ci)
                except Exception:
                    pass

            pass_locator = page.locator("input[name='password']:visible, input[type='password']:visible")
            if pass_locator.count():
                pass_locator.first.fill(clave)
            else:
                raise RuntimeError("No se encontro el campo de contrasena en el formulario de login del SRI.")

            captcha_resuelto = _resolver_captcha(page, "login")

            if PAUSE_AT_LOGIN:
                print(PAUSE_PROMPT, flush=True)
                try:
                    input(">>> Presiona Enter para continuar con el clic en 'Ingresar'...")
                except EOFError:
                    print("Entrada estandar no disponible; continuando automaticamente en 5 segundos.", flush=True)
                    time.sleep(0.2)

            if not _click_texto(page, "Ingresar"):
                try:
                    page.locator("input[name='login'], button[type='submit']").first.click()
                except Exception:
                    keyboard = getattr(page, "keyboard", None)
                    if keyboard:
                        try:
                            keyboard.press("Enter")
                        except Exception:
                            pass
                    else:
                        print("[WARN] No se pudo accionar el boton 'Ingresar'; el objeto page no expone teclado.")

            try:
                page.wait_for_load_state("networkidle", timeout=1000)
            except Exception:
                pass

            autenticado = False
            for intento in range(2):
                try:
                    page.wait_for_url(lambda url: "auth/realms" not in url, timeout=1000)
                    autenticado = True
                    break
                except PlaywrightTimeoutError:
                    if _resolver_autenticacion_persistente(page):
                        continue
                    if intento == 1:
                        raise RuntimeError("No fue posible completar el login del SRI (pantalla de autenticacion persistente).")

            if autenticado and "auth/realms" not in page.url:
                break

            recaptcha_activo = False
            if captcha_resuelto:
                try:
                    recaptcha_activo = _recaptcha_challenge_activo(page)
                except Exception:
                    recaptcha_activo = False

            if captcha_resuelto and (_captcha_visible(page, timeout=1000) or recaptcha_activo):
                captcha_retry += 1
                if captcha_retry >= CAPTCHA_MAX_ATTEMPTS:
                    raise RuntimeError("No fue posible completar el login del SRI (captcha).")
                print(f"[INFO] Reintentando login por captcha adicional ({captcha_retry}/{CAPTCHA_MAX_ATTEMPTS}).")
                continue

            raise RuntimeError("No fue posible completar el login del SRI (credenciales o captcha).")
        if "auth/realms" in page.url:
            if not _resolver_autenticacion_persistente(page):
                raise RuntimeError("No fue posible completar el login del SRI (pantalla de autenticacion persistente).")
            try:
                page.wait_for_url(lambda url: "auth/realms" not in url, timeout=1000)
            except PlaywrightTimeoutError:
                raise RuntimeError("No fue posible completar el login del SRI (pantalla de autenticacion persistente).")

    if "auth/realms" not in page.url:
        _cerrar_modal_encuesta(page)
        try:
            cookies_path.write_text(json.dumps(context.cookies()))
        except Exception:
            pass

# ============================================================
# ?? DESCARGA DE COMPROBANTES RECIBIDOS (TXT + XML + PDF)
# ============================================================
def _seleccionar_por_label(locator, *labels) -> bool:
    for etiqueta in labels:
        if etiqueta is None:
            continue
        try:
            locator.select_option(label=etiqueta)
            return True
        except Exception:
            try:
                locator.select_option(value=etiqueta)
                return True
            except Exception:
                continue
    return False

def _flujo_recibidos(page, destino: Path, anio: int, mes: int, dia: int, tipo: str, formatos: list):
    try:
        page.goto(RECIBIDOS_DIRECT_URL, wait_until="domcontentloaded", timeout=5000)
        page.wait_for_load_state("domcontentloaded", timeout=1000)
    except Exception:
        pass
    page.wait_for_selector("select#frmPrincipal\\:ano", state="visible", timeout=10000)

    selector_ano = page.locator("select#frmPrincipal\\:ano")
    if not _seleccionar_por_label(selector_ano, str(anio)):
        raise RuntimeError("No fue posible seleccionar el año solicitado en el SRI.")

    selector_mes = page.locator("select#frmPrincipal\\:mes")
    mes_texto = _mes_a_texto(mes)
    if not _seleccionar_por_label(selector_mes, mes_texto, f"{mes:02d}", str(mes)):
        raise RuntimeError("No fue posible seleccionar el mes solicitado en el SRI.")

    selector_dia = page.locator("select#frmPrincipal\\:dia")
    dia_labels = ("Todos", "0") if dia in (None, 0) else (str(dia), f"{dia:02d}")
    if not _seleccionar_por_label(selector_dia, *dia_labels):
        objetivo = "Todos" if dia in (None, 0) else str(dia)
        raise RuntimeError(f"No fue posible seleccionar el dia '{objetivo}' en el SRI.")

    _esperar_ajax(page)

    def _valor_actual_dia():
        try:
            return page.evaluate(
                """({selector}) => {
                    const el = document.querySelector(selector);
                    if (!el) { return null; }
                    const selected = el.options[el.selectedIndex];
                    const texto = selected ? (selected.textContent || selected.label || "").trim() : "";
                    return {
                        valor: (el.value || "").trim(),
                        texto
                    };
                }""",
                arg={"selector": "select#frmPrincipal\\:dia"},
            )
        except Exception:
            return None

    def _normalizar_dia(token: str) -> str:
        token = (token or "").strip()
        if not token:
            return ""
        if token.lower() == "todos":
            return "todos"
        if token.isdigit():
            numero = int(token)
            return "0" if numero == 0 else str(numero)
        return token.lower()

    esperados_norm = {_normalizar_dia(opcion) for opcion in dia_labels}

    def _coincide(info) -> bool:
        if not info:
            return False
        candidatos = [
            _normalizar_dia(info.get("valor")),
            _normalizar_dia(info.get("texto")),
        ]
        return any(c in esperados_norm for c in candidatos if c)

    valor_actual = _valor_actual_dia()
    if not _coincide(valor_actual):
        try:
            page.evaluate(
                """({selector, opciones}) => {
                    const normalizar = (token) => {
                        if (!token) { return ""; }
                        const trimmed = token.trim();
                        if (trimmed.toLowerCase() === "todos") { return "todos"; }
                        if (/^\\d+$/.test(trimmed)) {
                            const numero = parseInt(trimmed, 10);
                            return numero === 0 ? "0" : String(numero);
                        }
                        return trimmed.toLowerCase();
                    };
                    const el = document.querySelector(selector);
                    if (!el) { return false; }
                    const objetivos = opciones.map(normalizar);
                    for (const option of Array.from(el.options)) {
                        const lbl = normalizar(option.label || option.textContent || "");
                        const val = normalizar(option.value || "");
                        if (objetivos.includes(lbl) || objetivos.includes(val)) {
                            if (el.value !== option.value) {
                                el.value = option.value;
                                el.dispatchEvent(new Event('change', { bubbles: true }));
                            }
                            return true;
                        }
                    }
                    return false;
                }""",
                arg={"selector": "select#frmPrincipal\\:dia", "opciones": list(dia_labels)},
            )
        except Exception:
            pass
        _esperar_ajax(page)
        valor_actual = _valor_actual_dia()

    if not _coincide(valor_actual):
        objetivo = "Todos" if dia in (None, 0) else str(dia)
        raise RuntimeError(f"No se logro confirmar el dia '{objetivo}' en el SRI.")

    selector_tipo = page.locator("select#frmPrincipal\\:cmbTipoComprobante")
    tipo_visible = TIPOS_MAP.get(tipo, tipo)
    _seleccionar_por_label(selector_tipo, tipo_visible)

    _orden_tipo, _label_tipo, tipo_prefijo = _prefijo_tipo(tipo_visible or tipo)
    tipo_slug = _slug_tipo(tipo_visible or tipo)
    anio_dir = f"{anio:04d}"
    mes_dir = _mes_a_texto(mes)
    try:
        dia_int = int(dia)
    except (TypeError, ValueError):
        dia_int = None
    dia_dir = "Todos" if dia_int in (None, 0) else f"{dia_int:02d}"

    carpeta_tipo = destino / anio_dir / mes_dir / dia_dir / tipo_prefijo
    carpeta_tipo.mkdir(parents=True, exist_ok=True)
    txt_dir = carpeta_tipo / "TXT"
    xml_dir = carpeta_tipo / "XML"
    pdf_dir = carpeta_tipo / "PDF"

    boton_consultar = page.locator("#btnRecaptcha")
    boton_consultar.first.wait_for(state="visible", timeout=1000)

    max_intentos = CAPTCHA_MAX_ATTEMPTS if captcha_solver_enabled() else 1
    for intento_captcha in range(max_intentos):
        solucion_automatica = _resolver_captcha(page, "recibidos_consulta")

        try:
            captcha_visible = _captcha_visible(page, timeout=500)
        except Exception:
            captcha_visible = False

        recaptcha_activo = False
        if not captcha_visible:
            try:
                recaptcha_activo = _recaptcha_challenge_activo(page)
            except Exception:
                recaptcha_activo = False

        requiere_intervencion_manual = (not solucion_automatica) and (captcha_visible or recaptcha_activo)

        if requiere_intervencion_manual:
            if captcha_visible:
                _notificar_usuario_captcha("captcha de imagen", "recibidos_consulta")
                _esperar_captcha_manual_input(page, timeout=120000)
            else:
                _notificar_usuario_captcha("reCAPTCHA", "recibidos_consulta")
                _esperar_recaptcha_resuelto(page, timeout=120000)

        boton_consultar.first.click()
        try:
            page.wait_for_load_state("networkidle", timeout=1000)
        except Exception:
            pass
        time.sleep(0.2)

        if requiere_intervencion_manual:
            try:
                page.wait_for_load_state("networkidle", timeout=2000)
            except Exception:
                pass
            break

        recaptcha_activo = False
        try:
            recaptcha_activo = _recaptcha_challenge_activo(page)
        except Exception:
            recaptcha_activo = False
        if not (_captcha_visible(page, timeout=1000) or recaptcha_activo):
            break
        if intento_captcha + 1 >= max_intentos:
            raise RuntimeError("No fue posible superar el captcha al consultar comprobantes recibidos.")
        print(f"[INFO] Reintentando consulta por captcha adicional ({intento_captcha + 1}/{max_intentos}).")

    tabla_datos = page.locator("#frmPrincipal\\:tablaCompRecibidos_data")
    mensaje_vacio = page.locator(".ui-datatable-empty-message")
    alerta_parametros = page.locator(".ui-messages-info, .ui-messages-warn, .ui-messages-error")

    def _resultado_sin_datos(texto: str = "") -> dict:
        texto = (texto or "No se encontraron comprobantes para los filtros seleccionados.").strip()
        return {
            "estado": "sin_resultados",
            "mensaje": texto,
            "n_xml": 0,
            "n_pdf": 0,
        }

    def _texto_alerta() -> str:
        if alerta_parametros.count():
            try:
                return alerta_parametros.first.inner_text().strip()
            except Exception:
                return ""
        if mensaje_vacio.count():
            try:
                return mensaje_vacio.first.inner_text().strip()
            except Exception:
                return ""
        return ""

    alerta_texto = _texto_alerta()
    if alerta_texto:
        return _resultado_sin_datos(alerta_texto)
    try:
        tabla_datos.wait_for(state="visible", timeout=180000)
    except PlaywrightTimeoutError:
        alerta_texto = _texto_alerta()
        if alerta_texto:
            return _resultado_sin_datos(alerta_texto)
        try:
            if _captcha_visible(page, timeout=500):
                raise RuntimeError(
                    "El captcha se mostro nuevamente luego de consultarlo. Resuelvelo y presiona 'Consultar'."
                )
        except Exception:
            pass
        raise
    except Exception:
        alerta_texto = _texto_alerta()
        if alerta_texto:
            return _resultado_sin_datos(alerta_texto)
        raise

    alerta_texto = _texto_alerta()
    if alerta_texto:
        return _resultado_sin_datos(alerta_texto)

    n_xml = 0
    n_pdf = 0
    txt_path = None

    if "XML" in formatos:
        link_reporte = page.locator("a#frmPrincipal\\:lnkTxtlistado")
        if link_reporte.count():
            try:
                with page.expect_download(timeout=DOWNLOAD_TIMEOUT) as descarga_info:
                    link_reporte.first.click(no_wait_after=True)
                descarga = descarga_info.value
                txt_dir.mkdir(parents=True, exist_ok=True)
                txt_nombre = descarga.suggested_filename or f"recibidos_{anio}_{mes:02d}.txt"
                txt_path = txt_dir / txt_nombre
                descarga.save_as(str(txt_path))
            except Exception as err:
                print(f"[WARN] No se pudo descargar el reporte TXT/XML: {err}")
        else:
            print("[WARN] No se encontro el enlace 'Descargar reporte' para XML.")

    descargar_xml = "XML" in formatos
    descargar_pdf = "PDF" in formatos

    if descargar_xml:
        xml_dir.mkdir(parents=True, exist_ok=True)
    if descargar_pdf:
        pdf_dir.mkdir(parents=True, exist_ok=True)

    if descargar_xml or descargar_pdf:
        pagina = 1
        while True:
            filas = tabla_datos.locator("tr")
            total_filas = filas.count()
            for idx in range(total_filas):
                fila = filas.nth(idx)
                celdas = fila.locator("td")
                if not celdas.count():
                    continue
                razon_texto = celdas.nth(1).inner_text().strip()
                bloques = [segmento.strip() for segmento in razon_texto.splitlines() if segmento.strip()]
                razon_social = bloques[-1] if bloques else f"documento_{pagina}_{idx+1}"
                nombre_base = _sanear_nombre_archivo(razon_social)

                xml_guardado = False
                if descargar_xml and not xml_guardado:
                    xml_link_directo = fila.locator("a[id$=':lnkXml']")
                    if not xml_link_directo.count():
                        xml_link_directo = fila.locator("a[title*='xml' i], button[title*='xml' i]")
                    if not xml_link_directo.count():
                        icono_xml = fila.locator("img[title*='xml' i], img[alt*='xml' i]")
                        if icono_xml.count():
                            contenedor = icono_xml.first.locator("xpath=ancestor::a[1] | xpath=ancestor::button[1]")
                            if not contenedor.count():
                                contenedor = icono_xml.first.locator("xpath=ancestor::span[1]")
                            if contenedor.count():
                                xml_link_directo = contenedor.first
                    if xml_link_directo and xml_link_directo.count():
                        destino_base_directo = xml_dir / nombre_base
                        destino_xml_directo = _resolver_destino_unico(destino_base_directo, ".xml")
                        resultado_directo = _guardar_xml_desde_enlace(page, xml_link_directo, destino_xml_directo)
                        if resultado_directo:
                            n_xml += 1
                            xml_guardado = True

                if descargar_xml and not xml_guardado:
                    xml_link = None
                    selectores_xml = [
                        "a[id$=':lnkXml']",
                        "a[id*='lnkXml']",
                        "a[title*='xml' i]",
                        "button[id*='lnkXml']",
                        "button[title*='xml' i]",
                    ]
                    for selector in selectores_xml:
                        posible = fila.locator(selector)
                        if posible.count():
                            xml_link = posible.first
                            break
                    if not xml_link:
                        icono_xml = fila.locator("img[title*='xml' i], img[alt*='xml' i]")
                        if icono_xml.count():
                            contenedor = icono_xml.first.locator("xpath=ancestor::a[1]")
                            if contenedor.count():
                                xml_link = contenedor.first
                    if xml_link:
                        destino_xml = xml_dir / f"{nombre_base}.xml"
                        sufijo_xml = 1
                        while destino_xml.exists():
                            destino_xml = xml_dir / f"{nombre_base}_{sufijo_xml}.xml"
                            sufijo_xml += 1
                        try:
                            with page.expect_download(timeout=DOWNLOAD_TIMEOUT) as descarga_info:
                                xml_link.click(no_wait_after=True)
                            descarga_xml = descarga_info.value
                            sugerido = descarga_xml.suggested_filename or destino_xml.name
                            extension = Path(sugerido).suffix or ".xml"
                            destino_final = destino_xml.with_suffix(extension)
                            sufijo_xml = 1
                            while destino_final.exists():
                                destino_final = destino_xml.with_name(f"{destino_xml.stem}_{sufijo_xml}{extension}")
                                sufijo_xml += 1
                            descarga_xml.save_as(str(destino_final))
                            n_xml += 1
                            xml_guardado = True
                        except Exception as err:
                            print(f"[WARN] No se pudo descargar XML para '{razon_social}': {err}")

                if descargar_pdf:
                    link_pdf = fila.locator("a[id$=':lnkPdf']")
                    if not link_pdf.count():
                        link_pdf = fila.locator("a[title*='pdf' i], button[title*='pdf' i]")
                    if link_pdf.count():
                        destino_pdf = pdf_dir / f"{nombre_base}.pdf"
                        resultado_pdf = _guardar_pdf_desde_enlace(page, link_pdf.first, destino_pdf)
                        if resultado_pdf:
                            n_pdf += 1
                        else:
                            print(f"[WARN] No se pudo descargar PDF para '{razon_social}': no se obtuvo archivo.")

            boton_siguiente = page.locator("span.ui-paginator-next:not(.ui-state-disabled)")
            if boton_siguiente.count():
                boton_siguiente.first.click()
                try:
                    page.wait_for_load_state("networkidle", timeout=1000)
                except Exception:
                    pass
                time.sleep(0.2)
                pagina += 1
                continue
            break

    resultado = {
        "estado": "ok",
        "n_xml": n_xml,
        "n_pdf": n_pdf,
        "carpeta_tipo": str(carpeta_tipo),
        "tipo_slug": tipo_slug,
        "tipo_visible": tipo_visible,
        "txt_dir": str(txt_dir),
        "xml_dir": str(xml_dir),
        "pdf_dir": str(pdf_dir),
    }
    if txt_path:
        resultado["txt"] = str(txt_path)
    return resultado

# ============================================================
# ?? LECTURA DE TABLA PARA COMPROBANTES EMITIDOS (sin TXT)
# ============================================================


def _flujo_emitidos(
    page,
    destino: Path,
    fecha_emision: Optional[str],
    tipo: str,
    estado_autorizacion: Optional[str],
    establecimiento: Optional[str],
    punto_emision: Optional[str],
    formatos: list,
):
    fecha_emision = (fecha_emision or "").strip()
    estado_autorizacion = (estado_autorizacion or "").strip()
    establecimiento = (establecimiento or "").strip()
    punto_emision = (punto_emision or "").strip()
    if punto_emision:
        punto_emision = "".join(ch for ch in punto_emision if ch.isdigit())[:3]

    formatos_norm = [
        (fmt or "").strip().upper() for fmt in (formatos or []) if isinstance(fmt, str)
    ]
    descargar_pdf = "PDF" in formatos_norm
    descargar_xml = "XML" in formatos_norm

    tipo_visible = TIPOS_MAP.get(tipo, tipo)
    try:
        page.wait_for_load_state("domcontentloaded", timeout=1000)
    except Exception:
        pass

    if not _seleccionar_en_select(
        page,
        "select#frmPrincipal\:cmbTipoComprobante",
        tipo_visible,
        tipo,
    ):
        if not _seleccionar(page, "Tipo de comprobante", tipo_visible):
            print(f"[WARN] No se pudo seleccionar el tipo de comprobante '{tipo_visible}' en Emitidos.")

    estado_visible = ESTADOS_EMITIDOS_MAP.get(estado_autorizacion, estado_autorizacion)
    if estado_visible:
        if not _seleccionar_en_select(
            page,
            "select#frmPrincipal\:cmbEstadoAutorizacion",
            estado_visible,
        ):
            etiquetas_estado = [
                "Estado autorizacion",
                "Estado autorizacion",
                "Estado",
            ]
            seleccionado = any(_seleccionar(page, etiqueta, estado_visible) for etiqueta in etiquetas_estado)
            if not seleccionado:
                print(
                    f"[WARN] No se pudo seleccionar el estado de autorizacion '{estado_visible}' en Emitidos."
                )

    if fecha_emision:
        fecha_selector = "input#frmPrincipal\:calendarFechaDesde_input"
        fecha_ok = False
        try:
            fecha_loc = page.locator(fecha_selector)
            if fecha_loc.count():
                fecha_loc.first.fill("")
                fecha_loc.first.fill(fecha_emision)
                try:
                    fecha_loc.first.dispatch_event("input")
                    fecha_loc.first.dispatch_event("change")
                except Exception:
                    pass
                fecha_ok = True
        except Exception:
            fecha_ok = False
        if not fecha_ok:
            if not _rellenar_input_por_label(
                page,
                ["Fecha de emision", "Fecha emision", "Fecha"],
                fecha_emision,
                EMITIDOS_FECHA_SELECTORS,
            ):
                print(f"[WARN] No se pudo completar la fecha de emision con '{fecha_emision}' en Emitidos.")

    est_valor = establecimiento or ""
    if not est_valor or est_valor.lower() == "todos":
        est_valor = "Todos"
        if not _seleccionar_en_select(
            page,
            "select#frmPrincipal\:cmbEstablecimiento",
            est_valor,
        ):
            _rellenar_input_por_label(
                page,
                ["Establecimiento"],
                est_valor,
                EMITIDOS_ESTABLECIMIENTO_SELECTORS,
            )
    else:
        if not _seleccionar_en_select(
            page,
            "select#frmPrincipal\:cmbEstablecimiento",
            est_valor,
        ):
            if not _rellenar_input_por_label(
                page,
                ["Establecimiento"],
                est_valor,
                EMITIDOS_ESTABLECIMIENTO_SELECTORS,
            ):
                print(f"[WARN] No se pudo establecer el establecimiento '{est_valor}' en Emitidos.")

    punto_valor = punto_emision
    if est_valor and est_valor.lower() != "todos" and not punto_valor:
        raise RuntimeError("El punto de emision es obligatorio cuando se especifica un establecimiento.")
    if punto_valor:
        punto_selector = "input#frmPrincipal\:txtPuntoEmision"
        punto_ok = False
        try:
            punto_loc = page.locator(punto_selector)
            if punto_loc.count():
                punto_loc.first.fill("")
                punto_loc.first.fill(punto_valor)
                punto_ok = True
        except Exception:
            punto_ok = False
        if not punto_ok:
            if not _rellenar_input_por_label(
                page,
                ["Punto de emision", "Punto emision", "Punto"],
                punto_valor,
                EMITIDOS_PUNTO_SELECTORS,
            ):
                print(f"[WARN] No se pudo establecer el punto de emision '{punto_valor}' en Emitidos.")

    estado_nombre = (estado_visible or "Sin Estado").strip() or "Sin Estado"
    estado_normalizado = unicodedata.normalize("NFKD", estado_nombre).encode("ascii", "ignore").decode("ascii")
    estado_slug = re.sub(r"[^A-Za-z0-9]+", "_", estado_normalizado).strip("_") or "Sin_Estado"

    fecha_dt = _parse_datetime_local(fecha_emision) if fecha_emision else None
    if not fecha_dt:
        fecha_dt = datetime.now()
    anio_dir = f"{fecha_dt.year:04d}"
    mes_dir = _mes_a_texto(fecha_dt.month)
    dia_dir = f"{fecha_dt.day:02d}"

    _orden_tipo, _label_tipo, tipo_prefijo = _prefijo_tipo(tipo_visible or tipo)
    tipo_slug = _slug_tipo(tipo_visible or tipo)

    carpeta_estado = destino / estado_slug / anio_dir / mes_dir / dia_dir
    carpeta_estado.mkdir(parents=True, exist_ok=True)
    carpeta_tipo = carpeta_estado / tipo_prefijo
    carpeta_tipo.mkdir(parents=True, exist_ok=True)
    xml_dir = carpeta_tipo / "XML"
    if descargar_xml:
        xml_dir.mkdir(parents=True, exist_ok=True)
    pdf_dir = carpeta_tipo / "PDF"
    if descargar_pdf:
        pdf_dir.mkdir(parents=True, exist_ok=True)

    if not _click_texto(page, "Consultar"):
        try:
            page.locator("button[id$='consultar']").first.click()
        except Exception:
            try:
                page.locator("input[id$='consultar']").first.click()
            except Exception:
                page.keyboard.press("Enter")

            try:
                page.wait_for_load_state("networkidle", timeout=1000)
            except Exception:
                pass
    time.sleep(0.2)

    tabla_emitidos = page.locator("#frmPrincipal\\:tablaCompEmitidos_data")
    try:
        tabla_emitidos.wait_for(state="visible", timeout=1000)
    except Exception:
        pass

    html = page.content()
    rows = re.findall(
        r"<tr[^>]*>\s*(.*?)\s*</tr>",
        html,
        flags=re.DOTALL
    )

    encabezado_textos = ["EMITIDOS", "", "CLA VE ACCESO"]  # placeholder; se ajustara con row data
    data = []
    for r in rows:
        cols = re.findall(r"<td[^>]*>(.*?)</td>", r, flags=re.DOTALL)
        textos = [re.sub("<.*?>", "", c).strip() for c in cols]
        if len(textos) < 3:
            continue
        fecha_emision_col = textos[0]
        comprobante_raw = textos[1]
        partes_tipo = comprobante_raw.split()
        comprobante = partes_tipo[0] if partes_tipo else ""
        serie = " ".join(partes_tipo[1:]) if len(partes_tipo) > 1 else ""
        clave = textos[2]
        fecha_autorizacion = textos[3] if len(textos) > 3 else ""
        valor_sin_impuestos = textos[5] if len(textos) > 5 else ""
        iva_val = textos[6] if len(textos) > 6 else ""
        importe_total = textos[7] if len(textos) > 7 else ""
        if valor_sin_impuestos and not re.search(r"\d", valor_sin_impuestos):
            valor_sin_impuestos = ""
        if iva_val and not re.search(r"\d", iva_val):
            iva_val = ""
        if importe_total and not re.search(r"\d", importe_total):
            importe_total = ""

        data.append({
            "COMPROBANTE": comprobante,
            "SERIE_COMPROBANTE": serie,
            "CLAVE_ACCESO": clave,
            "FECHA_AUTORIZACION": fecha_autorizacion,
            "FECHA_EMISION": fecha_emision_col,
            "VALOR_SIN_IMPUESTOS": valor_sin_impuestos,
            "IVA": iva_val,
            "IMPORTE_TOTAL": importe_total,
        })

    n_pdf = 0
    n_xml = 0

    info_base = {
        "carpeta_tipo": str(carpeta_tipo),
        "tipo_slug": tipo_slug,
        "tipo_visible": tipo_visible,
        "fecha_filtro": fecha_emision,
        "estado_autorizacion": estado_visible,
        "establecimiento": est_valor,
        "punto_emision": punto_valor,
        "carpeta_estado": str(carpeta_estado),
        "xml_dir": str(xml_dir) if descargar_xml else "",
        "pdf_dir": str(pdf_dir) if descargar_pdf else "",
        "n_xml": n_xml,
        "n_pdf": n_pdf,
        "reporte_xml": "",
    }

    if not data:
        info_base.update({
            "estado": "sin_resultados",
            "mensaje": "No se encontraron filas en la tabla",
            "n_registros": 0,
        })
        return info_base

    if descargar_pdf or descargar_xml:
        try:
            tabla_emitidos.wait_for(state="visible", timeout=1000)
        except Exception:
            pass
        pagina = 1
        claves_guardadas = set()
        request_context = page.context.request

        while True:
            filas = tabla_emitidos.locator("tr")
            total_filas = filas.count()
            for idx in range(total_filas):
                fila = filas.nth(idx)
                celdas = fila.locator("td")
                if celdas.count() < 3:
                    continue
                try:
                    tipo_serie_texto = celdas.nth(1).inner_text().strip()
                except Exception:
                    tipo_serie_texto = ""
                try:
                    clave_texto = celdas.nth(2).inner_text().strip()
                except Exception:
                    clave_texto = ""
                try:
                    razon_texto = celdas.nth(4).inner_text().strip() if celdas.count() > 4 else ""
                except Exception:
                    razon_texto = ""

                tipo_serie_completo = " ".join(
                    fragment for fragment in [tipo_serie_texto, clave_texto] if fragment
                ).strip()
                nombre_base_pdf = _sanear_nombre_archivo(
                    tipo_serie_completo or razon_texto or f"emitido_{pagina}_{idx+1}"
                )

                if descargar_xml:
                    if clave_texto:
                        try:
                            resultado_xml = _descargar_xml_emitido_por_clave(
                                request_context,
                                clave_texto,
                                xml_dir,
                                nombre_base_pdf,
                                claves_guardadas,
                            )
                            if resultado_xml:
                                n_xml += 1
                        except Exception as err:
                            print(f"[WARN] No se pudo obtener XML SOAP para '{nombre_base_pdf}': {err}")
                    else:
                        print(f"[WARN] La fila '{nombre_base_pdf}' no tiene clave de acceso para solicitar el XML.")

                if descargar_pdf:
                    link_pdf = fila.locator("a[id$=':lnkPdf']")
                    if not link_pdf.count():
                        link_pdf = fila.locator("a[title*='pdf' i], button[title*='pdf' i]")
                    if not link_pdf.count():
                        continue

                    destino_pdf = pdf_dir / f"{nombre_base_pdf}.pdf"
                    resultado_pdf = _guardar_pdf_desde_enlace(page, link_pdf.first, destino_pdf)
                    if resultado_pdf:
                        n_pdf += 1
                    else:
                        print(f"[WARN] No se pudo descargar PDF para '{nombre_base_pdf}': no se obtuvo archivo.")

            boton_siguiente = page.locator("span.ui-paginator-next:not(.ui-state-disabled)")
            if boton_siguiente.count():
                boton_siguiente.first.click()
                try:
                    page.wait_for_load_state("networkidle", timeout=1000)
                except Exception:
                    pass
                time.sleep(0.2)
                pagina += 1
                continue
            break
        info_base["n_xml"] = n_xml
        info_base["n_pdf"] = n_pdf

        fecha_slug = re.sub(r"[^0-9]+", "", fecha_emision) or "consulta"
        if descargar_xml and n_xml > 0:
            xml_report_path = carpeta_tipo / f"emitidos_reporte_xml_{tipo_slug}_{fecha_slug}.xlsx"
            if xml_report_path.exists():
                try:
                    xml_report_path.unlink()
                except PermissionError:
                    sufijo_xml = 1
                    while True:
                        candidato = carpeta_tipo / f"emitidos_reporte_xml_{tipo_slug}_{fecha_slug}_{sufijo_xml}.xlsx"
                        if not candidato.exists():
                            xml_report_path = candidato
                            break
                        sufijo_xml += 1
            try:
                construir_reporte(xml_dir, xml_report_path)
                info_base["reporte_xml"] = str(xml_report_path)
            except Exception as err:
                print(f"[WARN] No se pudo construir el reporte XML de emitidos: {err}")

    columns_order = [
        "COMPROBANTE",
        "SERIE_COMPROBANTE",
        "CLAVE_ACCESO",
        "FECHA_AUTORIZACION",
        "FECHA_EMISION",
        "VALOR_SIN_IMPUESTOS",
        "IVA",
        "IMPORTE_TOTAL",
    ]
    df = pd.DataFrame(data, columns=columns_order)

    def _coerce_decimal_value(val):
        if isinstance(val, (int, float)):
            return float(val)
        parsed = _parse_decimal(val) if isinstance(val, str) else None
        return parsed if parsed is not None else val

    def _coerce_datetime_value(val):
        if isinstance(val, datetime):
            return val
        parsed = _parse_datetime_local(val) if isinstance(val, str) else None
        return parsed if parsed is not None else val

    for columna in ["VALOR_SIN_IMPUESTOS", "IVA", "IMPORTE_TOTAL"]:
        df[columna] = df[columna].apply(_coerce_decimal_value)

    for columna in ["FECHA_AUTORIZACION"]:
        df[columna] = df[columna].apply(_coerce_datetime_value)

    fecha_slug = re.sub(r"[^0-9]+", "", fecha_emision) or "consulta"
    excel_nombre_base = f"emitidos_reporte_{tipo_slug}_{fecha_slug}.xlsx"
    excel_path = carpeta_tipo / excel_nombre_base
    if excel_path.exists():
        try:
            excel_path.unlink()
        except PermissionError:
            sufijo_excel = 1
            while True:
                candidato = carpeta_tipo / f"emitidos_reporte_{tipo_slug}_{fecha_slug}_{sufijo_excel}.xlsx"
                if not candidato.exists():
                    excel_path = candidato
                    break
                sufijo_excel += 1

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        sheet_name = "Emitidos"
        df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=1)
        ws = writer.sheets[sheet_name]
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns_order))
        titulo = ws.cell(row=1, column=1, value="EMITIDOS")
        titulo.font = Font(bold=True, size=14)
        titulo.alignment = Alignment(horizontal="center", vertical="center")

        header_fill = PatternFill("solid", fgColor="305496")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[2]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.freeze_panes = "A3"

        text_columns = {"COMPROBANTE", "SERIE_COMPROBANTE", "CLAVE_ACCESO", "FECHA_EMISION"}
        numeric_columns = {"VALOR_SIN_IMPUESTOS", "IVA", "IMPORTE_TOTAL"}
        date_columns = {"FECHA_AUTORIZACION"}

        for idx, column in enumerate(columns_order, start=1):
            max_len = len(column)
            for cell_tuple in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=idx, max_col=idx):
                cell = cell_tuple[0]
                valor = cell.value
                if valor is None:
                    continue
                if isinstance(valor, datetime):
                    texto_len = len(valor.strftime("%d/%m/%Y %H:%M"))
                else:
                    texto_len = len(str(valor))
                if texto_len > max_len:
                    max_len = texto_len
            ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 45)

        for columna in text_columns:
            if columna not in columns_order:
                continue
            col_idx = columns_order.index(columna) + 1
            for cell_tuple in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                celda = cell_tuple[0]
                celda.number_format = "@"
                celda.alignment = Alignment(horizontal="left", vertical="center")

        for columna in numeric_columns:
            if columna not in columns_order:
                continue
            col_idx = columns_order.index(columna) + 1
            for cell_tuple in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                celda = cell_tuple[0]
                if isinstance(celda.value, (int, float)):
                    celda.number_format = "#,##0.00"
                    celda.alignment = Alignment(horizontal="right", vertical="center")

        for columna in date_columns:
            if columna not in columns_order:
                continue
            col_idx = columns_order.index(columna) + 1
            for cell_tuple in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                celda = cell_tuple[0]
                if isinstance(celda.value, datetime):
                    celda.number_format = "dd/mm/yyyy hh:mm"
                    celda.alignment = Alignment(horizontal="center", vertical="center")
    info_base.update({
        "estado": "ok",
        "n_registros": len(df),
        "reporte": str(excel_path),
    })
    return info_base

# ============================================================
# FUNCIoN PRINCIPAL
# ============================================================
def descargar_sri(
    ruc: str,
    clave: str,
    anio: int,
    mes: int,
    dia: int,
    tipo: str,
    formatos: list,
    destino: Path,
    origen: str = "Recibidos",
    ci_adicional: Optional[str] = None,
    fecha_emitidos: Optional[str] = None,
    estado_emitidos: Optional[str] = None,
    establecimiento: Optional[str] = None,
    punto_emision: Optional[str] = None,
):
    destino.mkdir(parents=True, exist_ok=True)
    destino_recibidos = destino / "Recibidos"
    destino_emitidos = destino / "Emitidos"
    destino_recibidos.mkdir(parents=True, exist_ok=True)
    destino_emitidos.mkdir(parents=True, exist_ok=True)
    destino_objetivo = destino
    cookies_path = Path(f"cookies_{ruc}.json")

    with sync_playwright() as p:
        launch_kwargs = dict(
            headless=HEADLESS,
            args=["--no-sandbox", "--disable-setuid-sandbox", "--disable-dev-shm-usage", "--disable-gpu"],
        )
        if SLOW_MO > 0:
            launch_kwargs["slow_mo"] = SLOW_MO
        browser = p.chromium.launch(**launch_kwargs)
        context = browser.new_context(accept_downloads=True)
        page = context.new_page()

        destino_url = PORTAL_HOME if origen in {"Recibidos", "Emitidos"} else URLS.get(origen, URLS["Recibidos"])
        _login(context, page, ruc, clave, cookies_path, destino_url, ci_adicional=ci_adicional)
        if "auth/realms" in page.url:
            raise RuntimeError("La autenticacion en el SRI fallo, se mantuvo en la pantalla de login.")
        modulo_page = None
        if origen in {"Recibidos", "Emitidos"}:
            modulo_page = _abrir_modulo_consultas(page, origen)
            _resolver_captcha(modulo_page, f"{origen.lower()}_Modulo")
            if origen == "Recibidos":
                destino_objetivo = destino_recibidos
                resultado = _flujo_recibidos(modulo_page, destino_objetivo, anio, mes, dia, tipo, formatos)
            else:
                destino_objetivo = destino_emitidos
                resultado = _flujo_emitidos(
                    modulo_page,
                    destino_objetivo,
                    fecha_emitidos,
                    tipo,
                    estado_emitidos,
                    establecimiento,
                    punto_emision,
                    formatos,
                )
        else:
            if page.url != destino_url:
                try:
                    page.goto(destino_url, timeout=1000)
                except Exception:
                    pass
                try:
                    page.wait_for_load_state("domcontentloaded", timeout=1000)
                except Exception:
                    pass
            _resolver_captcha(page, "recibidos_directo")
            destino_objetivo = destino_recibidos
            resultado = _flujo_recibidos(page, destino_objetivo, anio, mes, dia, tipo, formatos)

        if isinstance(resultado, dict):
            resultado.setdefault("carpeta_base", str(destino_objetivo))

        try:
            paginas_logout = []
            if modulo_page:
                paginas_logout.append(modulo_page)
            paginas_logout.append(page)
            for extra in context.pages:
                if extra not in paginas_logout:
                    paginas_logout.append(extra)
            for objetivo in paginas_logout:
                if _cerrar_sesion(objetivo):
                    break
        except Exception as err:
            print(f"[WARN] No se pudo cerrar la sesion del SRI: {err}")

        browser.close()
        return resultado










s