"""Cruce de Notas de Credito Emitidas vs Facturas modificadas.

Modulo aislado del flujo normal de descarga: lee NC desde una carpeta ya
descargada por la app, busca la Factura modificada (primero localmente,
luego en el portal del SRI si no esta), calcula el valor neto y exporta
todo a un Excel con columnas estandar + Estado + Observacion.

Diseño:
- Solo procesa NC Emitidas (cod=04) que modifican Facturas (codDocModificado=01).
- Reutiliza completamente las funciones de extraccion XML/PDF existentes
  (xml_extraction, pdf_extraction) — no reimplementa nada.
- La busqueda remota Playwright se hace en una sesion INDEPENDIENTE del
  flujo de Descarga: distintos cookies, distinto contexto. No interfiere
  con descargas en paralelo.
- Si una NC no puede procesarse (falta de campos, error de IO, etc.) NO
  rompe el reporte: queda como fila con Estado="Error" y Observacion
  explicando el motivo.

Punto de entrada principal: `generar_reporte_valor_neto(params, ...)`.
"""

from __future__ import annotations

import json
import logging
import re
import threading
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from robot.file_utils import _mes_a_texto

logger = logging.getLogger("robot.nc_vs_factura")


# =============================================================================
# Constantes de columnas del reporte final
# =============================================================================

COLUMNAS_REPORTE = [
    "RUC",
    "Fecha nota de credito",
    "Serie nota de credito",
    "Clave acceso nota de credito",
    "Valor total nota de credito",
    "Factura modificada",
    "Fecha factura modificada",
    "Clave acceso factura",
    "Valor total factura",
    "Valor neto",
    "Estado",
    "Observacion",
]

# Columnas forzadas a texto en Excel (claves de acceso largas + secuenciales
# que Excel interpretaria como numeros y perderia ceros a la izquierda).
COLUMNAS_TEXT_FORCE = {
    "RUC",
    "Serie nota de credito",
    "Clave acceso nota de credito",
    "Factura modificada",
    "Clave acceso factura",
}

# Columnas numericas (con formato monetario en el Excel).
COLUMNAS_NUMERICAS = {
    "Valor total nota de credito",
    "Valor total factura",
    "Valor neto",
}


# =============================================================================
# Utilidades de parseo y normalizacion
# =============================================================================


def _parse_fecha_es(valor: str) -> Optional[datetime]:
    """Parsea fechas en formato del SRI: DD/MM/YYYY o YYYY-MM-DD.

    Devuelve None si no matchea.
    """
    if not valor:
        return None
    valor = str(valor).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(valor, fmt)
        except ValueError:
            continue
    return None


def _safe_float(valor: Any) -> Optional[float]:
    """Convierte un valor a float aceptando coma decimal del SRI. None si no se puede."""
    if valor is None or valor == "":
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    txt = str(valor).strip().replace(",", ".")
    txt = re.sub(r"[^0-9.\-]", "", txt)
    try:
        return float(txt)
    except (ValueError, TypeError):
        return None


def _normalizar_secuencial(num_doc: str) -> str:
    """Normaliza un secuencial del SRI a 'NNN-NNN-NNNNNNNNN'.

    Acepta variantes: '002-002-000110921', '002002000110921',
    '2-2-110921' (sin padding). Devuelve string vacio si no parsea.
    """
    if not num_doc:
        return ""
    txt = str(num_doc).strip()
    # Quitar separadores y dejar solo digitos
    partes = re.split(r"[-_\s]+", txt)
    if len(partes) == 3:
        try:
            est = int(partes[0])
            pto = int(partes[1])
            sec = int(partes[2])
            return f"{est:03d}-{pto:03d}-{sec:09d}"
        except ValueError:
            pass
    # Sin separadores: 15 digitos = 3+3+9
    solo_dig = re.sub(r"\D", "", txt)
    if len(solo_dig) == 15:
        return f"{solo_dig[:3]}-{solo_dig[3:6]}-{solo_dig[6:]}"
    return txt  # devolver como vino si no se puede normalizar


# =============================================================================
# Localizacion de archivos en la carpeta del sistema
# =============================================================================


def _es_xml_de_nc(path: Path) -> bool:
    """Heuristica rapida: nombre del archivo empieza con 'Nota' o el tipo
    'NotaCredito' (el sistema usa _nombre_documento_mes con prefijo del tipo).
    """
    if path.suffix.lower() != ".xml":
        return False
    name = path.name.lower()
    return (
        name.startswith("nota")
        or "credito" in name
        or "nc_" in name
        or "_nc_" in name
    )


def _es_pdf_de_nc(path: Path) -> bool:
    if path.suffix.lower() != ".pdf":
        return False
    name = path.name.lower()
    return (
        name.startswith("nota")
        or "credito" in name
        or "nc_" in name
        or "_nc_" in name
    )


def _listar_notas_credito_en_carpeta(carpeta: Path) -> list[tuple[Path, str]]:
    """Devuelve una lista de tuplas (path, source) donde source es 'xml' o 'pdf'.

    Recorre recursivamente carpeta/ y se queda solo con archivos que parecen
    Notas de Credito. Si una NC tiene XML y PDF, se queda solo con el XML
    (preferimos XML por confiabilidad).
    """
    xml_files: dict[str, Path] = {}
    pdf_files: dict[str, Path] = {}
    for p in carpeta.rglob("*"):
        if not p.is_file():
            continue
        if _es_xml_de_nc(p):
            xml_files[p.stem] = p
        elif _es_pdf_de_nc(p):
            pdf_files[p.stem] = p

    resultado: list[tuple[Path, str]] = []
    for stem, xml_path in xml_files.items():
        resultado.append((xml_path, "xml"))
    for stem, pdf_path in pdf_files.items():
        if stem not in xml_files:
            resultado.append((pdf_path, "pdf"))
    return sorted(resultado, key=lambda t: t[0].name.lower())


def _inferir_base_descargas(nc_path: Path) -> Optional[Path]:
    """Dada la ruta de una NC dentro de la estructura del sistema, devuelve
    la carpeta [base]/[RUC] desde donde nace el arbol Emitidos/Recibidos.

    Estructura esperada:
        [base]/[RUC]/Emitidos/Autorizados/Notas de Credito/[Anio]/[Mes]/XML/file.xml
                              ^ buscamos este nivel y devolvemos su parent

    Devuelve None si no encuentra 'Emitidos' en el path.
    """
    for ancestro in nc_path.parents:
        if ancestro.name == "Emitidos":
            return ancestro.parent  # parent de Emitidos = carpeta [RUC] (o equivalente)
    return None


def _construir_ruta_factura_mes(
    base_ruc: Path, fecha_factura: datetime, formato: str = "XML"
) -> Path:
    """Construye la ruta esperada para Facturas Emitidas Autorizadas de un mes.

    [base_ruc]/Emitidos/Autorizados/Facturas/[YYYY]/[Mes_texto]/[formato]/
    """
    return (
        base_ruc
        / "Emitidos"
        / "Autorizados"
        / "Facturas"
        / f"{fecha_factura.year:04d}"
        / _mes_a_texto(fecha_factura.month)
        / formato
    )


def _buscar_factura_local(
    base_ruc: Path, secuencial: str, fecha_factura: datetime
) -> Optional[tuple[Path, str]]:
    """Busca la Factura emitida en la estructura local.

    Estrategia:
    1. Construir la ruta del mes correspondiente
    2. Hacer glob por secuencial (los archivos llevan el secuencial en el nombre,
       patron de _nombre_documento_mes: 'Factura__YYYYMMDD__...secuencial...').
    3. Preferir XML; fallback PDF.

    Devuelve (path, 'xml'|'pdf') o None.
    """
    if not secuencial:
        return None
    # El secuencial puede aparecer en el nombre con o sin guiones segun
    # el patron de _construir_nombre_xml_emitido. Probamos ambas formas.
    sec_normalizado = _normalizar_secuencial(secuencial)
    sec_sin_guiones = sec_normalizado.replace("-", "")
    patrones = [f"*{sec_normalizado}*", f"*{sec_sin_guiones}*"]

    # Si tenemos fecha, buscar solo en el mes correspondiente (mas rapido).
    # Si no, hacer un walk completo bajo Emitidos/Autorizados/Facturas.
    carpetas_busqueda: list[Path] = []
    if fecha_factura is not None:
        carpetas_busqueda.append(_construir_ruta_factura_mes(base_ruc, fecha_factura, "XML"))
        carpetas_busqueda.append(_construir_ruta_factura_mes(base_ruc, fecha_factura, "PDF"))
    else:
        raiz = base_ruc / "Emitidos" / "Autorizados" / "Facturas"
        if raiz.is_dir():
            carpetas_busqueda.extend([d for d in raiz.rglob("XML") if d.is_dir()])
            carpetas_busqueda.extend([d for d in raiz.rglob("PDF") if d.is_dir()])

    # Buscar XML primero
    for carpeta in carpetas_busqueda:
        if not carpeta.is_dir() or carpeta.name != "XML":
            continue
        for patron in patrones:
            for archivo in carpeta.glob(patron):
                if archivo.is_file() and archivo.suffix.lower() == ".xml":
                    return (archivo, "xml")
    # Fallback PDF
    for carpeta in carpetas_busqueda:
        if not carpeta.is_dir() or carpeta.name != "PDF":
            continue
        for patron in patrones:
            for archivo in carpeta.glob(patron):
                if archivo.is_file() and archivo.suffix.lower() == ".pdf":
                    return (archivo, "pdf")
    return None


# =============================================================================
# Extraccion de datos (delegada a las funciones existentes del robot)
# =============================================================================


def _extraer_datos_nc(nc_path: Path, source: str) -> dict:
    """Extrae los campos relevantes de una NC. Reutiliza extractors existentes.

    Devuelve dict con: ruc, fecha_emision, secuencial, clave_acceso,
    importe_total, num_doc_modificado, fecha_emision_doc_sustento,
    cod_doc_modificado, _err (mensaje de error si fallo).
    """
    out = {
        "ruc": "",
        "fecha_emision": "",
        "secuencial": "",
        "clave_acceso": "",
        "importe_total": None,
        "num_doc_modificado": "",
        "fecha_emision_doc_sustento": "",
        "cod_doc_modificado": "",
        "_err": "",
    }
    try:
        if source == "xml":
            from robot.xml_extraction import _extraer_datos_xml_nota_credito_emitido
            datos = _extraer_datos_xml_nota_credito_emitido(nc_path)
        else:
            from robot.pdf_extraction import (
                _extraer_datos_pdf_nota_credito_emitido,
            )
            datos = _extraer_datos_pdf_nota_credito_emitido(nc_path)
    except Exception as exc:
        out["_err"] = f"No se pudo leer la NC ({source.upper()}): {exc}"
        return out

    if not isinstance(datos, dict):
        out["_err"] = "Datos invalidos en la NC."
        return out

    out["ruc"] = str(datos.get("RUC Emisor") or "").strip()
    out["fecha_emision"] = str(datos.get("Fecha de Emisión") or "").strip()
    out["secuencial"] = str(datos.get("Secuencial") or "").strip()
    out["clave_acceso"] = str(datos.get("Clave de Acceso") or "").strip()
    out["importe_total"] = _safe_float(datos.get("Importe Total"))
    out["num_doc_modificado"] = str(datos.get("Número Documento Modificado") or "").strip()
    out["fecha_emision_doc_sustento"] = str(
        datos.get("Fecha Emisión Doc. Sustento") or ""
    ).strip()
    out["cod_doc_modificado"] = str(datos.get("Código Documento Modificado") or "").strip()
    return out


def _extraer_datos_factura(factura_path: Path, source: str) -> dict:
    """Extrae los campos relevantes de una Factura emitida.

    Devuelve dict con: ruc, fecha_emision, secuencial, clave_acceso,
    importe_total, _err.
    """
    out = {
        "ruc": "",
        "fecha_emision": "",
        "secuencial": "",
        "clave_acceso": "",
        "importe_total": None,
        "_err": "",
    }
    try:
        if source == "xml":
            from robot.xml_extraction import _extraer_datos_xml_factura_emitido
            datos = _extraer_datos_xml_factura_emitido(factura_path)
        else:
            from robot.pdf_extraction import _extraer_datos_pdf_factura_emitido
            datos = _extraer_datos_pdf_factura_emitido(factura_path)
    except Exception as exc:
        out["_err"] = f"No se pudo leer la Factura ({source.upper()}): {exc}"
        return out

    if not isinstance(datos, dict):
        out["_err"] = "Datos invalidos en la Factura."
        return out

    out["ruc"] = str(datos.get("RUC Emisor") or "").strip()
    out["fecha_emision"] = str(datos.get("Fecha de Emisión") or "").strip()
    out["secuencial"] = str(datos.get("Secuencial") or "").strip()
    out["clave_acceso"] = str(datos.get("Clave de Acceso") or "").strip()
    out["importe_total"] = _safe_float(datos.get("Importe Total"))
    return out


# =============================================================================
# Busqueda REMOTA en el portal SRI (Playwright)
# =============================================================================


def _normalizar_serie_para_match(texto: str) -> str:
    """Quita guiones, espacios y leading zeros para comparacion robusta."""
    if not texto:
        return ""
    return re.sub(r"[\s\-_]", "", str(texto)).lstrip("0")


def _esperar_overlay_cerrado(page, timeout_ms: int = 15000) -> bool:
    """Espera a que el dialogo bloqueante de PrimeFaces se oculte.

    El SRI muestra `<div id="dlgpopStatusPrime_modal" class="ui-widget-overlay">`
    como mascara de "procesando" durante CADA AJAX (Consultar, paginar,
    abrir Emitidos, cambiar filtros). Mientras esta visible, todos los
    clicks se interceptan con "intercepts pointer events" en Playwright.

    Devuelve True si el overlay no esta visible (o no existe). False si
    se agoto el timeout.
    """
    try:
        page.wait_for_function(
            """() => {
                const el = document.getElementById('dlgpopStatusPrime_modal');
                if (!el) return true;
                const style = window.getComputedStyle(el);
                if (style.display === 'none' || style.visibility === 'hidden') return true;
                if (el.offsetWidth === 0 && el.offsetHeight === 0) return true;
                return false;
            }""",
            timeout=timeout_ms,
        )
        return True
    except Exception:
        return False


def _set_fecha_emitidos_via_js(page, fecha_display: str) -> bool:
    """Setea el input de fecha desde via JavaScript directamente.

    PrimeFaces calendar mantiene estado interno que NO se actualiza con
    Playwright's `fill()` + `dispatch_event()` porque esos eventos tienen
    isTrusted=false. Setear `el.value = ...` + disparar eventos con
    `bubbles:true` desde el contexto de la pagina propia bypassea esa
    deteccion.

    Probado en los logs del 2026-06-18: con fill+dispatch_event, el SRI
    devolvia resultados de fechas aleatorias (no la pedida). Con este
    metodo, el filtro se aplica de verdad.
    """
    try:
        ok = page.evaluate(
            """(val) => {
                const el = document.getElementById('frmPrincipal:calendarFechaDesde_input');
                if (!el) return false;
                el.focus();
                el.value = '';
                el.dispatchEvent(new Event('input', {bubbles: true}));
                el.value = val;
                el.dispatchEvent(new Event('input', {bubbles: true}));
                el.dispatchEvent(new Event('change', {bubbles: true}));
                el.dispatchEvent(new Event('blur', {bubbles: true}));
                return true;
            }""",
            fecha_display,
        )
        return bool(ok)
    except Exception:
        return False


def _esperar_tabla_emitidos_lista(page, timeout_ms: int = 20000) -> bool:
    """Espera a que la tabla de Emitidos tenga resultados o muestre vacio.

    Despues de un Consultar, esperamos a que ocurra UNO de:
      - El tbody tiene al menos un <tr> con celdas (resultados)
      - Aparece el mensaje "No se encontraron registros" (sin resultados)

    `networkidle` solo no alcanza porque el partial AJAX de PrimeFaces
    puede completarse sin que el DOM termine de renderizarse.
    """
    try:
        page.wait_for_function(
            """() => {
                const t = document.getElementById('frmPrincipal:tablaCompEmitidos_data');
                if (!t) return false;
                if (t.querySelector('tr td')) return true;
                const empty = t.querySelector('.ui-datatable-empty-message');
                if (empty) return true;
                return false;
            }""",
            timeout=timeout_ms,
        )
        return True
    except Exception:
        return False


def _buscar_facturas_remoto(
    ruc: str,
    clave: str,
    pendientes: list[dict],
    *,
    cancel_event: Optional[threading.Event] = None,
    progress: Optional[Callable[[str], None]] = None,
) -> dict[str, dict]:
    """Busca en el portal SRI las facturas que no se encontraron localmente.

    `pendientes` es una lista de dicts con minimo: secuencial, fecha (datetime).
    Devuelve dict { secuencial_normalizado: {importe_total, clave_acceso, fecha,
    razon_social, _err} } con lo que SI encontro. Las que no encuentre quedan
    fuera del dict (el caller las marca como "Factura no encontrada en SRI").

    Estrategia:
    - Agrupa pendientes por fecha (1 consulta SRI por fecha unica)
    - Para cada fecha:
      - filtra fecha + tipo=Factura + estado=Autorizados
      - lee la tabla
      - para cada secuencial pendiente de esa fecha busca su fila
    - Maneja cancelacion via cancel_event en cada iteracion
    """
    encontradas: dict[str, dict] = {}
    if not pendientes:
        return encontradas

    def _emit(msg: str) -> None:
        if progress:
            try:
                progress(msg)
            except Exception:
                pass
        logger.info(msg)

    # Agrupar por fecha (YYYY-MM-DD) para minimizar consultas al portal.
    por_fecha: dict[str, list[dict]] = {}
    for item in pendientes:
        fecha_dt: Optional[datetime] = item.get("fecha")
        if fecha_dt is None:
            continue
        key = fecha_dt.strftime("%Y-%m-%d")
        por_fecha.setdefault(key, []).append(item)

    if not por_fecha:
        _emit("No hay fechas validas en las NC pendientes — se omite busqueda remota.")
        return encontradas

    _emit(
        f"Iniciando busqueda remota en SRI: {len(por_fecha)} fecha(s) distinta(s) "
        f"para resolver {len(pendientes)} Factura(s) faltante(s)."
    )

    from playwright.sync_api import sync_playwright  # import perezoso

    # Cookies aisladas para esta busqueda — NO mezclar con las de descarga.
    cookies_path = Path(f"cookies_nc_lookup_{ruc}.json")

    # Reusamos las funciones internas del robot existente.
    from robot.downloader import _login, _abrir_navegador
    from robot.browser import _abrir_modulo_consultas, _seleccionar_en_select
    from robot.config import PORTAL_HOME

    with sync_playwright() as p:
        context, browser, _persistent = _abrir_navegador(p)
        try:
            page = context.pages[0] if context.pages else context.new_page()
            _emit("Autenticando en el portal del SRI...")
            # IMPORTANTE: usamos PORTAL_HOME (entry SSO de Keycloak con
            # client_id=app-sri-claves-angular) en vez de un deep link como
            # RECUPERAR_COMPROBANTES_URL. Si navegamos directo al .jsf, el
            # SRI redirige a auth con un redirect_uri que NO coincide con
            # la sesion de SSO, y apenas pasa el login el portal nos rebota
            # de vuelta al login screen — esto generaba el error "pantalla
            # de autenticacion persistente" infinito.
            #
            # Defensiva: si una corrida anterior dejo cookies vencidas o
            # corruptas en cookies_nc_lookup_<ruc>.json, las borramos antes
            # del _login para que arranque limpio. Para Valor Neto (uso
            # one-shot), el costo de re-loguear es bajo y compensa evitar
            # bounces por estado cookie viejo.
            if cookies_path.exists():
                try:
                    cookies_path.unlink()
                except Exception:
                    pass
            _login(
                context, page, ruc, clave, cookies_path, PORTAL_HOME
            )
            _emit("Sesion del SRI lista.")

            for fecha_str, items in por_fecha.items():
                if cancel_event is not None and cancel_event.is_set():
                    _emit("Busqueda remota cancelada por el usuario.")
                    break

                fecha_dt = datetime.strptime(fecha_str, "%Y-%m-%d")
                fecha_display = fecha_dt.strftime("%d/%m/%Y")
                _emit(
                    f"Buscando {len(items)} Factura(s) emitidas el {fecha_display}..."
                )

                # Navegar al modulo Emitidos
                try:
                    _abrir_modulo_consultas(page, "Emitidos")
                except Exception as exc:
                    _emit(f"No se pudo abrir el modulo de Emitidos: {exc}")
                    continue

                # Aplicar filtros: fecha + Tipo=Factura + Estado=Autorizados
                try:
                    # Esperar a que el overlay de PrimeFaces se cierre
                    # antes de tocar inputs — sino las acciones se interceptan
                    # silenciosamente y la fecha queda con el valor anterior.
                    _esperar_overlay_cerrado(page, timeout_ms=10000)

                    # Setear la fecha via JS (fill+dispatch_event no es
                    # suficiente para PrimeFaces calendar — los eventos de
                    # Playwright tienen isTrusted=false y el widget los ignora).
                    fecha_ok = _set_fecha_emitidos_via_js(page, fecha_display)
                    if not fecha_ok:
                        # Fallback al metodo viejo si el JS no encontro el input.
                        fecha_loc = page.locator(
                            "input#frmPrincipal\\:calendarFechaDesde_input"
                        )
                        if fecha_loc.count():
                            fecha_loc.first.fill("")
                            fecha_loc.first.fill(fecha_display)
                            try:
                                fecha_loc.first.dispatch_event("input")
                                fecha_loc.first.dispatch_event("change")
                                fecha_loc.first.dispatch_event("blur")
                            except Exception:
                                pass

                    _seleccionar_en_select(
                        page,
                        "select#frmPrincipal\\:cmbTipoComprobante",
                        "Factura",
                    )
                    _seleccionar_en_select(
                        page,
                        "select#frmPrincipal\\:cmbEstadoAutorizacion",
                        "Autorizados",
                    )

                    # Click en "Consultar" — antes esperamos overlay cerrado.
                    _esperar_overlay_cerrado(page, timeout_ms=10000)
                    consultar_btn = page.locator(
                        "input[type='submit'][value='Consultar'], "
                        "button:has-text('Consultar')"
                    )
                    if consultar_btn.count():
                        try:
                            consultar_btn.first.click(timeout=8000)
                        except Exception as exc:
                            _emit(
                                f"  [filtros] {fecha_display}: click en Consultar "
                                f"fallo: {exc} (probable overlay no cerrado)."
                            )
                        # Despues del click esperamos a que:
                        #   1) el overlay de "procesando" aparezca y desaparezca
                        #   2) la tabla tenga resultados o el mensaje vacio
                        # Sin esto, leemos un DOM en transicion (filas=0 falsas).
                        _esperar_overlay_cerrado(page, timeout_ms=20000)
                        _esperar_tabla_emitidos_lista(page, timeout_ms=20000)

                    # Despues de aplicar filtros, intentamos subir el rows-per-page
                    # al maximo que ofrezca el SRI. Esto reduce el numero de
                    # paginas a recorrer (75 vs 10 = 7.5x menos clicks).
                    try:
                        rpp_loc = page.locator(
                            "#frmPrincipal\\:tablaCompEmitidos_paginator_top "
                            "select.ui-paginator-rpp-options, "
                            "select[id$='_rppDD']"
                        )
                        if rpp_loc.count():
                            opciones = rpp_loc.first.locator("option")
                            valores = []
                            for i in range(opciones.count()):
                                try:
                                    v = (opciones.nth(i).get_attribute("value") or "").strip()
                                    if v.isdigit():
                                        valores.append(int(v))
                                except Exception:
                                    continue
                            if valores:
                                mejor = max(valores)
                                rpp_loc.first.select_option(value=str(mejor))
                                _esperar_overlay_cerrado(page, timeout_ms=10000)
                                _esperar_tabla_emitidos_lista(page, timeout_ms=10000)
                    except Exception:
                        pass
                except Exception as exc:
                    _emit(f"Error aplicando filtros para {fecha_display}: {exc}")
                    continue

                # Pre-calcular las series normalizadas que estamos buscando.
                # Dos indices para matcheo robusto:
                #   buscados_por_serie: la version "Factura xxx-xxx-xxx" del cell.
                #   buscados_por_secuencial_9: los ultimos 9 digitos (= secuencial
                #     puro) extraidos del clave de acceso. Mas robusto cuando el
                #     cell de tipo+serie tiene markup raro.
                buscados_por_serie: dict[str, dict] = {}
                buscados_por_secuencial_9: dict[str, dict] = {}
                for item in items:
                    sec_orig = item.get("secuencial", "")
                    sec_norm = _normalizar_serie_para_match(sec_orig)
                    if sec_norm:
                        buscados_por_serie[sec_norm] = item
                    # secuencial puro (ultimos 9 digitos) para match por clave de acceso
                    digits = re.sub(r"\D", "", str(sec_orig))
                    if len(digits) >= 9:
                        buscados_por_secuencial_9[digits[-9:]] = item
                total_buscados = len(buscados_por_serie) or len(buscados_por_secuencial_9)
                vistos: set[str] = set()  # keys de items ya encontrados en cualquier pag

                # ===== Diagnostico inicial: confirmar que la tabla quedo
                # filtrada por la fecha objetivo. Si la primera fila de la
                # tabla tiene una fecha autorizacion (col 3) muy distinta a
                # la pedida, el filtro NO se aplico — el resto seria todo
                # inutil. Solo logeamos; no abortamos por si la heuristica
                # falla en algun caso raro.
                try:
                    primera = page.locator(
                        "#frmPrincipal\\:tablaCompEmitidos_data tr"
                    ).first
                    if primera.count():
                        fila_textos = primera.locator("td").all_inner_texts()
                        if len(fila_textos) >= 4:
                            fecha_autorizacion_raw = fila_textos[3].strip()
                            _emit(
                                f"  [diag] Fecha objetivo={fecha_display} | "
                                f"primera fila autorizacion={fecha_autorizacion_raw!r} | "
                                f"primera serie={fila_textos[1].strip()!r}"
                            )
                except Exception:
                    pass

                # Recorrer TODAS las paginas de la tabla. Antes solo
                # procesabamos pagina 1 — facturas en pagina 2+ salian como
                # "no encontrada" aun cuando estaban en el portal. Cortamos
                # cuando matcheamos todos los targets o agotamos el
                # paginador (con tope de seguridad).
                hits = 0
                pagina_actual = 1
                total_filas_revisadas = 0
                MAX_PAGINAS = 200  # tope defensivo; las consultas reales rara vez exceden ~30 paginas
                _emit(
                    f"  [paginacion] {fecha_display}: buscando {total_buscados} factura(s) — "
                    f"objetivos: {', '.join(list(buscados_por_serie.keys())[:5])}"
                    f"{'...' if len(buscados_por_serie) > 5 else ''}"
                )
                while True:
                    if cancel_event is not None and cancel_event.is_set():
                        break

                    try:
                        tabla = page.locator(
                            "#frmPrincipal\\:tablaCompEmitidos_data"
                        )
                        filas = tabla.locator("tr")
                        n_filas = filas.count()
                    except Exception as exc:
                        _emit(
                            f"  [paginacion] No se pudo leer la tabla en {fecha_display} "
                            f"(pag {pagina_actual}): {exc}"
                        )
                        break

                    hits_antes = hits
                    series_vistas_muestra: list[str] = []
                    primer_serie_pagina = ""
                    for idx in range(n_filas):
                        if cancel_event is not None and cancel_event.is_set():
                            break
                        fila = filas.nth(idx)
                        celdas = fila.locator("td")
                        try:
                            textos_celdas = celdas.all_inner_texts()
                        except Exception:
                            textos_celdas = []
                        if len(textos_celdas) < 8:
                            continue
                        tipo_serie_text = textos_celdas[1].strip()
                        clave_text = textos_celdas[2].strip()
                        importe_text = textos_celdas[7].strip()
                        total_filas_revisadas += 1
                        if idx == 0:
                            primer_serie_pagina = tipo_serie_text
                        if len(series_vistas_muestra) < 3:
                            series_vistas_muestra.append(tipo_serie_text)

                        # Match #1: por tipo+serie normalizada del cell.
                        # Acepta cualquier prefijo (no solo "Factura") — solo
                        # importan los digitos serie post normalizacion.
                        item_matched = None
                        serie_norm = _normalizar_serie_para_match(
                            tipo_serie_text.replace("Factura", "").replace("factura", "")
                        )
                        if serie_norm in buscados_por_serie:
                            item_matched = buscados_por_serie[serie_norm]
                        # Match #2 (fallback): extraer los digitos del secuencial
                        # de la clave de acceso (posiciones 30-38 del clave de
                        # 49 digitos = los 9 digitos del secuencial).
                        if item_matched is None:
                            clave_digits = re.sub(r"\D", "", clave_text)
                            if len(clave_digits) == 49:
                                sec_de_clave = clave_digits[30:39]
                                if sec_de_clave in buscados_por_secuencial_9:
                                    item_matched = buscados_por_secuencial_9[sec_de_clave]
                        if item_matched is not None:
                            key = _normalizar_secuencial(item_matched.get("secuencial", ""))
                            if key in vistos:
                                continue  # ya lo encontramos en una pag anterior
                            vistos.add(key)
                            encontradas[key] = {
                                "importe_total": _safe_float(importe_text),
                                "clave_acceso": clave_text,
                                "tipo_serie_raw": tipo_serie_text,
                            }
                            hits += 1

                    _emit(
                        f"  [pag {pagina_actual}] {fecha_display}: filas={n_filas} | "
                        f"matches en esta pag={hits - hits_antes} (total {hits}/{total_buscados}) | "
                        f"primera fila='{primer_serie_pagina[:50]}'"
                    )

                    # Si ya encontramos todos los targets, no tiene sentido
                    # seguir paginando.
                    if hits >= total_buscados:
                        _emit(
                            f"  [paginacion] {fecha_display}: matcheamos los "
                            f"{total_buscados} targets, corto paginacion."
                        )
                        break

                    # Tope defensivo
                    if pagina_actual >= MAX_PAGINAS:
                        _emit(
                            f"  [paginacion] {fecha_display}: se alcanzo tope de "
                            f"{MAX_PAGINAS} paginas — corto la busqueda."
                        )
                        break

                    # ====== Click en "siguiente pagina" con verificacion REAL =======
                    # Estrategia:
                    #  1) capturar el texto de la primera fila ANTES del click
                    #     (col 2 = clave de acceso, unica por factura).
                    #  2) probar selectores en orden de especificidad. Si el
                    #     elemento existe pero esta disabled o no existe el
                    #     siguiente, salimos: estamos en la ultima pag.
                    #  3) hacer click.
                    #  4) wait_for_function hasta que el primer-row clave cambie
                    #     (o la tabla quede vacia). networkidle solo no alcanza
                    #     porque PrimeFaces hace partial AJAX y el wait puede
                    #     completarse antes del re-render.
                    #  5) Si tras N segundos el contenido no cambia → el click
                    #     no avanzo realmente. Logeamos y abortamos.
                    clave_primera_fila_prev = ""
                    try:
                        if n_filas > 0:
                            primera_celdas = filas.first.locator("td")
                            if primera_celdas.count() >= 3:
                                clave_primera_fila_prev = (
                                    primera_celdas.nth(2).inner_text(timeout=1500)
                                    or ""
                                ).strip()
                    except Exception:
                        pass

                    # Localizar boton siguiente — probamos varias variantes.
                    boton_clickeable = None
                    for next_sel in (
                        "a.ui-paginator-next:not(.ui-state-disabled)",
                        "span.ui-paginator-next:not(.ui-state-disabled)",
                        ".ui-paginator-next:not(.ui-state-disabled)",
                        "#frmPrincipal\\:tablaCompEmitidos_paginator_bottom "
                        "a.ui-paginator-next:not(.ui-state-disabled)",
                    ):
                        try:
                            loc = page.locator(next_sel)
                            if loc.count():
                                boton_clickeable = loc.first
                                break
                        except Exception:
                            continue

                    if boton_clickeable is None:
                        _emit(
                            f"  [paginacion] {fecha_display}: no hay boton 'siguiente' "
                            f"habilitado en pag {pagina_actual}. Fin natural de paginacion."
                        )
                        break

                    try:
                        boton_clickeable.scroll_into_view_if_needed(timeout=2000)
                    except Exception:
                        pass
                    # CRITICO: esperar a que el overlay #dlgpopStatusPrime_modal
                    # de PrimeFaces se cierre antes de clickear. Si esta arriba,
                    # Playwright reintenta 13+ veces hasta el timeout y aborta.
                    overlay_libre = _esperar_overlay_cerrado(page, timeout_ms=15000)
                    if not overlay_libre:
                        _emit(
                            f"  [paginacion] {fecha_display}: overlay de PrimeFaces "
                            f"sigue activo tras 15s en pag {pagina_actual}. "
                            f"Intento click con force=True."
                        )
                    try:
                        boton_clickeable.click(timeout=6000, force=not overlay_libre)
                    except Exception as exc:
                        _emit(
                            f"  [paginacion] {fecha_display}: click en 'siguiente' "
                            f"fallo en pag {pagina_actual}: {exc}. Abandono paginacion."
                        )
                        break

                    # Esperar a que la primera fila REALMENTE cambie su clave.
                    # Si tras 12s sigue igual, el click no surtio efecto.
                    cambio_ok = False
                    try:
                        page.wait_for_function(
                            """(prev) => {
                                const t = document.getElementById('frmPrincipal:tablaCompEmitidos_data');
                                if (!t) return false;
                                const filaUno = t.querySelector('tr');
                                if (!filaUno) return true;
                                const celdas = filaUno.querySelectorAll('td');
                                if (celdas.length < 3) return true;
                                const clave = (celdas[2].innerText || '').trim();
                                if (!clave) return false;
                                return clave !== prev;
                            }""",
                            arg=clave_primera_fila_prev,
                            timeout=12000,
                        )
                        cambio_ok = True
                    except Exception:
                        cambio_ok = False

                    if not cambio_ok:
                        _emit(
                            f"  [paginacion] {fecha_display}: click en 'siguiente' "
                            f"NO avanzo la pagina (primera fila sigue siendo "
                            f"{clave_primera_fila_prev[:20]}...). Abandono paginacion. "
                            f"Si esto pasa siempre, el selector de 'siguiente' "
                            f"esta mal o el SRI cambio el markup."
                        )
                        break

                    # Esperar a que el overlay de "procesando" termine antes
                    # de leer la nueva pagina, sino podemos leer DOM en transicion.
                    _esperar_overlay_cerrado(page, timeout_ms=10000)
                    try:
                        page.wait_for_load_state("networkidle", timeout=3000)
                    except Exception:
                        pass
                    time.sleep(0.2)
                    pagina_actual += 1

                _emit(
                    f"Fecha {fecha_display}: {hits}/{len(items)} Factura(s) "
                    f"encontradas en la tabla del SRI "
                    f"(recorridas {pagina_actual} pag, {total_filas_revisadas} filas revisadas)."
                )
                # Log diagnostico: si quedaron sin encontrar, listamos los
                # secuenciales para que el usuario pueda chequear manualmente.
                if hits < len(items):
                    no_encontrados = []
                    for it in items:
                        k = _normalizar_secuencial(it.get("secuencial", ""))
                        if k not in vistos:
                            no_encontrados.append(it.get("secuencial", "?"))
                    if no_encontrados:
                        muestra = ", ".join(no_encontrados[:8])
                        suf = "..." if len(no_encontrados) > 8 else ""
                        _emit(
                            f"  No matcheadas en {fecha_display}: {muestra}{suf} "
                            f"(revisa el log: si las paginas avanzaron y filas_revisadas "
                            f"es alto pero hits es bajo, hay un bug de matching. "
                            f"Si filas_revisadas es bajo, hay un bug de paginacion.)"
                        )

            # Guardar cookies para futuras consultas
            try:
                cookies_path.write_text(
                    json.dumps(context.cookies()), encoding="utf-8"
                )
            except Exception:
                pass
        finally:
            try:
                context.close()
            except Exception:
                pass
            try:
                if browser is not None:
                    browser.close()
            except Exception:
                pass

    return encontradas


# =============================================================================
# Orquestador principal
# =============================================================================


def _construir_fila(
    nc_data: dict,
    factura_data: Optional[dict],
    estado: str,
    observacion: str,
) -> dict:
    """Construye una fila del Excel a partir de los datos de NC + Factura (opcional)."""
    valor_nc = nc_data.get("importe_total")
    valor_factura = factura_data.get("importe_total") if factura_data else None
    valor_neto = None
    if valor_factura is not None and valor_nc is not None:
        valor_neto = round(valor_factura - valor_nc, 2)

    return {
        "RUC": nc_data.get("ruc", ""),
        "Fecha nota de credito": nc_data.get("fecha_emision", ""),
        "Serie nota de credito": nc_data.get("secuencial", ""),
        "Clave acceso nota de credito": nc_data.get("clave_acceso", ""),
        "Valor total nota de credito": valor_nc if valor_nc is not None else "",
        "Factura modificada": _normalizar_secuencial(
            nc_data.get("num_doc_modificado", "")
        ),
        "Fecha factura modificada": nc_data.get("fecha_emision_doc_sustento", ""),
        "Clave acceso factura": factura_data.get("clave_acceso", "") if factura_data else "",
        "Valor total factura": valor_factura if valor_factura is not None else "",
        "Valor neto": valor_neto if valor_neto is not None else "",
        "Estado": estado,
        "Observacion": observacion,
    }


def generar_reporte_valor_neto(
    *,
    carpeta_nc: str | Path,
    ruc: str,
    clave: str,
    salida_excel: str | Path,
    cancel_event: Optional[threading.Event] = None,
    progress: Optional[Callable[[str], None]] = None,
) -> dict:
    """Punto de entrada unico del modulo. Lee NC de `carpeta_nc`, cruza contra
    Facturas (local + remoto si hace falta), genera Excel en `salida_excel`.

    Retorna resumen: {
        ok: bool,
        total_nc: int,
        encontradas_local: int,
        encontradas_remoto: int,
        no_encontradas: int,
        errores: int,
        excel_path: str,
        message: str,
    }
    """
    def _emit(msg: str) -> None:
        if progress:
            try:
                progress(msg)
            except Exception:
                pass
        logger.info(msg)

    carpeta_nc_path = Path(carpeta_nc).expanduser()
    if not carpeta_nc_path.is_dir():
        return {
            "ok": False,
            "total_nc": 0,
            "encontradas_local": 0,
            "encontradas_remoto": 0,
            "no_encontradas": 0,
            "errores": 0,
            "excel_path": "",
            "message": f"La carpeta no existe: {carpeta_nc_path}",
        }

    _emit(f"Explorando carpeta de Notas de Credito: {carpeta_nc_path}")
    notas = _listar_notas_credito_en_carpeta(carpeta_nc_path)
    if not notas:
        return {
            "ok": False,
            "total_nc": 0,
            "encontradas_local": 0,
            "encontradas_remoto": 0,
            "no_encontradas": 0,
            "errores": 0,
            "excel_path": "",
            "message": "No se encontraron Notas de Credito en la carpeta indicada.",
        }
    _emit(f"Notas de Credito detectadas: {len(notas)}")

    # Inferir base [RUC] del arbol de descargas (usa la primera NC como pista).
    base_descargas = _inferir_base_descargas(notas[0][0])
    if base_descargas is None:
        # Fallback: la propia carpeta seleccionada es la base.
        base_descargas = carpeta_nc_path
    _emit(f"Base de descargas inferida: {base_descargas}")

    rows: list[dict] = []
    pendientes_remoto: list[dict] = []
    errores = 0
    encontradas_local = 0

    for nc_path, source in notas:
        if cancel_event is not None and cancel_event.is_set():
            _emit("Proceso cancelado por el usuario.")
            break

        nc_data = _extraer_datos_nc(nc_path, source)
        if nc_data["_err"]:
            errores += 1
            rows.append(
                _construir_fila(
                    nc_data, None, "Error", f"{nc_data['_err']} (archivo: {nc_path.name})"
                )
            )
            continue

        # Filtrar solo NC que modifican Facturas (cod_doc_modificado == "01")
        cod = (nc_data.get("cod_doc_modificado") or "").strip().lstrip("0") or "0"
        if cod != "1":
            rows.append(
                _construir_fila(
                    nc_data,
                    None,
                    "Omitido",
                    f"Esta NC modifica un comprobante tipo '{cod}', no una Factura (01).",
                )
            )
            continue

        secuencial_factura = _normalizar_secuencial(nc_data.get("num_doc_modificado", ""))
        fecha_factura = _parse_fecha_es(nc_data.get("fecha_emision_doc_sustento", ""))

        if not secuencial_factura:
            rows.append(
                _construir_fila(
                    nc_data,
                    None,
                    "Error",
                    "La NC no contiene 'Numero Documento Modificado' valido.",
                )
            )
            errores += 1
            continue

        # Busqueda LOCAL
        encontrado_local = _buscar_factura_local(
            base_descargas, secuencial_factura, fecha_factura
        )
        if encontrado_local is not None:
            factura_path, factura_source = encontrado_local
            factura_data = _extraer_datos_factura(factura_path, factura_source)
            if factura_data["_err"]:
                rows.append(
                    _construir_fila(
                        nc_data,
                        None,
                        "Error",
                        f"Factura encontrada localmente pero no se pudo leer: "
                        f"{factura_data['_err']}",
                    )
                )
                errores += 1
                continue
            rows.append(
                _construir_fila(
                    nc_data,
                    factura_data,
                    "OK (local)",
                    f"Factura encontrada localmente en {factura_path.name}.",
                )
            )
            encontradas_local += 1
        else:
            # Marcar para busqueda remota
            pendientes_remoto.append(
                {
                    "secuencial": secuencial_factura,
                    "fecha": fecha_factura,
                    "nc_data": nc_data,
                }
            )

    # === Busqueda REMOTA para las que no se encontraron localmente ===
    encontradas_remoto = 0
    no_encontradas = 0
    encontradas_dict: dict[str, dict] = {}

    if pendientes_remoto and (cancel_event is None or not cancel_event.is_set()):
        if not ruc or not clave:
            # Sin credenciales no podemos consultar el SRI: marcar todas como
            # "no encontradas localmente — falta credencial para consulta remota".
            _emit(
                "No se proporcionaron credenciales — se omite busqueda remota. "
                "Las Facturas faltantes quedaran marcadas como 'no encontradas'."
            )
        else:
            try:
                encontradas_dict = _buscar_facturas_remoto(
                    ruc,
                    clave,
                    pendientes_remoto,
                    cancel_event=cancel_event,
                    progress=progress,
                )
            except Exception as exc:
                _emit(f"Busqueda remota fallo: {exc}")
                encontradas_dict = {}

    # Construir filas de las pendientes
    for item in pendientes_remoto:
        nc_data = item["nc_data"]
        key = _normalizar_secuencial(item["secuencial"])
        info_remoto = encontradas_dict.get(key)
        if info_remoto:
            factura_data = {
                "importe_total": info_remoto.get("importe_total"),
                "clave_acceso": info_remoto.get("clave_acceso", ""),
            }
            rows.append(
                _construir_fila(
                    nc_data,
                    factura_data,
                    "OK (remoto)",
                    f"Factura encontrada en el portal del SRI "
                    f"(tabla emitidos, columna tipo y serie).",
                )
            )
            encontradas_remoto += 1
        else:
            rows.append(
                _construir_fila(
                    nc_data,
                    None,
                    "Factura no encontrada",
                    f"No se encontro la Factura {item['secuencial']} ni localmente "
                    f"ni en el portal del SRI.",
                )
            )
            no_encontradas += 1

    # === Generar Excel ===
    salida_path = Path(salida_excel).expanduser()
    salida_path.parent.mkdir(parents=True, exist_ok=True)
    _escribir_excel(rows, salida_path)
    _emit(f"Excel generado: {salida_path}")

    return {
        "ok": True,
        "total_nc": len(rows),
        "encontradas_local": encontradas_local,
        "encontradas_remoto": encontradas_remoto,
        "no_encontradas": no_encontradas,
        "errores": errores,
        "excel_path": str(salida_path),
        "message": (
            f"Reporte generado con {len(rows)} fila(s). "
            f"Local: {encontradas_local}, Remoto: {encontradas_remoto}, "
            f"No encontradas: {no_encontradas}, Errores: {errores}."
        ),
    }


def _escribir_excel(rows: list[dict], path: Path) -> None:
    """Genera el Excel con formato basico (header + monetario + text-force)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Valor Neto NC vs Facturas"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Header
    for col_idx, nombre in enumerate(COLUMNAS_REPORTE, start=1):
        cell = ws.cell(row=1, column=col_idx, value=nombre)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Anchos sugeridos
    anchos = {
        "RUC": 15,
        "Fecha nota de credito": 14,
        "Serie nota de credito": 18,
        "Clave acceso nota de credito": 52,
        "Valor total nota de credito": 16,
        "Factura modificada": 18,
        "Fecha factura modificada": 14,
        "Clave acceso factura": 52,
        "Valor total factura": 16,
        "Valor neto": 14,
        "Estado": 22,
        "Observacion": 60,
    }
    for col_idx, nombre in enumerate(COLUMNAS_REPORTE, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = anchos.get(nombre, 16)

    # Filas de datos
    for row_idx, fila in enumerate(rows, start=2):
        for col_idx, nombre in enumerate(COLUMNAS_REPORTE, start=1):
            valor = fila.get(nombre, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            if nombre in COLUMNAS_TEXT_FORCE and valor != "":
                cell.number_format = "@"
            if nombre in COLUMNAS_NUMERICAS and valor != "":
                cell.number_format = "#,##0.00"
            cell.alignment = Alignment(vertical="center", wrap_text=(nombre == "Observacion"))

    # Congelar header
    ws.freeze_panes = "A2"
    wb.save(str(path))
