from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


TEXT_COLUMNS = [
    "rucEmisor",
    "numeroComprobante",
    "establecimiento",
    "puntoEmision",
    "secuencial",
    "claveAcceso",
]


DEFAULT_COLUMNS = [
    "archivoPdf",
    "tipoDocumento",
    "rucEmisor",
    "razonSocialEmisor",
    "nombreComercial",
    "direccionMatrizEmisor",
    "direccionSucursalEmisor",
    "contribuyenteEspecial",
    "agenteRetencion",
    "obligadoContabilidad",
    "tipoContribuyenteRIMPE",
    "numeroComprobante",
    "establecimiento",
    "puntoEmision",
    "secuencial",
    "claveAcceso",
    "fechaEmision",
    "razonSocialComprador",
    "identificacionComprador",
    "valorTotal",
]


def write_rows_to_excel(rows: list[dict], excel_path: str | Path, columns: list[str] | None = None) -> Path:
    excel_path = Path(excel_path)
    columns = columns or DEFAULT_COLUMNS
    df = pd.DataFrame(rows or [])
    for column in columns:
        if column not in df.columns:
            df[column] = ""
    df = df[columns].copy()

    for column in TEXT_COLUMNS:
        if column in df.columns:
            df[column] = df[column].map(lambda value: "" if value is None else str(value).strip())

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Detalle PDFs", index=False)
        worksheet = writer.sheets["Detalle PDFs"]

        column_index = {cell.value: idx + 1 for idx, cell in enumerate(worksheet[1])}
        for text_column in TEXT_COLUMNS:
            idx = column_index.get(text_column)
            if not idx:
                continue
            col_letter = get_column_letter(idx)
            for row_idx in range(2, worksheet.max_row + 1):
                cell = worksheet[f"{col_letter}{row_idx}"]
                cell.number_format = "@"
                if cell.value is not None:
                    cell.value = str(cell.value)
                cell.alignment = Alignment(horizontal="left")
            worksheet.column_dimensions[col_letter].width = max(16, worksheet.column_dimensions[col_letter].width or 0)

    return excel_path
