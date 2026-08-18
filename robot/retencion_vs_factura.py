# =====================================================
# PARTE 3: COMPROBANTES DE RETENCION EMITIDOS vs FACTURAS RECIBIDAS
# =====================================================
# Cruza cada comprobante de retencion emitido contra la factura recibida que
# le sirve de sustento.
#
# POR QUE NO SE ARMA LA CLAVE DE ACCESO
# -------------------------------------
# En Notas de Credito vs Facturas (robot/nc_vs_factura.py) se calcula la clave
# de la factura porque la NC modifica una factura del MISMO contribuyente: el
# codigo numerico (8 digitos libres) se puede copiar de la propia NC.
#
# Aca no sirve: la factura la emitio un TERCERO (el proveedor), asi que su
# codigo numerico es inadivinable. Y `numAutDocSustento` -que traeria la clave
# servida- vive solo en el XML de la retencion, que para meses viejos no se
# puede obtener: el portal solo ofrece PDF y el WS publico de autorizacion
# devuelve `numeroComprobantes=0` pasados ~30 dias (el mismo umbral que ya
# codifica `_debe_omitir_soap_xml` en robot/workflows.py).
#
# Por eso el cruce se hace por RUC del proveedor + numero de factura, que el
# RIDE de la retencion si trae siempre. La `Fecha Emision` del sustento indica
# ademas en que mes buscar la factura, sin necesidad de barrer un rango.
#
# DE DONDE SALEN LOS DATOS DE CADA LADO
# -------------------------------------
# Los dos flujos de descarga usan mecanismos distintos, y eso define que se
# puede pedir de cada lado:
#
#   Emitidos  -> XML por el WS publico (workflows.py: _descargar_xml_emitido_
#                por_clave). Tiene ventana de ~30 dias, asi que para meses
#                viejos solo queda el PDF. De ahi el extractor de RIDE.
#   Recibidos -> XML por el enlace de la propia tabla del portal
#                (workflows.py: "a[id$=':lnkXml']"). NO pasa por el WS y no
#                tiene ventana: los XML de facturas recibidas estan siempre.
#
# El listado de Recibidos trae, ademas de la clave de acceso, el SUBTOTAL, el
# IVA y el IMPORTE TOTAL de cada factura en columnas separadas. Con eso alcanza
# para todo el reporte -incluida la verificacion de sobre que base se calculo la
# retencion- SIN descargar el XML de cada factura. El XML se sigue aceptando y
# gana cuando existe, pero no es necesario.
# =====================================================

from __future__ import annotations

import logging
import re
import threading
import unicodedata
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path
from typing import Callable, Iterable, Optional

logger = logging.getLogger(__name__)

# Codigos de documento sustento del SRI. Solo "01" cruza contra facturas.
COD_FACTURA = "01"
DOC_SUSTENTO_LABELS = {
    "01": "FACTURA",
    "02": "NOTA DE VENTA",
    "03": "LIQUIDACION DE COMPRA",
    "04": "NOTA DE CREDITO",
    "05": "NOTA DE DEBITO",
    "06": "GUIA DE REMISION",
    "07": "COMPROBANTE DE RETENCION",
    "12": "DOCUMENTOS IFIS",
    "15": "COMPROBANTE DE VENTA POR REEMBOLSO",
    "16": "DOCUMENTOS AUTORIZADOS EN IMPORTACIONES",
    "19": "COMPROBANTE DE PAGO DE CUOTAS O APORTES",
    "41": "COMPROBANTE DE VENTA POR REEMBOLSO",
    "47": "NOTA DE CREDITO POR REEMBOLSO",
    "48": "NOTA DE DEBITO POR REEMBOLSO",
}

_MESES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
    7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre",
    12: "Diciembre",
}

# Los dos sentidos del cruce. Cambian de donde salen las facturas, no como se
# las encuentra: quien emitio la factura es en ambos casos el sujeto retenido.
#
#   emitidas  -- el contribuyente retuvo a su proveedor.
#                La factura se la emitio el proveedor -> esta en Recibidos.
#   recibidas -- a el le retuvieron sobre una venta.
#                La factura la emitio el mismo         -> esta en Emitidos.
SENTIDO_EMITIDAS = "emitidas"
SENTIDO_RECIBIDAS = "recibidas"


def _normalizar_sentido(sentido: object) -> str:
    texto = _norm(sentido)
    return SENTIDO_RECIBIDAS if texto.startswith("recib") else SENTIDO_EMITIDAS


def _origen_facturas(sentido: str) -> str:
    """Modulo del portal donde vive la factura sustento de cada sentido."""
    return "Emitidos" if _normalizar_sentido(sentido) == SENTIDO_RECIBIDAS else "Recibidos"


# =============================================================================
# Helpers de texto y numeros
# =============================================================================


def _norm(texto: object) -> str:
    """Minusculas, sin acentos y con espacios colapsados, para comparar."""
    s = str(texto or "")
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s).strip().lower()


def _a_float(valor: object) -> Optional[float]:
    """Convierte '1.234,56' o '1234.56' a float. None si no es numero."""
    s = str(valor or "").strip()
    if not s:
        return None
    s = re.sub(r"[^\d,.\-]", "", s)
    if not s:
        return None
    if "," in s and "." in s:
        # El separador decimal es el ultimo que aparece.
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _normalizar_numero_doc(valor: object) -> str:
    """Lleva un numero de comprobante a 'EEE-PPP-SSSSSSSSS'.

    Acepta '001001000004748', '001-001-000004748' o el numero partido en
    pedazos por el wrap de la celda del PDF ('0010010000047' + '48').
    """
    digitos = re.sub(r"\D", "", str(valor or ""))
    if not digitos:
        return ""
    if len(digitos) < 15:
        digitos = digitos.zfill(15)
    elif len(digitos) > 15:
        digitos = digitos[-15:]
    return f"{digitos[:3]}-{digitos[3:6]}-{digitos[6:]}"


def _numero_desde_partes(estab: object, pto: object, secuencial: object) -> str:
    """Arma 'EEE-PPP-SSSSSSSSS' a partir de las tres partes por separado.

    No sirve concatenar y normalizar: los reportes del portal traen las partes
    SIN ceros a la izquierda (estab=1, pto=901, sec=1509), asi que hay que
    rellenar cada una a su largo antes de unirlas.
    """
    e = re.sub(r"\D", "", str(estab or "")).zfill(3)[-3:]
    p = re.sub(r"\D", "", str(pto or "")).zfill(3)[-3:]
    s = re.sub(r"\D", "", str(secuencial or "")).zfill(9)[-9:]
    if not any((e.strip("0"), p.strip("0"), s.strip("0"))):
        return ""
    return f"{e}-{p}-{s}"


def _canonizar_identificacion(valor: object) -> str:
    """Lleva cedula y RUC de una misma persona natural a la misma forma.

    Un RIDE puede identificar al sujeto retenido con la CEDULA (10 digitos)
    mientras que en el listado esa persona factura con su RUC (la cedula mas
    '001'). Comparando los digitos crudos, `1716891658` y `1716891658001`
    nunca coinciden y la factura queda como no encontrada.

    Solo se recorta cuando es RUC de persona natural: el tercer digito indica
    el tipo de contribuyente, y 6 (publico) y 9 (sociedad) no se tocan porque
    ahi el '001' es parte del establecimiento, no un sufijo de la cedula.
    """
    digitos = re.sub(r"\D", "", str(valor or ""))
    if len(digitos) == 13 and digitos.endswith("001") and digitos[2] not in "69":
        return digitos[:10]
    return digitos


def _clave_match(ruc: object, numero: object) -> str:
    """Clave de cruce entre retencion y factura: identificacion + numero."""
    ruc_s = _canonizar_identificacion(ruc)
    num_s = re.sub(r"\D", "", str(numero or ""))
    if not ruc_s or not num_s:
        return ""
    return f"{ruc_s}|{num_s.zfill(15)[-15:]}"


def _parse_fecha(valor: object) -> Optional[datetime]:
    s = str(valor or "").strip()
    if not s:
        return None
    s = s.split(" ")[0]
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


# =============================================================================
# Extraccion posicional del RIDE (PDF) de una retencion
# =============================================================================
# El texto plano del RIDE sale desordenado: leerlo en orden de lectura mezcla
# la base con el porcentaje y parte el numero de factura en dos. Por eso todo
# se resuelve por coordenadas.
#
# Las columnas NO se hardcodean: se derivan de la posicion de las cabeceras del
# propio PDF, asi que el extractor se adapta a cada emisor.

# Cada columna con sus posibles textos de cabecera, en orden de aparicion.
_ANCLAS_TABLA = [
    ("comprobante", ("comprobante",)),
    ("numero", ("numero",)),
    ("fecha_emision", ("fecha",)),
    ("ejercicio", ("ejercicio",)),
    ("base", ("base",)),
    ("impuesto", ("impuesto",)),
    ("porcentaje", ("porcentaje",)),
    ("valor", ("valorretenido", "valor")),
]


def _palabras_pdf(pdf_path: Path) -> list[tuple[float, float, float, float, str]]:
    """Devuelve [(x0, y0, x1, y1, texto)] de la primera pagina.

    Usa pdfplumber, que es lo que ya usan pdf_extraction.py y downloader.py, y
    cae a PyMuPDF si no estuviera. Las dos librerias figuran en requirements
    pero PyMuPDF no siempre termina instalada en el entorno donde corre la app
    -- y cuando falta, sin este fallback el reporte sale vacio con un unico
    aviso de 'No module named fitz' por archivo.
    """
    errores = []

    try:
        import pdfplumber

        with pdfplumber.open(str(pdf_path)) as doc:
            if not doc.pages:
                return []
            palabras = doc.pages[0].extract_words()
        # 'top'/'bottom' son la misma orientacion que el y0/y1 de PyMuPDF.
        return [
            (float(w["x0"]), float(w["top"]), float(w["x1"]), float(w["bottom"]), w["text"])
            for w in palabras
        ]
    except ImportError as err:
        errores.append(f"pdfplumber: {err}")
    except Exception as err:
        errores.append(f"pdfplumber: {err}")

    try:
        import fitz

        with fitz.open(pdf_path) as doc:
            if not doc.page_count:
                return []
            palabras = doc[0].get_text("words")
        return [(w[0], w[1], w[2], w[3], w[4]) for w in palabras]
    except Exception as err:
        errores.append(f"PyMuPDF: {err}")

    raise RuntimeError(
        "No se pudo leer el PDF con ninguna libreria disponible. " + " | ".join(errores)
    )


def _centro_x(palabra) -> float:
    return (palabra[0] + palabra[2]) / 2.0


def _agrupar_lineas(palabras, tol_y: float = 3.0) -> list[list]:
    """Agrupa palabras en lineas visuales, ordenadas de arriba a abajo.

    Se agrupa por cercania y no redondeando la coordenada: en el RIDE la
    etiqueta y su valor pueden diferir un par de puntos en y (437.7 vs 436.0) y
    un redondeo los parte en dos lineas distintas.
    """
    lineas: list[list] = []
    actual: list = []
    for palabra in sorted(palabras, key=lambda w: w[1]):
        if actual and palabra[1] - actual[0][1] > tol_y:
            lineas.append(sorted(actual, key=lambda w: w[0]))
            actual = []
        actual.append(palabra)
    if actual:
        lineas.append(sorted(actual, key=lambda w: w[0]))
    return lineas


def _valor_junto_a(
    palabras, etiqueta: str, patron: Optional[str] = None, tol_y: float = 3.0
) -> str:
    """Devuelve el texto que esta a la derecha de `etiqueta`, en su misma linea.

    En el RIDE las etiquetas van en una columna y los valores en otra, alineados
    por linea ('Identificacion' a la izquierda, el RUC a la derecha).

    La etiqueta puede ser de varias palabras ('Razon Social / Nombres y
    Apellidos:'), asi que se compara contra la linea completa y no palabra por
    palabra. `patron` filtra el valor cuando la etiqueta se repite en el
    documento: 'Fecha' aparece tanto en los datos del sujeto retenido como en
    la cabecera de la tabla de sustento.
    """
    objetivo = _norm(etiqueta)

    for grupo in _agrupar_lineas(palabras, tol_y):
        acumulado = ""
        for i, palabra in enumerate(grupo):
            acumulado = _norm(f"{acumulado} {palabra[4]}")
            if acumulado != objetivo:
                continue
            valor = " ".join(w[4] for w in grupo[i + 1:]).strip()
            if not valor:
                break
            if patron and not re.search(patron, valor):
                break
            if patron:
                encontrado = re.search(patron, valor)
                return encontrado.group(0) if encontrado else valor
            return valor
    return ""


def _razon_social_emisor(palabras) -> str:
    """Razon social de quien EMITE la retencion, del recuadro superior izquierdo.

    No viene como etiqueta-valor sino suelta dentro del recuadro, asi que se
    toma la primera linea de texto de esa caja: en los RIDE es siempre el nombre
    del emisor, arriba del nombre comercial y de 'Direccion Matriz'.

    Importa sobre todo en retenciones RECIBIDAS, donde el emisor es el cliente
    que retuvo (en las emitidas es el propio contribuyente).
    """
    ancla = [w for w in palabras if _norm(w[4]).startswith("direccion")]
    if not ancla:
        return ""
    y_tope = min(w[1] for w in ancla)

    # El RIDE tiene dos cajas lado a lado: la del emisor a la izquierda y la de
    # autorizacion a la derecha. El borde es donde arranca 'R.U.C.:'. Sin este
    # corte se cuela el titulo 'COMPROBANTE DE RETENCION', que esta a la misma
    # altura pero en la caja de al lado.
    ruc_x = [w[0] for w in palabras if _norm(w[4]).startswith("r.u.c")]
    x_limite = min(ruc_x) if ruc_x else max((w[2] for w in palabras), default=600) / 2

    candidatas = [w for w in palabras if w[1] < y_tope and w[0] < x_limite]
    for linea in _agrupar_lineas(candidatas):
        texto = " ".join(w[4] for w in linea).strip()
        plano = _norm(texto)
        # Saltar el encabezado del documento y las etiquetas del recuadro.
        if not texto or len(plano) < 4:
            continue
        if any(
            t in plano
            for t in (
                "comprobante de retencion", "r.u.c", "ruc:", "numero de autorizacion",
                "clave de acceso", "ambiente", "emision", "fecha y hora",
                "no tiene logo", "produccion",
            )
        ):
            continue
        if re.fullmatch(r"[\d\s./:-]+", texto):  # fechas, claves, numeros sueltos
            continue
        return texto
    return ""


def _limites_columnas(palabras) -> Optional[dict]:
    """Ubica la cabecera de la tabla de sustento y devuelve los cortes en x.

    Devuelve {'orden': [...], 'cortes': {col: (x_min, x_max)}, 'y': y_cabecera}
    o None si la tabla no aparece.
    """
    # La cabecera se ancla en 'Comprobante' con 'Numero' a su derecha y cerca.
    y_cabecera = None
    for x0, y0, x1, y1, texto in palabras:
        if _norm(texto) != "comprobante":
            continue
        hay_numero = any(
            _norm(t) == "numero" and abs(py0 - y0) <= 12 and px0 > x1
            for px0, py0, px1, py1, t in palabras
        )
        if hay_numero:
            y_cabecera = y0
            break
    if y_cabecera is None:
        return None

    banda = [w for w in palabras if abs(w[1] - y_cabecera) <= 12]

    centros: dict[str, float] = {}
    for col, variantes in _ANCLAS_TABLA:
        for variante in variantes:
            candidatos = [w for w in banda if _norm(w[4]) == variante]
            if candidatos:
                # El de mas a la izquierda: 'Retencion' aparece dos veces.
                elegido = min(candidatos, key=_centro_x)
                centros[col] = _centro_x(elegido)
                break
    faltan = [c for c, _ in _ANCLAS_TABLA if c not in centros]
    if faltan:
        logger.warning(f"Cabeceras no ubicadas en el RIDE: {faltan}")
        return None

    orden = [c for c, _ in _ANCLAS_TABLA]
    orden.sort(key=lambda c: centros[c])
    cortes: dict[str, tuple[float, float]] = {}
    for i, col in enumerate(orden):
        izq = float("-inf") if i == 0 else (centros[orden[i - 1]] + centros[col]) / 2
        der = float("inf") if i == len(orden) - 1 else (centros[col] + centros[orden[i + 1]]) / 2
        cortes[col] = (izq, der)
    return {"orden": orden, "cortes": cortes, "y": y_cabecera}


def _columna_de(palabra, cortes: dict) -> Optional[str]:
    cx = _centro_x(palabra)
    for col, (izq, der) in cortes.items():
        if izq <= cx < der:
            return col
    return None


def _fin_de_tabla(palabras, y_cabecera: float) -> float:
    """La tabla termina donde empieza 'Informacion Adicional'."""
    ys = [
        w[1]
        for w in palabras
        if w[1] > y_cabecera and _norm(w[4]) in {"informacion", "adicional"}
    ]
    return min(ys) if ys else float("inf")


def _extraer_documentos_pdf(palabras) -> list[dict]:
    """Arma la lista de documentos sustento con sus lineas de retencion."""
    layout = _limites_columnas(palabras)
    if not layout:
        return []
    cortes = layout["cortes"]
    y_ini = layout["y"] + 8
    y_fin = _fin_de_tabla(palabras, layout["y"])

    celdas: dict[str, list[tuple[float, float, str]]] = {c: [] for c in cortes}
    for palabra in palabras:
        x0, y0, x1, y1, texto = palabra
        if not (y_ini <= y0 < y_fin) or not texto.strip():
            continue
        col = _columna_de(palabra, cortes)
        if col:
            celdas[col].append((y0, x0, texto))

    # Cada documento se ancla en su FECHA DE EMISION, que es un unico token por
    # fila. No sirve anclar en el nombre del comprobante: 'NOTA DE VENTA' son
    # tres palabras y 'DOCUMENTOS IFIS' viene partido en dos lineas, asi que un
    # documento se contaria varias veces.
    docs_raw = sorted(
        (t for t in celdas["fecha_emision"] if re.match(r"\d{1,2}/\d{1,2}/\d{2,4}", t[2])),
        key=lambda t: t[0],
    )
    if not docs_raw:
        docs_raw = sorted(
            (t for t in celdas["ejercicio"] if re.match(r"\d{1,2}/\d{4}", t[2])),
            key=lambda t: t[0],
        )
    if not docs_raw:
        return []

    # Cada documento ocupa una banda vertical. Con un solo documento la banda es
    # toda la tabla; con varios se corta a mitad de camino entre ellos. Es
    # necesario porque la celda del documento va centrada respecto de sus lineas
    # de impuesto: la primera linea puede quedar POR ENCIMA del nombre.
    bandas: list[tuple[float, float]] = []
    for i, (y_doc, _, _) in enumerate(docs_raw):
        arriba = y_ini if i == 0 else (docs_raw[i - 1][0] + y_doc) / 2
        abajo = y_fin if i == len(docs_raw) - 1 else (y_doc + docs_raw[i + 1][0]) / 2
        bandas.append((arriba, abajo))

    def _en_banda(col: str, banda: tuple[float, float]) -> list[tuple[float, float, str]]:
        return sorted(
            (t for t in celdas[col] if banda[0] <= t[0] < banda[1]),
            key=lambda t: (t[0], t[1]),
        )

    documentos = []
    for (y_doc, _, _), banda in zip(docs_raw, bandas):
        tipo = " ".join(t for _, _, t in _en_banda("comprobante", banda))
        numero = "".join(t for _, _, t in _en_banda("numero", banda))
        fechas = [t for _, _, t in _en_banda("fecha_emision", banda)]
        ejercicios = [t for _, _, t in _en_banda("ejercicio", banda)]

        # Las lineas de impuesto se anclan en la columna de valor retenido, que
        # siempre es de una sola linea (a diferencia de 'Impuesto a la Renta').
        lineas = []
        for y_val, _, texto_val in _en_banda("valor", banda):
            def _mismo_alto(col: str, tol: float = 3.0) -> str:
                return " ".join(
                    t for y, _, t in _en_banda(col, banda) if abs(y - y_val) <= tol
                )

            etiqueta = " ".join(
                t for y, _, t in _en_banda("impuesto", banda) if abs(y - y_val) <= 6.5
            )
            lineas.append(
                {
                    "impuesto": etiqueta.strip(),
                    "base": _a_float(_mismo_alto("base")),
                    "porcentaje": _a_float(_mismo_alto("porcentaje")),
                    "valor_retenido": _a_float(texto_val),
                }
            )

        documentos.append(
            {
                "tipo": tipo.strip().upper(),
                "numero": _normalizar_numero_doc(numero),
                "fecha_emision": fechas[0] if fechas else "",
                "ejercicio_fiscal": ejercicios[0] if ejercicios else "",
                "lineas": lineas,
            }
        )
    return documentos


def extraer_retencion_pdf(pdf_path: Path) -> Optional[dict]:
    """Lee un RIDE de comprobante de retencion. None si no se puede parsear."""
    try:
        palabras = _palabras_pdf(Path(pdf_path))
    except Exception as err:
        logger.warning(f"No se pudo abrir {Path(pdf_path).name}: {err}")
        return None
    if not palabras:
        return None

    texto_plano = " ".join(w[4] for w in palabras)
    claves = re.findall(r"\d{49}", texto_plano)

    return {
        "origen": "pdf",
        "archivo": str(pdf_path),
        "clave_acceso": claves[0] if claves else "",
        "numero": _normalizar_numero_doc(
            _valor_junto_a(palabras, "No.", r"\d{3}-\d{3}-\d{9}")
        ),
        "fecha_emision": _valor_junto_a(palabras, "Fecha", r"\d{1,2}/\d{1,2}/\d{4}"),
        "ruc_emisor": _valor_junto_a(palabras, "R.U.C.:", r"\d{10,13}"),
        "razon_social_emisor": _razon_social_emisor(palabras),
        "identificacion_sujeto": _valor_junto_a(
            palabras, "Identificacion", r"\d{10,13}"
        ),
        "razon_social_sujeto": _valor_junto_a(
            palabras, "Razon Social / Nombres y Apellidos:"
        ),
        "documentos": _extraer_documentos_pdf(palabras),
    }


# =============================================================================
# Extraccion del XML de una retencion (cuando esta disponible)
# =============================================================================


def extraer_retencion_xml(xml_path: Path) -> Optional[dict]:
    """Lee el XML de una retencion agrupando cada docSustento con SUS impuestos.

    robot/pdf_extraction.py toma un solo docSustento con `.find()` pero acumula
    los impuestos de todos con `.findall()`; si la retencion cubre dos facturas,
    los valores de la segunda se le atribuyen a la primera. Aca se agrupa bien.
    """
    try:
        texto = Path(xml_path).read_text(encoding="utf-8", errors="replace")
        # El XML del SRI suele venir envuelto en <autorizacion><comprobante>.
        interno = re.search(r"<comprobante>\s*<!\[CDATA\[([\s\S]*?)\]\]>", texto)
        if interno:
            texto = interno.group(1)
        root = ET.fromstring(texto)
    except Exception as err:
        logger.warning(f"No se pudo leer XML {Path(xml_path).name}: {err}")
        return None

    info = root.find(".//infoCompRetencion")
    tributaria = root.find(".//infoTributaria")

    def _txt(nodo, tag: str) -> str:
        if nodo is None:
            return ""
        return (nodo.findtext(tag) or "").strip()

    documentos = []
    for doc in root.findall(".//docsSustento/docSustento"):
        cod = _txt(doc, "codDocSustento")
        lineas = []
        for ret in doc.findall("./retenciones/retencion"):
            codigo = _txt(ret, "codigo")
            lineas.append(
                {
                    "impuesto": {"1": "RENTA", "2": "IVA", "6": "ISD"}.get(codigo, codigo),
                    "base": _a_float(_txt(ret, "baseImponible")),
                    "porcentaje": _a_float(_txt(ret, "porcentajeRetener")),
                    "valor_retenido": _a_float(_txt(ret, "valorRetenido")),
                }
            )
        documentos.append(
            {
                "tipo": DOC_SUSTENTO_LABELS.get(cod, cod),
                "cod_doc_sustento": cod,
                "numero": _normalizar_numero_doc(_txt(doc, "numDocSustento")),
                "fecha_emision": _txt(doc, "fechaEmisionDocSustento"),
                "ejercicio_fiscal": "",
                # El XML si trae la clave de la factura; cuando existe se usa
                # directo y no hace falta cruzar por numero.
                "num_aut_doc_sustento": _txt(doc, "numAutDocSustento"),
                "total_sin_impuestos": _a_float(_txt(doc, "totalSinImpuestos")),
                "importe_total": _a_float(_txt(doc, "importeTotal")),
                "lineas": lineas,
            }
        )

    return {
        "origen": "xml",
        "archivo": str(xml_path),
        "clave_acceso": _txt(tributaria, "claveAcceso"),
        "numero": _numero_desde_partes(
            _txt(tributaria, "estab"), _txt(tributaria, "ptoEmi"), _txt(tributaria, "secuencial")
        ),
        "fecha_emision": _txt(info, "fechaEmision"),
        "ruc_emisor": _txt(tributaria, "ruc"),
        "razon_social_emisor": _txt(tributaria, "razonSocial"),
        "identificacion_sujeto": _txt(info, "identificacionSujetoRetenido"),
        "razon_social_sujeto": _txt(info, "razonSocialSujetoRetenido"),
        "documentos": documentos,
    }


def cargar_retenciones(carpeta: Path) -> list[dict]:
    """Lee todas las retenciones de una carpeta. Prefiere XML sobre PDF."""
    carpeta = Path(carpeta)
    retenciones: list[dict] = []
    vistos: set[str] = set()

    for xml_path in sorted(carpeta.rglob("*.xml")):
        datos = extraer_retencion_xml(xml_path)
        if datos and datos.get("documentos"):
            retenciones.append(datos)
            if datos.get("clave_acceso"):
                vistos.add(datos["clave_acceso"])

    for pdf_path in sorted(carpeta.rglob("*.pdf")):
        if "retencion" not in _norm(pdf_path.name):
            continue
        datos = extraer_retencion_pdf(pdf_path)
        if not datos:
            continue
        if datos.get("clave_acceso") and datos["clave_acceso"] in vistos:
            continue  # ya lo tenemos por XML, que es mas confiable
        retenciones.append(datos)
    return retenciones


# =============================================================================
# Indice de facturas recibidas
# =============================================================================


def _factura_desde_xml(xml_path: Path) -> Optional[dict]:
    try:
        texto = Path(xml_path).read_text(encoding="utf-8", errors="replace")
        interno = re.search(r"<comprobante>\s*<!\[CDATA\[([\s\S]*?)\]\]>", texto)
        if interno:
            texto = interno.group(1)
        root = ET.fromstring(texto)
    except Exception:
        return None

    tributaria = root.find(".//infoTributaria")
    factura = root.find(".//infoFactura")
    if tributaria is None or factura is None:
        return None

    def _txt(nodo, tag: str) -> str:
        return (nodo.findtext(tag) or "").strip() if nodo is not None else ""

    ruc = _txt(tributaria, "ruc")
    numero = _numero_desde_partes(
        _txt(tributaria, "estab"), _txt(tributaria, "ptoEmi"), _txt(tributaria, "secuencial")
    )

    # El IVA desglosado solo esta en el XML: el listado del portal (y el TXT del
    # modo rapido) traen unicamente el importe total. Hace falta para comprobar
    # que la base de la retencion de IVA es el IVA real de la factura.
    iva_valor = 0.0
    iva_base = 0.0
    hay_iva = False
    for imp in root.findall(".//totalConImpuestos/totalImpuesto"):
        if (imp.findtext("codigo") or "").strip() != "2":  # 2 = IVA
            continue
        hay_iva = True
        iva_valor += _a_float(imp.findtext("valor")) or 0.0
        iva_base += _a_float(imp.findtext("baseImponible")) or 0.0

    return {
        "ruc_emisor": ruc,
        "razon_social_emisor": _txt(tributaria, "razonSocial"),
        "numero": numero,
        "clave_acceso": _txt(tributaria, "claveAcceso"),
        "fecha_emision": _txt(factura, "fechaEmision"),
        "total_sin_impuestos": _a_float(_txt(factura, "totalSinImpuestos")),
        "total_descuento": _a_float(_txt(factura, "totalDescuento")),
        "iva_factura": round(iva_valor, 2) if hay_iva else None,
        "base_iva_factura": round(iva_base, 2) if hay_iva else None,
        "importe_total": _a_float(_txt(factura, "importeTotal")),
        "origen": "xml",
        "archivo": str(xml_path),
    }


# Se comparan con _norm_col, que unifica espacios y guiones bajos: el TXT del
# portal usa RUC_EMISOR y los reportes internos "RUC Emisor".
_ALIAS_RUC = {"ruc emisor", "ruc", "identificacion emisor"}
_ALIAS_SERIE = {"serie comprobante", "serie", "numero comprobante", "num comprobante"}
_ALIAS_CLAVE = {"clave de acceso", "clave acceso", "claveacceso", "numero autorizacion"}
_ALIAS_TOTAL = {"importe total", "valor total", "total", "importe"}
_ALIAS_FECHA_EMI = {"fecha emision", "fecha de emision"}
_ALIAS_RAZON = {"razon social emisor", "razon social", "nombre emisor"}
# El listado de Recibidos del portal trae el desglose, no solo el total: sus
# columnas de subtotal e IVA alcanzan para verificar sobre que base se calculo
# la retencion, sin necesidad de descargar el XML de cada factura.
_ALIAS_SUBTOTAL = {
    "valor sin impuestos", "valor sin impuesto", "total sin impuestos",
    "subtotal", "base imponible", "valor neto",
}
_ALIAS_IVA = {
    "iva", "valor iva", "monto iva", "total iva", "impuesto iva",
    "impuesto al valor agregado",
}
# El reporte de Emitidos publica la serie en tres columnas separadas y sin ceros
# a la izquierda, en vez de una sola ya armada.
_ALIAS_ESTAB = {"establecimiento", "estab"}
_ALIAS_PTO = {"punto de emision", "punto emision", "ptoemi", "pto emision"}
_ALIAS_SEC = {"secuencial", "nro secuencial"}


def _norm_col(texto: object) -> str:
    """Normaliza un encabezado de columna tratando _ y - como espacios.

    El TXT del portal titula sus columnas con guion bajo (`VALOR_SIN_IMPUESTOS`)
    y los reportes internos con espacios (`Valor Sin Impuestos`). Sin unificar
    los separadores, las columnas del listado no matchean ningun alias y el
    cruce sale sin subtotal, IVA ni importe.
    """
    return re.sub(r"[\s_\-]+", " ", _norm(texto)).strip()


def _factura_desde_excel(path: Path) -> list[dict]:
    """Lee el Excel del modo rapido (reporte TXT) de facturas recibidas."""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as err:
        logger.warning(f"No se pudo abrir {Path(path).name}: {err}")
        return []

    facturas: list[dict] = []
    for ws in wb.worksheets:
        filas = ws.iter_rows(values_only=True)
        try:
            encabezado = next(filas)
        except StopIteration:
            continue
        idx = {_norm_col(h): i for i, h in enumerate(encabezado) if h}

        def _col(alias: set[str]) -> Optional[int]:
            for nombre, i in idx.items():
                if nombre in alias:
                    return i
            return None

        i_ruc, i_serie = _col(_ALIAS_RUC), _col(_ALIAS_SERIE)
        # El reporte de Emitidos no trae la serie armada sino establecimiento,
        # punto de emision y secuencial en columnas aparte.
        i_estab, i_pto, i_sec = _col(_ALIAS_ESTAB), _col(_ALIAS_PTO), _col(_ALIAS_SEC)
        tiene_partes = None not in (i_estab, i_pto, i_sec)
        if i_ruc is None or (i_serie is None and not tiene_partes):
            continue
        i_clave, i_total = _col(_ALIAS_CLAVE), _col(_ALIAS_TOTAL)
        i_fecha, i_razon = _col(_ALIAS_FECHA_EMI), _col(_ALIAS_RAZON)
        i_subtotal, i_iva = _col(_ALIAS_SUBTOTAL), _col(_ALIAS_IVA)
        if i_subtotal is None or i_iva is None:
            logger.info(
                f"{Path(path).name}: el listado no trae subtotal/IVA "
                f"({sorted(idx)}). Solo se podra verificar base x porcentaje."
            )

        def _num(fila, i):
            return _a_float(fila[i]) if i is not None and i < len(fila) else None

        def _celda(fila, i):
            return str(fila[i] or "").strip() if i is not None and i < len(fila) else ""

        for fila in filas:
            if not fila or i_ruc >= len(fila):
                continue
            ruc = _celda(fila, i_ruc)
            if i_serie is not None:
                numero = _normalizar_numero_doc(_celda(fila, i_serie))
            else:
                numero = _numero_desde_partes(
                    _celda(fila, i_estab), _celda(fila, i_pto), _celda(fila, i_sec)
                )
            if not ruc or not numero:
                continue
            facturas.append(
                {
                    "ruc_emisor": ruc,
                    "razon_social_emisor": _celda(fila, i_razon),
                    "numero": numero,
                    "clave_acceso": _celda(fila, i_clave),
                    "fecha_emision": _celda(fila, i_fecha),
                    "total_sin_impuestos": _num(fila, i_subtotal),
                    "total_descuento": None,
                    "iva_factura": _num(fila, i_iva),
                    "base_iva_factura": None,
                    "importe_total": _num(fila, i_total),
                    "origen": "listado",
                    "archivo": str(path),
                }
            )
    return facturas


def construir_indice_facturas(carpetas: Iterable[Path]) -> dict[str, dict]:
    """Indexa las facturas recibidas por RUC del emisor + numero.

    Acepta XML sueltos y los Excel que produce el modo rapido. Si la misma
    factura aparece en ambos, gana el XML: trae subtotal y descuento.
    """
    indice: dict[str, dict] = {}
    for carpeta in carpetas:
        carpeta = Path(carpeta)
        if not carpeta.exists():
            continue
        candidatos: list[dict] = []
        if carpeta.is_file():
            archivos = [carpeta]
        else:
            archivos = sorted(carpeta.rglob("*.xml")) + sorted(carpeta.rglob("*.xlsx"))
        for archivo in archivos:
            if archivo.suffix.lower() == ".xml":
                factura = _factura_desde_xml(archivo)
                if factura:
                    candidatos.append(factura)
            elif archivo.suffix.lower() == ".xlsx" and not archivo.name.startswith("~$"):
                candidatos.extend(_factura_desde_excel(archivo))

        for factura in candidatos:
            clave = _clave_match(factura["ruc_emisor"], factura["numero"])
            if not clave:
                continue
            previo = indice.get(clave)
            if previo is None or (previo["origen"] != "xml" and factura["origen"] == "xml"):
                indice[clave] = factura
    return indice


def descargar_listados_facturas(
    *,
    ruc: str,
    clave: str,
    destino: Path,
    periodos: Iterable[tuple[int, int]],
    sentido: str = SENTIDO_EMITIDAS,
    progress: Optional[Callable[[str], None]] = None,
) -> list[dict]:
    """Trae del portal el LISTADO de facturas de los meses indicados.

    Consulta por mes, no por factura. Aunque el form de Recibidos permita buscar
    por clave de acceso, hacerlo comprobante por comprobante serian N envios del
    formulario -N veces la exposicion al captcha- para obtener lo mismo que una
    sola consulta mensual devuelve de una.

    Siempre en modo rapido: baja el TXT que publica el portal en vez del PDF o
    el XML de cada factura. Ese listado ya trae subtotal, IVA, importe total y
    clave de acceso, que es todo lo que el cruce necesita.

    Un unico `descargar_sri` por anio, cubriendo de su mes menor al mayor, para
    que sea una sola sesion y un solo login. Ojo: en Emitidos el portal filtra
    por UN dia, asi que internamente consulta dia por dia; sigue siendo una
    sola sesion, pero tarda mas que en Recibidos.

    Las descargas van a `destino/<ruc>`, la misma convencion que usa el modulo
    de Descarga de Comprobantes, para que despues se encuentren desde el flujo
    local sin volver a bajarlas.
    """
    from robot.downloader import descargar_sri

    sentido = _normalizar_sentido(sentido)
    origen = _origen_facturas(sentido)
    ruc_limpio = re.sub(r"\D", "", str(ruc or ""))
    destino_ruc = Path(destino) / ruc_limpio if ruc_limpio else Path(destino)

    def _emit(msg: str) -> None:
        if progress:
            try:
                progress(msg)
            except Exception:
                pass
        logger.info(msg)

    por_anio: dict[int, list[int]] = {}
    for anio, mes in sorted(set(periodos)):
        por_anio.setdefault(int(anio), []).append(int(mes))

    if origen == "Emitidos":
        _emit(
            "Facturas emitidas: el portal filtra por dia, asi que consulta dia "
            "por dia. Tarda mas que en Recibidos, pero no descarga comprobantes."
        )

    resultados = []
    for anio, meses in sorted(por_anio.items()):
        inicio, fin = min(meses), max(meses)
        pedidos = ", ".join(_MESES[m] for m in sorted(set(meses)))
        etiqueta = "recibidas" if origen == "Recibidos" else "emitidas"
        if fin > inicio:
            _emit(f"Consultando facturas {etiqueta} {pedidos} de {anio} (rango {inicio}-{fin})...")
        else:
            _emit(f"Consultando facturas {etiqueta} de {pedidos} {anio}...")
        resultados.append(
            descargar_sri(
                ruc=ruc,
                clave=clave,
                anio=anio,
                mes=inicio,
                mes_fin=fin if fin > inicio else None,
                dia=0,
                tipo="Factura",
                formatos=[],          # modo rapido: sin PDF ni XML
                destino=destino_ruc,
                origen=origen,
                modo_rapido=True,
            )
        )
    return resultados


def inferir_base_ruc(carpeta: Path) -> Optional[Path]:
    """Sube por el arbol buscando la carpeta [RUC] que contiene 'Recibidos'.

    Las descargas quedan como [base]/[RUC]/[Emitidos|Recibidos]/[Tipo]/[anio]/
    [Mes]/[XML|PDF], asi que desde la carpeta de retenciones se puede ubicar
    sola la de facturas recibidas.
    """
    carpeta = Path(carpeta).resolve()
    for candidata in [carpeta, *carpeta.parents]:
        if (candidata / "Recibidos").is_dir() or (candidata / "Emitidos").is_dir():
            return candidata
    return None


def rutas_facturas_sugeridas(
    base_ruc: Path, fechas: Iterable[datetime], sentido: str = SENTIDO_EMITIDAS
) -> list[Path]:
    """Carpetas de facturas para los meses de las facturas sustento.

    La fecha sale del propio RIDE, asi que se buscan solo los meses que hacen
    falta en vez de barrer un rango a ciegas.
    """
    base_ruc = Path(base_ruc)
    origen = _origen_facturas(sentido)
    rutas: list[Path] = []
    vistas: set[Path] = set()
    for fecha in fechas:
        if not fecha:
            continue
        carpeta = base_ruc / origen / "Factura" / f"{fecha.year:04d}" / _MESES[fecha.month]
        if carpeta not in vistas:
            vistas.add(carpeta)
            rutas.append(carpeta)
    return rutas


# =============================================================================
# COLUMNA CALCULADA: DIAS ENTRE LA FACTURA Y LA RETENCION
# =============================================================================
# El reporte es descriptivo: pone los datos de la retencion y los de su factura
# uno al lado del otro. Lo unico que se calcula son los dias transcurridos entre
# la emision de la factura y la de la retencion.
#
# La fecha de la factura sale del comprobante encontrado, que es la fuente
# autorizada; si no se lo encontro, se cae a la que declara el RIDE de la
# retencion en su documento sustento. Asi la columna se llena igual en las filas
# sin factura, que es donde el dato suele importar mas.
#
# Para agregar mas columnas calculadas: sumar el nombre a COLUMNAS_OPERACION y
# devolverlo desde calcular_operacion(). El Excel se arma con lo que devuelva.

COLUMNAS_OPERACION: list[str] = [
    "Dias entre factura y retencion",
]


def calcular_operacion(
    retencion: dict, documento: dict, factura: Optional[dict]
) -> dict:
    """Devuelve {nombre_columna: valor} para agregar al final de la fila.

    Los dias son `fecha de la retencion - fecha de la factura`, asi que un
    valor positivo significa que la retencion se emitio despues de la factura,
    que es lo normal. Un negativo indica que la retencion es anterior a la
    factura que dice sustentar.
    """
    fecha_retencion = _parse_fecha(retencion.get("fecha_emision"))
    fecha_factura = _parse_fecha((factura or {}).get("fecha_emision")) or _parse_fecha(
        documento.get("fecha_emision")
    )
    if not fecha_retencion or not fecha_factura:
        return {"Dias entre factura y retencion": ""}
    return {"Dias entre factura y retencion": (fecha_retencion - fecha_factura).days}


# =============================================================================
# Armado del reporte
# =============================================================================


def _linea_por_impuesto(documento: dict, *nombres: str) -> Optional[dict]:
    for linea in documento.get("lineas", []):
        etiqueta = _norm(linea.get("impuesto"))
        if any(nombre in etiqueta for nombre in nombres):
            return linea
    return None


def _observacion_no_encontrada(
    retencion: dict,
    documento: dict,
    origen_facturas: str,
    facturas_del_mes: Optional[int] = None,
) -> str:
    """Explica por que una factura no aparecio, segun se haya revisado o no su mes.

    Son dos situaciones distintas y la accion que sigue tambien:

    - El mes NO se indexo: falta descargarlo, y punto.
    - El mes SI se indexo y la factura igual no esta: el portal no la tiene.
      Lo mas comun es que no sea electronica -- una factura preimpresa sustenta
      la retencion igual pero nunca figura en el listado.

    Decir siempre "revisa que este descargado el mes" mandaba a revisar
    justamente lo unico que ya estaba bien.
    """
    fecha = documento.get("fecha_emision") or "la fecha que declara la retencion"
    identificacion = re.sub(r"\D", "", str(retencion.get("identificacion_sujeto") or ""))

    if not facturas_del_mes:
        return (
            f"No se indexo ninguna factura de {origen_facturas} de {fecha}. "
            "Falta descargar ese mes: volve a generar el reporte con la busqueda "
            "automatica marcada, o indica la carpeta donde estan."
        )

    pista_cedula = (
        " El sujeto retenido esta identificado con cedula y no con RUC, lo que "
        "refuerza esa posibilidad."
        if len(identificacion) == 10
        else ""
    )
    return (
        f"El listado de {origen_facturas} de {fecha} si se reviso "
        f"({facturas_del_mes} factura(s)) y esta no figura. Lo mas probable es "
        "que no sea electronica: una factura preimpresa sustenta la retencion "
        f"igual, pero no aparece en el portal.{pista_cedula} Tambien puede estar "
        "anulada o tener una fecha de emision distinta a la declarada."
    )


def _construir_fila(
    retencion: dict, documento: dict, factura: Optional[dict], estado: str, observacion: str
) -> dict:
    iva = _linea_por_impuesto(documento, "iva") or {}
    renta = _linea_por_impuesto(documento, "renta") or {}

    # Etiquetas neutrales: en retenciones emitidas el sujeto retenido es el
    # proveedor y en las recibidas es el propio contribuyente, asi que se nombra
    # por el rol que cada uno cumple en el comprobante y no por la relacion
    # comercial. Quien emitio la factura es SIEMPRE el sujeto retenido.
    fila = {
        "RUC agente de retencion": retencion.get("ruc_emisor", ""),
        "Razon social agente de retencion": retencion.get("razon_social_emisor", ""),
        "Numero retencion": retencion.get("numero", ""),
        "Fecha retencion": retencion.get("fecha_emision", ""),
        "Clave acceso retencion": retencion.get("clave_acceso", ""),
        "RUC sujeto retenido": retencion.get("identificacion_sujeto", ""),
        "Razon social sujeto retenido": retencion.get("razon_social_sujeto", ""),
        "Tipo documento sustento": documento.get("tipo", ""),
        "Numero factura": documento.get("numero", ""),
        "Fecha factura (segun retencion)": documento.get("fecha_emision", ""),
        "Ejercicio fiscal": documento.get("ejercicio_fiscal", ""),
        "Base retencion IVA": iva.get("base", ""),
        "Porcentaje retencion IVA": iva.get("porcentaje", ""),
        "Valor retenido IVA": iva.get("valor_retenido", ""),
        "Base retencion Renta": renta.get("base", ""),
        "Porcentaje retencion Renta": renta.get("porcentaje", ""),
        "Valor retenido Renta": renta.get("valor_retenido", ""),
        "Clave acceso factura": factura.get("clave_acceso", "") if factura else "",
        "Fecha factura (segun listado)": factura.get("fecha_emision", "") if factura else "",
        "Subtotal factura": factura.get("total_sin_impuestos", "") if factura else "",
        "IVA factura": factura.get("iva_factura", "") if factura else "",
        "Importe total factura": factura.get("importe_total", "") if factura else "",
        "Origen datos factura": factura.get("origen", "") if factura else "",
        "Estado": estado,
        "Observacion": observacion,
    }
    fila.update(calcular_operacion(retencion, documento, factura))
    return fila


def _escribir_excel(filas_cruce: list[dict], filas_otros: list[dict], path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    encabezado_fill = PatternFill("solid", fgColor="1F4E78")
    encabezado_font = Font(color="FFFFFF", bold=True)

    def _hoja(titulo: str, filas: list[dict], vacio: str) -> None:
        ws = wb.create_sheet(titulo[:31])
        if not filas:
            ws.cell(row=1, column=1, value=vacio)
            return
        columnas = list(filas[0].keys())
        for i, nombre in enumerate(columnas, start=1):
            celda = ws.cell(row=1, column=i, value=nombre)
            celda.fill = encabezado_fill
            celda.font = encabezado_font
            celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for r, fila in enumerate(filas, start=2):
            for c, nombre in enumerate(columnas, start=1):
                valor = fila.get(nombre, "")
                celda = ws.cell(row=r, column=c, value=valor)
                # Claves y RUC como texto: si no, Excel los pasa a notacion
                # cientifica y les come digitos.
                if any(t in nombre for t in ("Clave", "RUC", "Numero")):
                    celda.number_format = "@"
                elif isinstance(valor, float):
                    celda.number_format = "#,##0.00"
        for i, nombre in enumerate(columnas, start=1):
            ancho = max(len(str(nombre)), *(len(str(f.get(nombre, ""))) for f in filas))
            ws.column_dimensions[get_column_letter(i)].width = min(max(ancho + 2, 12), 45)
        ws.freeze_panes = "A2"

    _hoja("Retenciones vs Facturas", filas_cruce, "Sin retenciones con factura de sustento.")
    _hoja(
        "Sustento no factura",
        filas_otros,
        "Todos los sustentos eran facturas.",
    )
    wb.remove(wb["Sheet"])
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def generar_reporte_retenciones(
    *,
    carpeta_retenciones: str | Path,
    salida_excel: str | Path,
    carpetas_facturas: Optional[Iterable[str | Path]] = None,
    base_ruc: Optional[str | Path] = None,
    ruc: Optional[str] = None,
    clave: Optional[str] = None,
    destino_descargas: Optional[str | Path] = None,
    sentido: str = SENTIDO_EMITIDAS,
    cancel_event: Optional[threading.Event] = None,
    progress: Optional[Callable[[str], None]] = None,
) -> dict:
    """Punto de entrada del modulo.

    Lee las retenciones de `carpeta_retenciones` (XML o PDF), busca cada factura
    de sustento en `carpetas_facturas` -o, si se pasa `base_ruc`, en las carpetas
    de Recibidos/Factura de los meses que indiquen las propias retenciones- y
    escribe el Excel en `salida_excel`.

    Devuelve un resumen: ok, total_retenciones, total_documentos, con_factura,
    sin_factura, no_facturas, excel_path, message.
    """

    def _emit(msg: str) -> None:
        if progress:
            try:
                progress(msg)
            except Exception:
                pass
        logger.info(msg)

    def _cancelado() -> bool:
        return bool(cancel_event and cancel_event.is_set())

    resumen = {
        "ok": False,
        "total_retenciones": 0,
        "total_documentos": 0,
        "con_factura": 0,
        "sin_factura": 0,
        # Desglose de `sin_factura`: el mes se reviso y no estaba, o falta bajarlo.
        "sin_factura_electronica": 0,
        "mes_sin_descargar": 0,
        "no_facturas": 0,
        "excel_path": "",
        "message": "",
    }

    sentido = _normalizar_sentido(sentido)
    origen_facturas = _origen_facturas(sentido)
    resumen["sentido"] = sentido

    carpeta = Path(carpeta_retenciones).expanduser()
    if not carpeta.is_dir():
        resumen["message"] = f"La carpeta de retenciones no existe: {carpeta}"
        return resumen

    _emit(
        f"Retenciones {sentido}: las facturas de sustento se buscan en "
        f"{origen_facturas}."
    )
    _emit("Leyendo comprobantes de retencion...")
    retenciones = cargar_retenciones(carpeta)
    resumen["total_retenciones"] = len(retenciones)
    if not retenciones:
        resumen["message"] = (
            "No se encontraron comprobantes de retencion legibles en la carpeta."
        )
        return resumen
    _emit(f"{len(retenciones)} retencion(es) leida(s).")

    if _cancelado():
        resumen["message"] = "Cancelado por el usuario."
        return resumen

    # Meses en los que viven las facturas sustento. Salen del propio RIDE, asi
    # que no hay que barrer un rango a ciegas.
    fechas_sustento = [
        _parse_fecha(doc.get("fecha_emision"))
        for ret in retenciones
        for doc in ret.get("documentos", [])
    ]
    periodos = sorted({(f.year, f.month) for f in fechas_sustento if f})

    # Carpetas de facturas: explicitas, o deducidas de las fechas de sustento.
    rutas = [Path(c).expanduser() for c in (carpetas_facturas or [])]

    # Si hay credenciales, se trae el listado del portal. Una consulta por mes
    # -no por factura- para minimizar la exposicion al captcha.
    if ruc and clave:
        if not periodos:
            _emit("No se pudo deducir el mes de ninguna factura sustento.")
        else:
            carpeta_descarga = Path(
                destino_descargas or (carpeta.parent / "_facturas_listado")
            ).expanduser()
            meses_txt = ", ".join(f"{_MESES[m]} {a}" for a, m in periodos)
            _emit(
                f"Las retenciones apuntan a facturas de: {meses_txt}. "
                f"Se consultara {origen_facturas} en el portal."
            )
            try:
                descargar_listados_facturas(
                    ruc=ruc,
                    clave=clave,
                    destino=carpeta_descarga,
                    periodos=periodos,
                    sentido=sentido,
                    progress=progress,
                )
                # descargar_listados_facturas guarda bajo destino/<ruc>.
                rutas.append(carpeta_descarga)
            except Exception as err:
                logger.warning(f"Fallo la consulta al portal: {err}")
                _emit(
                    f"No se pudo consultar el portal ({err}). Se sigue con lo "
                    "que haya en disco."
                )
        if _cancelado():
            resumen["message"] = "Cancelado por el usuario."
            return resumen
    raiz = Path(base_ruc).expanduser() if base_ruc else inferir_base_ruc(carpeta)
    if raiz:
        if not base_ruc:
            _emit(f"Carpeta del RUC deducida: {raiz}")
        rutas.extend(rutas_facturas_sugeridas(raiz, fechas_sustento, sentido))
    elif not rutas:
        _emit(
            f"No se ubico la carpeta de facturas de {origen_facturas}. Indicala a mano o "
            "usa una carpeta de retenciones que cuelgue de la carpeta del RUC."
        )
    if rutas:
        _emit(f"Indexando facturas de {origen_facturas} en {len(rutas)} carpeta(s)...")
    indice = construir_indice_facturas(rutas)
    _emit(f"{len(indice)} factura(s) de {origen_facturas} indexada(s).")

    # Meses que el indice llego a cubrir. Sirve para distinguir dos situaciones
    # que hoy se confundian: que falte descargar el mes, o que el mes se haya
    # revisado y la factura igual no este -- caso tipico de una factura
    # preimpresa, que sustenta la retencion pero no figura en el portal.
    facturas_por_mes: dict[tuple[int, int], int] = {}
    for factura_indexada in indice.values():
        fecha_ind = _parse_fecha(factura_indexada.get("fecha_emision"))
        if fecha_ind:
            periodo = (fecha_ind.year, fecha_ind.month)
            facturas_por_mes[periodo] = facturas_por_mes.get(periodo, 0) + 1
    if facturas_por_mes:
        _emit(
            "Meses cubiertos: "
            + ", ".join(
                f"{_MESES[m]} {a} ({n})"
                for (a, m), n in sorted(facturas_por_mes.items())
            )
        )

    filas_cruce: list[dict] = []
    filas_otros: list[dict] = []

    for retencion in retenciones:
        if _cancelado():
            resumen["message"] = "Cancelado por el usuario."
            return resumen
        for documento in retencion.get("documentos", []):
            resumen["total_documentos"] += 1
            tipo = _norm(documento.get("tipo"))
            cod = documento.get("cod_doc_sustento") or ""

            es_factura = cod == COD_FACTURA if cod else tipo.startswith("factura")
            if not es_factura:
                resumen["no_facturas"] += 1
                filas_otros.append(
                    _construir_fila(
                        retencion,
                        documento,
                        None,
                        "Sustento no es factura",
                        f"El documento sustento es '{documento.get('tipo')}', "
                        "no cruza contra facturas recibidas.",
                    )
                )
                continue

            clave = _clave_match(
                retencion.get("identificacion_sujeto"), documento.get("numero")
            )
            factura = indice.get(clave)
            if factura:
                resumen["con_factura"] += 1
                filas_cruce.append(
                    _construir_fila(retencion, documento, factura, "Factura encontrada", "")
                )
            else:
                resumen["sin_factura"] += 1
                fecha_doc = _parse_fecha(documento.get("fecha_emision"))
                del_mes = (
                    facturas_por_mes.get((fecha_doc.year, fecha_doc.month), 0)
                    if fecha_doc
                    else 0
                )
                # Si el mes se reviso y la factura no esta, el portal no la
                # tiene: no es lo mismo que no haberlo descargado.
                if del_mes:
                    resumen["sin_factura_electronica"] += 1
                    estado_fila = "Sin factura electronica"
                else:
                    resumen["mes_sin_descargar"] += 1
                    estado_fila = "Factura no encontrada"
                filas_cruce.append(
                    _construir_fila(
                        retencion,
                        documento,
                        None,
                        estado_fila,
                        _observacion_no_encontrada(
                            retencion, documento, origen_facturas, del_mes
                        ),
                    )
                )

    salida = Path(salida_excel).expanduser()
    _emit("Escribiendo Excel...")
    _escribir_excel(filas_cruce, filas_otros, salida)

    resumen["ok"] = True
    resumen["excel_path"] = str(salida)
    partes = [f"{resumen['con_factura']} con factura"]
    if resumen["mes_sin_descargar"]:
        partes.append(f"{resumen['mes_sin_descargar']} con el mes sin descargar")
    if resumen["sin_factura_electronica"]:
        partes.append(
            f"{resumen['sin_factura_electronica']} sin factura electronica "
            "(el mes se reviso y no figuran)"
        )
    if resumen["no_facturas"]:
        partes.append(f"{resumen['no_facturas']} con sustento que no es factura")
    resumen["message"] = ", ".join(partes) + "."
    if not COLUMNAS_OPERACION:
        resumen["message"] += " Falta definir la operacion matematica."
    _emit(resumen["message"])
    return resumen
