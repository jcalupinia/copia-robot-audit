from pathlib import Path
from collections import Counter, defaultdict
from datetime import datetime
import json
import re
import pandas as pd
import xml.etree.ElementTree as ET
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.colors import ColorChoice
from openpyxl.chart.label import DataLabelList

DOC_TYPES = {
    '01': 'Factura',
    '03': 'Liquidacion de compra',
    '04': 'Nota de credito',
    '05': 'Nota de debito',
    '07': 'Comprobante de retencion',
}

DOC_TYPE_COUNTER_KEYS = {
    '01': 'facturas',
    '03': 'liquidaciones',
    '04': 'nc',
    '05': 'nd',
    '07': 'retenciones',
}

IMPUESTO_CODIGO_LABEL = {
    '2': 'IVA',
    '3': 'ICE',
    '5': 'IRBPNR',
}

RETENCION_TIPO_IMPUESTO = {
    '1': 'Renta',
    '2': 'IVA',
    '6': 'ISD',
}

FORMA_PAGO_LABEL = {
    '01': 'Sin utilizacion del sistema financiero',
    '15': 'Compensacion de deudas',
    '16': 'Tarjeta de debito',
    '17': 'Dinero electronico',
    '18': 'Tarjeta prepago',
    '19': 'Tarjeta de credito',
    '20': 'Otros con utilizacion del sistema financiero',
    '21': 'Endoso de titulos',
}

IVA_CODIGO_PORCENTAJE_LABEL = {
    '0': 'IVA 0%',
    '2': 'IVA 12%',
    '3': 'IVA 14%',
    '4': 'IVA 15%',
    '5': 'IVA 5%',
    '6': 'No objeto IVA',
    '7': 'Exento IVA',
}

AMBIENTE_LABEL = {
    '1': '1 - Pruebas',
    '2': '2 - Produccion',
}

TIPO_EMISION_LABEL = {
    '1': '1 - Emision normal',
    '2': '2 - Emision por indisponibilidad del sistema',
}

TIPO_IDENTIFICACION_LABEL = {
    '04': '04 - RUC',
    '05': '05 - Cedula',
    '06': '06 - Pasaporte',
    '07': '07 - Consumidor final',
    '08': '08 - Identificacion del exterior',
    '09': '09 - Placa',
}

DOC_CODE_LABEL = {
    '01': '01 - FACTURA',
    '03': '03 - LIQUIDACION DE COMPRA',
    '04': '04 - NOTA DE CREDITO',
    '05': '05 - NOTA DE DEBITO',
    '07': '07 - COMPROBANTE DE RETENCION',
}

IVA_TARIFA_LABEL = {
    '2': '12%',
    '3': '14%',
    '4': '15%',
    '5': '5%',
}

TAX_FULL_NAMES = {
    'IVA': 'Impuesto al valor agregado',
    'ICE': 'Impuesto a los consumos especiales',
    'IRBPNR': 'Impuesto redimible a las botellas plasticas no retornables',
}

COMMON_COLUMN_TITLES = {
    'CONSULTA_ID': 'ID de consulta',
    'FECHA_EJECUCION': 'Fecha y hora de ejecucion',
    'RANGO_FECHAS': 'Rango de fechas consultado',
    'TOTAL_DOCUMENTOS': 'Total de documentos',
    'ARCHIVO_XML': 'Archivo XML',
    'RUTA_XML': 'Ruta del archivo XML',
    'COD_DOC': 'Codigo del documento',
    'DESCRIPCION_DOC': 'Descripcion del comprobante',
    'CLAVE_ACCESO': 'Clave de acceso',
    'SERIE_COMPROBANTE': 'Serie del comprobante',
    'ESTADO_AUTORIZACION': 'Estado de autorizacion',
    'NUMERO_AUTORIZACION': 'Numero de autorizacion',
    'FECHA_AUTORIZACION': 'Fecha de autorizacion',
    'AMBIENTE': 'Ambiente',
    'ESTAB': 'Establecimiento',
    'PTO_EMI': 'Punto de emision',
    'SECUENCIAL': 'Secuencial',
    'RUC_EMISOR': 'RUC del emisor',
    'RAZON_SOCIAL_EMISOR': 'Razon social del emisor',
    'NOMBRE_COMERCIAL_EMISOR': 'Nombre comercial del emisor',
    'DIR_MATRIZ': 'Direccion de la matriz',
    'DIR_ESTABLECIMIENTO': 'Direccion del establecimiento',
    'TIPO_IDENTIFICACION_COMPRADOR': 'Tipo de identificacion del receptor',
    'IDENTIFICACION_COMPRADOR': 'Identificacion del receptor',
    'RAZON_SOCIAL_COMPRADOR': 'Nombre o razon social del receptor',
    'DIRECCION_COMPRADOR': 'Direccion del receptor',
    'FECHA_EMISION': 'Fecha de emision',
    'TOTAL_SIN_IMPUESTOS': 'Total sin impuestos',
    'TOTAL_DESCUENTO': 'Total de descuentos',
    'IMPORTE_TOTAL': 'Importe total',
    'MONEDA': 'Moneda',
    'MOTIVO': 'Motivo',
    'VALOR_MODIFICACION': 'Valor de la modificacion',
    'VALOR_TOTAL': 'Valor total',
    'TOTAL_RETENIDO': 'Total retenido',
    'PERIODO_FISCAL': 'Periodo fiscal',
    'EJERCICIO_FISCAL': 'Ejercicio fiscal',
    'LINE_NUM': 'Numero de linea',
    'CODIGO_PRINCIPAL': 'Codigo principal',
    'DESCRIPCION': 'Descripcion del item',
    'CANTIDAD': 'Cantidad',
    'PRECIO_UNITARIO': 'Precio unitario',
    'PRECIO_TOTAL_SIN_IMPUESTO': 'Precio total sin impuestos',
    'CENTRO_COSTO': 'Centro de costo',
    'CTA_CONTABLE': 'Cuenta contable',
    'NIVEL': 'Nivel del impuesto',
    'CODIGO': 'Codigo del impuesto',
    'CODIGO_PORCENTAJE': 'Codigo de porcentaje',
    'TARIFA': 'Tarifa',
    'BASE_IMPONIBLE': 'Base imponible',
    'VALOR': 'Valor del impuesto',
    'PAGO_NUM': 'Numero de pago',
    'FORMA_PAGO': 'Forma de pago (codigo)',
    'FORMA_PAGO_DESC': 'Descripcion de la forma de pago',
    'PLAZO': 'Plazo',
    'UNIDAD_TIEMPO': 'Unidad de tiempo',
    'TIPO_IMPUESTO': 'Tipo de impuesto',
    'CODIGO_RETENCION': 'Codigo de retencion',
    'PORCENTAJE_RETENER': 'Porcentaje a retener',
    'VALOR_RETENIDO': 'Valor retenido',
    'DOC_SUSTENTO_TIPO': 'Tipo del documento de sustento',
    'DOC_SUSTENTO_ESTAB': 'Establecimiento del documento de sustento',
    'DOC_SUSTENTO_PTO': 'Punto de emision del documento de sustento',
    'DOC_SUSTENTO_SECUENCIAL': 'Secuencial del documento de sustento',
    'FECHA_EMISION_DOC_SUSTENTO': 'Fecha del documento de sustento',
    'ERROR': 'Descripcion del error',
}


def _titulo_columna_legible(nombre: str, extras: dict | None = None) -> str:
    clave = nombre.upper()
    extras = {k.upper(): v for k, v in (extras or {}).items()}
    if clave in extras:
        return extras[clave]
    if clave in COMMON_COLUMN_TITLES:
        return COMMON_COLUMN_TITLES[clave]

    m = re.match(r"^FORMA_PAGO_(\d+)$", clave)
    if m:
        return f"Forma de pago {m.group(1)}"

    m = re.match(r"^TOTAL_PAGO_(\d+)$", clave)
    if m:
        return f"Monto del pago {m.group(1)}"

    m = re.match(r"^PLAZO_(\d+)$", clave)
    if m:
        return f"Plazo del pago {m.group(1)}"

    m = re.match(r"^UNIDAD_TIEMPO_(\d+)$", clave)
    if m:
        return f"Unidad de tiempo del pago {m.group(1)}"

    m = re.match(r"^(IVA|ICE|IRBPNR)_([0-9]{1,3})_(BASE|VALOR)$", clave)
    if m:
        impuesto, porcentaje, campo = m.groups()
        impuesto_nombre = TAX_FULL_NAMES.get(impuesto, impuesto.title())
        porcentaje_txt = f" {porcentaje}%"
        campo_txt = "base imponible" if campo == "BASE" else "valor"
        return f"{impuesto_nombre}{porcentaje_txt} {campo_txt}"

    m = re.match(r"^(IVA|ICE|IRBPNR)_(BASE|VALOR)$", clave)
    if m:
        impuesto, campo = m.groups()
        impuesto_nombre = TAX_FULL_NAMES.get(impuesto, impuesto.title())
        campo_txt = "base imponible" if campo == "BASE" else "valor"
        return f"{impuesto_nombre} {campo_txt}"

    m = re.match(r"^(IVA|ICE|IRBPNR)_(CODIGO_PORCENTAJE|TARIFA|BASE|VALOR)$", clave)
    if m:
        impuesto, atributo = m.groups()
        impuesto_nombre = TAX_FULL_NAMES.get(impuesto, impuesto.title())
        atributo_txt = {
            'CODIGO_PORCENTAJE': 'codigo de porcentaje',
            'TARIFA': 'tarifa',
            'BASE': 'base imponible',
            'VALOR': 'valor',
        }[atributo]
        return f"{impuesto_nombre} {atributo_txt}"

    m = re.match(r"^IMP_([0-9A-Z]+)_([0-9A-Z]+)_(BASE|VALOR)$", clave)
    if m:
        codigo, porcentaje, campo = m.groups()
        campo_txt = "base imponible" if campo == "BASE" else "valor"
        return f"Impuesto codigo {codigo} porcentaje {porcentaje} {campo_txt}"

    palabras = [p.lower() for p in clave.split('_')]
    texto = " ".join(palabras).strip()
    return texto.capitalize() if texto else clave


def _renombrar_dataframe(df: pd.DataFrame, extras: dict | None = None) -> pd.DataFrame:
    if df.empty and len(df.columns) == 0:
        return df
    extras = extras or {}
    nuevos = {col: _titulo_columna_legible(col, extras) for col in df.columns}
    return df.rename(columns=nuevos)


def _strip_namespaces(elem: ET.Element):
    for el in elem.iter():
        if '}' in el.tag:
            el.tag = el.tag.split('}', 1)[1]
        if ':' in el.tag:
            el.tag = el.tag.split(':', 1)[1]
        if el.attrib:
            el.attrib = {attr.split('}', 1)[-1].split(':', 1)[-1]: val for attr, val in el.attrib.items()}


def _clean_inner_xml(xml_text: str) -> ET.Element:
    if not xml_text:
        raise ValueError('El nodo <comprobante> esta vacio.')
    contenido = xml_text.strip()
    if contenido.startswith('<![CDATA['):
        contenido = contenido[9:-3]
    try:
        elemento = ET.fromstring(contenido)
    except ET.ParseError as err:
        raise ValueError(f'XML malformado en comprobante: {err}') from err
    _strip_namespaces(elemento)
    return elemento


def _safe_str(texto: str) -> str:
    return (texto or '').strip()


def _safe_float(valor: str) -> float:
    if valor is None:
        return 0.0
    texto = str(valor).strip().replace(' ', '')
    if not texto:
        return 0.0
    candidatos = [
        texto,
        texto.replace('.', '').replace(',', '.'),
        texto.replace(',', ''),
    ]
    for candidato in candidatos:
        try:
            return float(candidato)
        except ValueError:
            continue
    return 0.0


def _safe_datetime(texto: str) -> datetime | None:
    if not texto:
        return None
    limpio = re.sub(r'\s+', ' ', texto.strip())
    formatos = ('%d/%m/%Y %H:%M:%S', '%d/%m/%Y %H:%M', '%d/%m/%Y')
    for fmt in formatos:
        try:
            return datetime.strptime(limpio, fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(limpio)
    except ValueError:
        return None


def _doc_type_label(cod_doc: str) -> str:
    return DOC_TYPES.get(cod_doc, f'Desconocido ({cod_doc})')


def _serie(estab: str, pto_emi: str, secuencial: str) -> str:
    partes = [_safe_str(estab), _safe_str(pto_emi), _safe_str(secuencial)]
    return '-'.join(p for p in partes if p)

def _tax_column_names(codigo: str, codigo_pct: str):
    if codigo == '2':  # IVA
        label = IVA_CODIGO_PORCENTAJE_LABEL.get(codigo_pct, f'IVA {codigo_pct}%')
        slug = label.replace('%', '').replace(' ', '_').upper()
    elif codigo == '3':
        slug = 'ICE'
    elif codigo == '5':
        slug = 'IRBPNR'
    else:
        codigo_pct = codigo_pct or '0'
        slug = f'IMP_{codigo}_{codigo_pct}'
    base_col = f'{slug}_BASE'
    valor_col = f'{slug}_VALOR'
    return base_col, valor_col


def _parse_recibido_xml(xml_path: Path):
    try:
        root = ET.parse(xml_path).getroot()
    except Exception as err:
        return (
            None,
            [],
            [],
            [],
            [],
            [],
            {'archivo_xml': xml_path.name, 'clave_acceso': '', 'error': str(err)},
            {},
        )

    meta = {
        'estado_autorizacion': '',
        'numero_autorizacion': '',
        'fecha_autorizacion': None,
        'ambiente': '',
    }

    try:
        if root.tag.lower().endswith('autorizacion'):
            meta['estado_autorizacion'] = _safe_str(root.findtext('estado'))
            meta['numero_autorizacion'] = _safe_str(root.findtext('numeroAutorizacion'))
            meta['ambiente'] = _safe_str(root.findtext('ambiente'))
            meta['fecha_autorizacion'] = _safe_str(root.findtext('fechaAutorizacion'))
            doc_node = _clean_inner_xml(root.findtext('comprobante', ''))
        else:
            _strip_namespaces(root)
            doc_node = root
    except Exception as err:
        return (
            None,
            [],
            [],
            [],
            [],
            [],
            {
                'archivo_xml': xml_path.name,
                'clave_acceso': '',
                'error': f'Error extrayendo comprobante: {err}',
            },
            {},
        )

    info_trib = doc_node.find('infoTributaria')
    if info_trib is None:
        return (
            None,
            [],
            [],
            [],
            [],
            [],
            {
                'archivo_xml': xml_path.name,
                'clave_acceso': '',
                'error': 'Nodo infoTributaria no encontrado.',
            },
            {},
        )

    cod_doc = _safe_str(info_trib.findtext('codDoc'))
    estab = _safe_str(info_trib.findtext('estab'))
    pto_emi = _safe_str(info_trib.findtext('ptoEmi'))
    secuencial = _safe_str(info_trib.findtext('secuencial'))
    clave_acceso = _safe_str(info_trib.findtext('claveAcceso'))

    info_node = None
    for candidato in (
        'infoFactura',
        'infoLiquidacionCompra',
        'infoNotaCredito',
        'infoNotaDebito',
        'infoCompRetencion',
    ):
        info_node = doc_node.find(candidato)
        if info_node is not None:
            break

    cabecera = {
        'ARCHIVO_XML': xml_path.name,
        'RUTA_XML': str(xml_path),
        'COD_DOC': cod_doc,
        'DESCRIPCION_DOC': _doc_type_label(cod_doc),
        'CLAVE_ACCESO': clave_acceso,
        'SERIE_COMPROBANTE': _serie(estab, pto_emi, secuencial),
        'ESTADO_AUTORIZACION': meta['estado_autorizacion'],
        'NUMERO_AUTORIZACION': meta['numero_autorizacion'],
        'FECHA_AUTORIZACION': meta['fecha_autorizacion'],
        'AMBIENTE': meta['ambiente'],
        'TIPO_EMISION': _safe_str(info_trib.findtext('tipoEmision')),
        'CONTRIBUYENTE_RIMPE': _safe_str(info_trib.findtext('contribuyenteRimpe')),
        'VERSION_XML': doc_node.attrib.get('version', ''),
        'ESTAB': estab,
        'PTO_EMI': pto_emi,
        'SECUENCIAL': secuencial,
        'RUC_EMISOR': _safe_str(info_trib.findtext('ruc')),
        'RAZON_SOCIAL_EMISOR': _safe_str(info_trib.findtext('razonSocial')),
        'NOMBRE_COMERCIAL_EMISOR': _safe_str(info_trib.findtext('nombreComercial')),
        'DIR_MATRIZ': _safe_str(info_trib.findtext('dirMatriz')),
        'DIR_ESTABLECIMIENTO': '',
        'OBLIGADO_CONTABILIDAD': '',
        'TIPO_IDENTIFICACION_COMPRADOR': '',
        'IDENTIFICACION_COMPRADOR': '',
        'RAZON_SOCIAL_COMPRADOR': '',
        'DIRECCION_COMPRADOR': '',
        'FECHA_EMISION': '',
        'TOTAL_SIN_IMPUESTOS': 0.0,
        'TOTAL_DESCUENTO': 0.0,
        'PROPINA': 0.0,
        'IMPORTE_TOTAL': 0.0,
        'MONEDA': '',
        'MOTIVO': '',
        'MOTIVO_MODIFICACION': '',
        'MOTIVOS_DESCRIPCION': '',
        'COD_DOC_MODIFICADO': '',
        'NUM_DOC_MODIFICADO': '',
        'FECHA_EMISION_DOC_SUSTENTO': '',
        'VALOR_MODIFICACION': 0.0,
        'VALOR_MODIFICACION_XML': 0.0,
        'VALOR_TOTAL': 0.0,
        'TOTAL_RETENIDO': 0.0,
        'PERIODO_FISCAL': '',
        'EJERCICIO_FISCAL': '',
        'INFO_ADICIONAL_JSON': '',
    }

    detalles_rows = []
    impuestos_rows = []
    pagos_rows = []
    adicionales_rows = []
    retenciones_rows = []
    cabecera_tax_columns = {}

    if info_node is not None:
        cabecera['DIR_ESTABLECIMIENTO'] = _safe_str(info_node.findtext('dirEstablecimiento'))
        cabecera['OBLIGADO_CONTABILIDAD'] = _safe_str(info_node.findtext('obligadoContabilidad'))
        cabecera['COD_DOC_MODIFICADO'] = _safe_str(info_node.findtext('codDocModificado'))
        cabecera['NUM_DOC_MODIFICADO'] = _safe_str(info_node.findtext('numDocModificado'))
        cabecera['FECHA_EMISION_DOC_SUSTENTO'] = _safe_str(info_node.findtext('fechaEmisionDocSustento'))
        motivo_text = _safe_str(info_node.findtext('motivo'))
        if motivo_text:
            cabecera['MOTIVO_MODIFICACION'] = motivo_text
        fecha_emision_text = _safe_str(info_node.findtext('fechaEmision'))
        fecha_emision_dt = _safe_datetime(fecha_emision_text)
        cabecera['FECHA_EMISION'] = (
            fecha_emision_dt.date().isoformat() if fecha_emision_dt else fecha_emision_text
        )
        cabecera['TOTAL_SIN_IMPUESTOS'] = _safe_float(info_node.findtext('totalSinImpuestos'))
        cabecera['TOTAL_DESCUENTO'] = _safe_float(info_node.findtext('totalDescuento'))
        cabecera['PROPINA'] = _safe_float(info_node.findtext('propina'))
        cabecera['IMPORTE_TOTAL'] = _safe_float(info_node.findtext('importeTotal'))
        cabecera['MONEDA'] = _safe_str(info_node.findtext('moneda'))

        if cod_doc == '04':
            cabecera['MOTIVO'] = _safe_str(info_node.findtext('motivo'))
            cabecera['VALOR_MODIFICACION'] = _safe_float(info_node.findtext('valorModificacion'))
            if not cabecera['IMPORTE_TOTAL']:
                cabecera['IMPORTE_TOTAL'] = cabecera['VALOR_MODIFICACION']
        if cod_doc == '05':
            cabecera['VALOR_TOTAL'] = _safe_float(info_node.findtext('valorTotal'))
            if not cabecera['IMPORTE_TOTAL']:
                cabecera['IMPORTE_TOTAL'] = cabecera['VALOR_TOTAL']
        if cod_doc == '07':
            cabecera['PERIODO_FISCAL'] = _safe_str(info_node.findtext('periodoFiscal'))
            cabecera['EJERCICIO_FISCAL'] = _safe_str(info_node.findtext('ejercicioFiscal'))

        motivos = doc_node.findall('.//motivos/motivo')
        if motivos:
            motivos_descripciones = []
            motivos_razones = []
            motivos_valores = []
            for motivo in motivos:
                razon = _safe_str(motivo.findtext('razon'))
                valor_text = _safe_str(motivo.findtext('valor'))
                valor = _safe_float(valor_text)
                if razon:
                    motivos_razones.append(razon)
                if valor_text:
                    motivos_valores.append(valor)
                if razon or valor_text:
                    motivos_descripciones.append(
                        f"Razon: {razon or 'No Disponible'}, Valor: {valor_text or 'No Disponible'}"
                    )
            if motivos_razones:
                cabecera['MOTIVO_MODIFICACION'] = " | ".join(motivos_razones)
            if motivos_valores:
                cabecera['VALOR_MODIFICACION_XML'] = sum(motivos_valores)
            if motivos_descripciones:
                cabecera['MOTIVOS_DESCRIPCION'] = " | ".join(motivos_descripciones)

        for campo in (
            'tipoIdentificacionComprador',
            'tipoIdentificacionProveedor',
            'tipoIdentificacionSujetoRetenido',
        ):
            valor = _safe_str(info_node.findtext(campo))
            if valor:
                cabecera['TIPO_IDENTIFICACION_COMPRADOR'] = valor
                break

        for campo in (
            'identificacionComprador',
            'identificacionProveedor',
            'identificacionSujetoRetenido',
        ):
            valor = _safe_str(info_node.findtext(campo))
            if valor:
                cabecera['IDENTIFICACION_COMPRADOR'] = valor
                break

        for campo in (
            'razonSocialComprador',
            'razonSocialProveedor',
            'razonSocialSujetoRetenido',
        ):
            valor = _safe_str(info_node.findtext(campo))
            if valor:
                cabecera['RAZON_SOCIAL_COMPRADOR'] = valor
                break

        for campo in (
            'direccionComprador',
            'direccionProveedor',
            'direccionSujetoRetenido',
        ):
            valor = _safe_str(info_node.findtext(campo))
            if valor:
                cabecera['DIRECCION_COMPRADOR'] = valor
                break

        totales = info_node.find('totalConImpuestos')
        if totales is not None:
            for imp in totales.findall('totalImpuesto'):
                codigo = _safe_str(imp.findtext('codigo'))
                codigo_pct = _safe_str(imp.findtext('codigoPorcentaje'))
                base = _safe_float(imp.findtext('baseImponible'))
                valor = _safe_float(imp.findtext('valor'))
                tarifa = _safe_float(imp.findtext('tarifa'))
                impuestos_rows.append(
                    {
                        'NIVEL': 'DOCUMENTO',
                        'CLAVE_ACCESO': clave_acceso,
                        'LINE_NUM': None,
                        'CODIGO': codigo,
                        'CODIGO_PORCENTAJE': codigo_pct,
                        'TARIFA': tarifa,
                        'BASE_IMPONIBLE': base,
                        'VALOR': valor,
                    }
                )
                base_col, valor_col = _tax_column_names(codigo, codigo_pct)
                cabecera_tax_columns[base_col] = cabecera_tax_columns.get(base_col, 0.0) + base
                cabecera_tax_columns[valor_col] = cabecera_tax_columns.get(valor_col, 0.0) + valor

        pagos = info_node.find('pagos')
        pago_num = 0
        if pagos is not None:
            for pago in pagos.findall('pago'):
                pago_num += 1
                forma = _safe_str(pago.findtext('formaPago'))
                total = _safe_float(pago.findtext('total'))
                plazo = _safe_str(pago.findtext('plazo'))
                unidad_tiempo = _safe_str(pago.findtext('unidadTiempo'))
                pagos_rows.append(
                    {
                        'CLAVE_ACCESO': clave_acceso,
                        'PAGO_NUM': pago_num,
                        'FORMA_PAGO': forma,
                        'FORMA_PAGO_DESC': FORMA_PAGO_LABEL.get(forma, ''),
                        'TOTAL': total,
                        'PLAZO': plazo,
                        'UNIDAD_TIEMPO': unidad_tiempo,
                    }
                )
                cabecera[f'FORMA_PAGO_{pago_num}'] = forma
                cabecera[f'TOTAL_PAGO_{pago_num}'] = total
                cabecera[f'PLAZO_{pago_num}'] = plazo
                cabecera[f'UNIDAD_TIEMPO_{pago_num}'] = unidad_tiempo

        info_adicional = info_node.find('infoAdicional')
        adicionales_items = []
        if info_adicional is not None:
            for campo in info_adicional.findall('campoAdicional'):
                nombre = _safe_str(campo.attrib.get('nombre'))
                valor = _safe_str(campo.text)
                adicionales_items.append({'nombre': nombre, 'valor': valor})
                adicionales_rows.append(
                    {
                        'CLAVE_ACCESO': clave_acceso,
                        'ARCHIVO_XML': xml_path.name,
                        'NOMBRE': nombre,
                        'VALOR': valor,
                    }
                )
        cabecera['INFO_ADICIONAL_JSON'] = json.dumps(adicionales_items, ensure_ascii=False) if adicionales_items else ''

        detalles = doc_node.findall('.//detalles/detalle')
        for idx, detalle in enumerate(detalles, start=1):
            codigo_principal = _safe_str(detalle.findtext('codigoPrincipal'))
            codigo_aux = _safe_str(detalle.findtext('codigoAuxiliar'))
            descripcion = ' '.join(_safe_str(detalle.findtext('descripcion')).split())
            cantidad = _safe_float(detalle.findtext('cantidad'))
            precio_unitario = _safe_float(detalle.findtext('precioUnitario'))
            descuento = _safe_float(detalle.findtext('descuento'))
            precio_total = _safe_float(detalle.findtext('precioTotalSinImpuesto'))

            detalle_row = {
                'CLAVE_ACCESO': clave_acceso,
                'ARCHIVO_XML': xml_path.name,
                'LINE_NUM': idx,
                'CODIGO_PRINCIPAL': codigo_principal,
                'CODIGO_AUXILIAR': codigo_aux,
                'DESCRIPCION': descripcion,
                'CANTIDAD': cantidad,
                'PRECIO_UNITARIO': precio_unitario,
                'DESCUENTO': descuento,
                'PRECIO_TOTAL_SIN_IMPUESTO': precio_total,
                'CENTRO_COSTO': _safe_str(detalle.findtext('centroCosto')),
                'CTA_CONTABLE': '',
                'IVA_CODIGO_PORCENTAJE': '',
                'IVA_TARIFA': 0.0,
                'IVA_BASE': 0.0,
                'IVA_VALOR': 0.0,
                'ICE_CODIGO_PORCENTAJE': '',
                'ICE_TARIFA': 0.0,
                'ICE_BASE': 0.0,
                'ICE_VALOR': 0.0,
                'IRBPNR_CODIGO_PORCENTAJE': '',
                'IRBPNR_TARIFA': 0.0,
                'IRBPNR_BASE': 0.0,
                'IRBPNR_VALOR': 0.0,
            }

            impuestos_linea = detalle.findall('impuestos/impuesto')
            for imp in impuestos_linea:
                codigo = _safe_str(imp.findtext('codigo'))
                codigo_pct = _safe_str(imp.findtext('codigoPorcentaje'))
                tarifa = _safe_float(imp.findtext('tarifa'))
                base = _safe_float(imp.findtext('baseImponible'))
                valor = _safe_float(imp.findtext('valor'))

                impuestos_rows.append(
                    {
                        'NIVEL': 'LINEA',
                        'CLAVE_ACCESO': clave_acceso,
                        'LINE_NUM': idx,
                        'CODIGO': codigo,
                        'CODIGO_PORCENTAJE': codigo_pct,
                        'TARIFA': tarifa,
                        'BASE_IMPONIBLE': base,
                        'VALOR': valor,
                    }
                )

                etiqueta = IMPUESTO_CODIGO_LABEL.get(codigo)
                if etiqueta == 'IVA':
                    detalle_row['IVA_CODIGO_PORCENTAJE'] = codigo_pct
                    detalle_row['IVA_TARIFA'] = tarifa
                    detalle_row['IVA_BASE'] += base
                    detalle_row['IVA_VALOR'] += valor
                elif etiqueta == 'ICE':
                    detalle_row['ICE_CODIGO_PORCENTAJE'] = codigo_pct
                    detalle_row['ICE_TARIFA'] = tarifa
                    detalle_row['ICE_BASE'] += base
                    detalle_row['ICE_VALOR'] += valor
                elif etiqueta == 'IRBPNR':
                    detalle_row['IRBPNR_CODIGO_PORCENTAJE'] = codigo_pct
                    detalle_row['IRBPNR_TARIFA'] = tarifa
                    detalle_row['IRBPNR_BASE'] += base
                    detalle_row['IRBPNR_VALOR'] += valor

            detalles_rows.append(detalle_row)

        if cod_doc == '07':
            impuestos_ret = doc_node.findall('.//impuestos/impuesto')
            total_retenido = 0.0
            for imp in impuestos_ret:
                codigo = _safe_str(imp.findtext('codigo'))
                codigo_ret = _safe_str(imp.findtext('codigoRetencion'))
                base = _safe_float(imp.findtext('baseImponible'))
                porcentaje = _safe_float(imp.findtext('porcentajeRetener'))
                valor = _safe_float(imp.findtext('valorRetenido'))
                total_retenido += valor
                numero_doc = _safe_str(imp.findtext('numeroDocumento'))
                fecha_doc = _safe_str(imp.findtext('fechaEmisionDocSustento'))
                tipo_doc = _safe_str(imp.findtext('tipoDocumento'))

                retenciones_rows.append(
                    {
                        'CLAVE_ACCESO': clave_acceso,
                        'ARCHIVO_XML': xml_path.name,
                        'TIPO_IMPUESTO': RETENCION_TIPO_IMPUESTO.get(codigo, codigo),
                        'CODIGO_RETENCION': codigo_ret,
                        'BASE_IMPONIBLE': base,
                        'PORCENTAJE_RETENER': porcentaje,
                        'VALOR_RETENIDO': valor,
                        'DOC_SUSTENTO_TIPO': tipo_doc,
                        'DOC_SUSTENTO_ESTAB': estab,
                        'DOC_SUSTENTO_PTO': pto_emi,
                        'DOC_SUSTENTO_SECUENCIAL': numero_doc,
                        'FECHA_EMISION_DOC_SUSTENTO': fecha_doc,
                        'PERIODO_FISCAL': cabecera['PERIODO_FISCAL'],
                    }
                )
            cabecera['TOTAL_RETENIDO'] = total_retenido

    return (
        cabecera,
        detalles_rows,
        impuestos_rows,
        pagos_rows,
        adicionales_rows,
        retenciones_rows,
        None,
        cabecera_tax_columns,
    )


def construir_reporte(
    carpeta_mes: Path,
    excel_salida: Path,
    estado_autorizacion_default: str | None = None,
    xml_files: list[Path] | None = None,
):
    if xml_files is None:
        xml_files = sorted(carpeta_mes.rglob("*.xml"))
    else:
        vistos = set()
        normalizados: list[Path] = []
        for item in xml_files:
            try:
                ruta = item if isinstance(item, Path) else Path(item)
            except Exception:
                continue
            if ruta in vistos:
                continue
            vistos.add(ruta)
            normalizados.append(ruta)
        xml_files = normalizados
    if not xml_files:
        print("No se encontraron archivos XML en la carpeta.")
        return

    cabecera_rows = []
    detalle_rows = []
    impuestos_rows = []
    pagos_rows = []
    adicionales_rows = []
    retenciones_rows = []
    errores_rows = []
    impuesto_columnas = set()

    for xml_path in xml_files:
        (
            cabecera,
            detalles,
            impuestos,
            pagos,
            adicionales,
            retenciones,
            error_entry,
            cabecera_tax_cols,
        ) = _parse_recibido_xml(xml_path)

        if error_entry:
            errores_rows.append(error_entry)
            continue
        if estado_autorizacion_default:
            estado_actual = (cabecera.get("ESTADO_AUTORIZACION") or "").strip()
            if not estado_actual:
                cabecera["ESTADO_AUTORIZACION"] = estado_autorizacion_default

        for key, value in cabecera_tax_cols.items():
            cabecera[key] = value
            impuesto_columnas.add(key)

        cabecera_rows.append(cabecera)
        detalle_rows.extend(detalles)
        impuestos_rows.extend(impuestos)
        pagos_rows.extend(pagos)
        adicionales_rows.extend(adicionales)
        retenciones_rows.extend(retenciones)

    if not cabecera_rows and not errores_rows:
        print("No fue posible procesar los XML (sin datos validos).")
        return

    fecha_ejecucion = datetime.now()

    resumen_counter = Counter()
    fechas_emision = []
    monto_total = 0.0
    monto_retenido = 0.0

    for cab in cabecera_rows:
        cod_doc = cab.get("COD_DOC")
        resumen_counter[DOC_TYPE_COUNTER_KEYS.get(cod_doc, "otros")] += 1
        importe = float(cab.get("IMPORTE_TOTAL") or 0.0)
        if cod_doc == "07":
            monto_retenido += float(cab.get("TOTAL_RETENIDO") or 0.0)
        else:
            monto_total += importe
        fecha_emision = cab.get("FECHA_EMISION")
        if fecha_emision:
            fechas_emision.append(fecha_emision)

    total_documentos = len(cabecera_rows)
    if fechas_emision:
        fechas_ordenadas = sorted(fechas_emision)
        inicio = fechas_ordenadas[0]
        fin = fechas_ordenadas[-1]
        rango_fechas = f"{inicio} a {fin}"
    else:
        rango_fechas = ""
        inicio = datetime.now().date().isoformat()
        fin = inicio

    consulta_id = f"RECIBIDOS_{inicio}"

    monto_total_impt = monto_total + monto_retenido

    if errores_rows:
        primeros = ", ".join(e["archivo_xml"] for e in errores_rows[:5])
        observaciones = f"Errores en {len(errores_rows)} XML: {primeros}"
    else:
        observaciones = ""

    resumen_row = {
        "consulta_id": consulta_id,
        "fecha_ejecucion": fecha_ejecucion,
        "rango_fechas": rango_fechas,
        "total_documentos": total_documentos,
        "facturas": resumen_counter.get("facturas", 0),
        "liquidaciones": resumen_counter.get("liquidaciones", 0),
        "nc": resumen_counter.get("nc", 0),
        "nd": resumen_counter.get("nd", 0),
        "retenciones": resumen_counter.get("retenciones", 0),
        "monto_total_impt": monto_total_impt,
        "observaciones": observaciones,
    }

    df_resumen = pd.DataFrame([resumen_row])
    resumen_drop = [
        "facturas",
        "liquidaciones",
        "nc",
        "nd",
        "retenciones",
        "monto_total_impt",
        "observaciones",
    ]
    df_resumen = df_resumen.drop(columns=[col for col in resumen_drop if col in df_resumen.columns], errors="ignore")
    df_resumen = _renombrar_dataframe(df_resumen)

    base_columns = [
        "ARCHIVO_XML",
        "RUTA_XML",
        "COD_DOC",
        "DESCRIPCION_DOC",
        "CLAVE_ACCESO",
        "SERIE_COMPROBANTE",
        "ESTADO_AUTORIZACION",
        "NUMERO_AUTORIZACION",
        "FECHA_AUTORIZACION",
        "AMBIENTE",
        "VERSION_XML",
        "ESTAB",
        "PTO_EMI",
        "SECUENCIAL",
        "RUC_EMISOR",
        "RAZON_SOCIAL_EMISOR",
        "NOMBRE_COMERCIAL_EMISOR",
        "DIR_MATRIZ",
        "DIR_ESTABLECIMIENTO",
        "TIPO_IDENTIFICACION_COMPRADOR",
        "IDENTIFICACION_COMPRADOR",
        "RAZON_SOCIAL_COMPRADOR",
        "DIRECCION_COMPRADOR",
        "FECHA_EMISION",
        "TOTAL_SIN_IMPUESTOS",
        "TOTAL_DESCUENTO",
        "PROPINA",
        "IMPORTE_TOTAL",
        "MONEDA",
        "MOTIVO",
        "VALOR_MODIFICACION",
        "VALOR_TOTAL",
        "TOTAL_RETENIDO",
        "PERIODO_FISCAL",
        "EJERCICIO_FISCAL",
        "INFO_ADICIONAL_JSON",
    ]

    pago_columns = sorted(
        {
            col
            for cab in cabecera_rows
            for col in cab.keys()
            if col.startswith(("FORMA_PAGO_", "TOTAL_PAGO_", "PLAZO_", "UNIDAD_TIEMPO_"))
        }
    )

    df_cabecera = pd.DataFrame(
        cabecera_rows,
        columns=base_columns + sorted(impuesto_columnas) + pago_columns,
    )
    df_cabecera = df_cabecera.drop(
        columns=[col for col in ("VERSION_XML", "INFO_ADICIONAL_JSON", "PROPINA") if col in df_cabecera.columns],
        errors="ignore",
    )
    df_cabecera = _renombrar_dataframe(df_cabecera)

    detalle_columns = [
        "CLAVE_ACCESO",
        "ARCHIVO_XML",
        "LINE_NUM",
        "CODIGO_PRINCIPAL",
        "CODIGO_AUXILIAR",
        "DESCRIPCION",
        "CANTIDAD",
        "PRECIO_UNITARIO",
        "DESCUENTO",
        "PRECIO_TOTAL_SIN_IMPUESTO",
        "CENTRO_COSTO",
        "CTA_CONTABLE",
        "IVA_CODIGO_PORCENTAJE",
        "IVA_TARIFA",
        "IVA_BASE",
        "IVA_VALOR",
        "ICE_CODIGO_PORCENTAJE",
        "ICE_TARIFA",
        "ICE_BASE",
        "ICE_VALOR",
        "IRBPNR_CODIGO_PORCENTAJE",
        "IRBPNR_TARIFA",
        "IRBPNR_BASE",
        "IRBPNR_VALOR",
    ]
    df_detalle = (
        pd.DataFrame(detalle_rows, columns=detalle_columns)
        if detalle_rows
        else pd.DataFrame(columns=detalle_columns)
    )
    detalle_drop = ["CODIGO_AUXILIAR", "DESCUENTO", "IVA_TARIFA", "IVA_BASE", "IVA_VALOR"]
    df_detalle = df_detalle.drop(
        columns=[col for col in detalle_drop if col in df_detalle.columns],
        errors="ignore",
    )
    df_detalle = _renombrar_dataframe(df_detalle)

    impuestos_columns = [
        "NIVEL",
        "CLAVE_ACCESO",
        "LINE_NUM",
        "CODIGO",
        "CODIGO_PORCENTAJE",
        "TARIFA",
        "BASE_IMPONIBLE",
        "VALOR",
    ]
    df_impuestos = (
        pd.DataFrame(impuestos_rows, columns=impuestos_columns)
        if impuestos_rows
        else pd.DataFrame(columns=impuestos_columns)
    )
    df_impuestos = df_impuestos.drop(columns=[col for col in ("LINE_NUM",) if col in df_impuestos.columns], errors="ignore")
    df_impuestos = _renombrar_dataframe(df_impuestos)

    pagos_columns = [
        "CLAVE_ACCESO",
        "PAGO_NUM",
        "FORMA_PAGO",
        "FORMA_PAGO_DESC",
        "TOTAL",
        "PLAZO",
        "UNIDAD_TIEMPO",
    ]
    df_pagos = (
        pd.DataFrame(pagos_rows, columns=pagos_columns)
        if pagos_rows
        else pd.DataFrame(columns=pagos_columns)
    )
    df_pagos = _renombrar_dataframe(df_pagos, {"TOTAL": "Monto del pago"})

    adicionales_columns = ["CLAVE_ACCESO", "ARCHIVO_XML", "NOMBRE", "VALOR"]
    df_adicionales = (
        pd.DataFrame(adicionales_rows, columns=adicionales_columns)
        if adicionales_rows
        else pd.DataFrame(columns=adicionales_columns)
    )

    retenciones_columns = [
        "CLAVE_ACCESO",
        "ARCHIVO_XML",
        "TIPO_IMPUESTO",
        "CODIGO_RETENCION",
        "BASE_IMPONIBLE",
        "PORCENTAJE_RETENER",
        "VALOR_RETENIDO",
        "DOC_SUSTENTO_TIPO",
        "DOC_SUSTENTO_ESTAB",
        "DOC_SUSTENTO_PTO",
        "DOC_SUSTENTO_SECUENCIAL",
        "FECHA_EMISION_DOC_SUSTENTO",
        "PERIODO_FISCAL",
    ]
    df_retenciones = (
        pd.DataFrame(retenciones_rows, columns=retenciones_columns)
        if retenciones_rows
        else pd.DataFrame(columns=retenciones_columns)
    )
    df_retenciones = _renombrar_dataframe(df_retenciones)

    errores_columns = ["archivo_xml", "clave_acceso", "error"]
    df_errores = (
        pd.DataFrame(errores_rows, columns=errores_columns)
        if errores_rows
        else pd.DataFrame(columns=errores_columns)
    )
    df_errores = _renombrar_dataframe(df_errores)

    diccionarios_rows = []
    for codigo, descripcion in DOC_TYPES.items():
        diccionarios_rows.append({"catalogo": "codDoc", "codigo": codigo, "descripcion": descripcion})
    for codigo, descripcion in IMPUESTO_CODIGO_LABEL.items():
        diccionarios_rows.append({"catalogo": "impuesto.codigo", "codigo": codigo, "descripcion": descripcion})
    for codigo, descripcion in IVA_CODIGO_PORCENTAJE_LABEL.items():
        diccionarios_rows.append({"catalogo": "iva.codigoPorcentaje", "codigo": codigo, "descripcion": descripcion})
    for codigo, descripcion in FORMA_PAGO_LABEL.items():
        diccionarios_rows.append({"catalogo": "formaPago", "codigo": codigo, "descripcion": descripcion})
    df_diccionarios = pd.DataFrame(diccionarios_rows)

    resumen_xml05_columns = [
        "Estado",
        "Número de Autorización",
        "Fecha de Autorización",
        "Ambiente",
        "Razón Social Emisor",
        "Dir. Establecimiento",
        "Obligado Contabilidad",
        "Tipo Identificación Comprador",
        "Identificación Comprador",
        "Tipo Emisión",
        "Nombre Comercial",
        "Código del Documento",
        "Establecimiento",
        "Punto de Emisión",
        "Secuencial",
        "Dirección Matriz",
        "Contribuyente RIMPE",
        "RUC Emisor",
        "Clave de Acceso",
        "Fecha de Emisión",
        "Razón Social Comprador",
        "Dirección Comprador",
        "Moneda",
        "Plazo Pago",
        "Unidad Tiempo Pago",
        "Descripciones",
        "Forma Pago",
        "Total Sin Impuestos",
        "Base Gravada",
        "Base No Gravada",
        "Tarifas IVA",
        "Monto IVA",
        "Total Descuento",
        "Propina",
        "Importe Total",
        "Total Pago",
        "Código Documento Modificado",
        "Número Documento Modificado",
        "Fecha Emisión Doc. Sustento",
        "Motivo",
        "Valor Modificación",
        "Campos Adicionales",
        "Base Gravada 12%",
        "Monto IVA 12%",
    ]

    detalle_por_clave = defaultdict(list)
    for row in detalle_rows:
        clave = row.get("CLAVE_ACCESO")
        descripcion = _safe_str(row.get("DESCRIPCION"))
        if clave and descripcion:
            detalle_por_clave[clave].append(descripcion)

    iva_bases = defaultdict(lambda: defaultdict(float))
    iva_valores = defaultdict(lambda: defaultdict(float))
    for row in impuestos_rows:
        if row.get("NIVEL") != "DOCUMENTO":
            continue
        if row.get("CODIGO") != "2":
            continue
        clave = row.get("CLAVE_ACCESO")
        if not clave:
            continue
        codigo_pct = _safe_str(row.get("CODIGO_PORCENTAJE"))
        iva_bases[clave][codigo_pct] += _safe_float(row.get("BASE_IMPONIBLE"))
        iva_valores[clave][codigo_pct] += _safe_float(row.get("VALOR"))

    pagos_por_clave = defaultdict(list)
    for row in pagos_rows:
        clave = row.get("CLAVE_ACCESO")
        if clave:
            pagos_por_clave[clave].append(row)
    for pagos in pagos_por_clave.values():
        pagos.sort(key=lambda r: r.get("PAGO_NUM") or 0)

    adicionales_por_clave = defaultdict(list)
    for row in adicionales_rows:
        clave = row.get("CLAVE_ACCESO")
        if clave:
            adicionales_por_clave[clave].append(row)

    def _texto_o_no_disponible(valor: str) -> str:
        texto = _safe_str(valor)
        return texto if texto else "No Disponible"

    def _formato_fecha(valor: str) -> str:
        texto = _safe_str(valor)
        if not texto:
            return "No Disponible"
        fecha_dt = _safe_datetime(texto)
        return fecha_dt.strftime("%d/%m/%Y") if fecha_dt else texto

    resumen_xml05_rows = []
    for cab in cabecera_rows:
        clave = cab.get("CLAVE_ACCESO", "")
        iva_base = iva_bases.get(clave, {})
        iva_valor = iva_valores.get(clave, {})
        base_gravada = sum(iva_base.get(code, 0.0) for code in ("2", "3", "4", "5"))
        base_no_gravada = sum(iva_base.get(code, 0.0) for code in ("0", "6", "7"))
        monto_iva = sum(iva_valor.get(code, 0.0) for code in ("2", "3", "4", "5"))
        base_gravada_12 = iva_base.get("2", 0.0)
        monto_iva_12 = iva_valor.get("2", 0.0)
        tarifas_act = [IVA_TARIFA_LABEL[code] for code in ("2", "3", "4", "5") if iva_base.get(code, 0.0) > 0]
        if tarifas_act:
            tarifas_iva = ", ".join(tarifas_act)
        elif base_no_gravada > 0:
            tarifas_iva = "0%"
        else:
            tarifas_iva = "No Disponible"

        pagos = pagos_por_clave.get(clave, [])
        if pagos:
            primero = pagos[0]
            forma_code = _safe_str(primero.get("FORMA_PAGO"))
            forma_desc = _safe_str(primero.get("FORMA_PAGO_DESC")) or FORMA_PAGO_LABEL.get(forma_code, "")
            forma_pago = f"{forma_code or 'No Disponible'} - {forma_desc or 'No Disponible'}"
            plazo_pago = _texto_o_no_disponible(primero.get("PLAZO"))
            unidad_tiempo = _texto_o_no_disponible(primero.get("UNIDAD_TIEMPO"))
        else:
            forma_pago = "No Disponible - No Disponible"
            plazo_pago = "No Disponible"
            unidad_tiempo = "No Disponible"
        total_pago = sum(_safe_float(row.get("TOTAL")) for row in pagos)

        descripcion_items = cab.get("MOTIVOS_DESCRIPCION") or " | ".join(detalle_por_clave.get(clave, []))
        descripcion_items = descripcion_items if descripcion_items else "No Disponible"

        adicionales_items = []
        if cab.get("INFO_ADICIONAL_JSON"):
            try:
                adicionales_items = json.loads(cab.get("INFO_ADICIONAL_JSON") or "")
            except Exception:
                adicionales_items = []
        if adicionales_items:
            campos_adicionales = "; ".join(
                f"{_safe_str(item.get('nombre'))}: {_safe_str(item.get('valor'))}"
                for item in adicionales_items
                if _safe_str(item.get("nombre")) or _safe_str(item.get("valor"))
            )
            if not campos_adicionales:
                campos_adicionales = "No Disponible"
        else:
            campos_adicionales = "No Disponible"

        valor_modificacion = cab.get("VALOR_MODIFICACION_XML") or cab.get("VALOR_MODIFICACION") or 0.0

        resumen_xml05_rows.append(
            {
                "Estado": _texto_o_no_disponible(cab.get("ESTADO_AUTORIZACION")),
                "Número de Autorización": _texto_o_no_disponible(cab.get("NUMERO_AUTORIZACION")),
                "Fecha de Autorización": _texto_o_no_disponible(cab.get("FECHA_AUTORIZACION")),
                "Ambiente": _texto_o_no_disponible(
                    AMBIENTE_LABEL.get(_safe_str(cab.get("AMBIENTE")), _safe_str(cab.get("AMBIENTE")))
                ),
                "Razón Social Emisor": _texto_o_no_disponible(cab.get("RAZON_SOCIAL_EMISOR")),
                "Dir. Establecimiento": _texto_o_no_disponible(cab.get("DIR_ESTABLECIMIENTO")),
                "Obligado Contabilidad": _texto_o_no_disponible(cab.get("OBLIGADO_CONTABILIDAD")),
                "Tipo Identificación Comprador": _texto_o_no_disponible(
                    TIPO_IDENTIFICACION_LABEL.get(
                        _safe_str(cab.get("TIPO_IDENTIFICACION_COMPRADOR")),
                        _safe_str(cab.get("TIPO_IDENTIFICACION_COMPRADOR")),
                    )
                ),
                "Identificación Comprador": _texto_o_no_disponible(cab.get("IDENTIFICACION_COMPRADOR")),
                "Tipo Emisión": _texto_o_no_disponible(
                    TIPO_EMISION_LABEL.get(_safe_str(cab.get("TIPO_EMISION")), _safe_str(cab.get("TIPO_EMISION")))
                ),
                "Nombre Comercial": _texto_o_no_disponible(cab.get("NOMBRE_COMERCIAL_EMISOR")),
                "Código del Documento": _texto_o_no_disponible(
                    DOC_CODE_LABEL.get(_safe_str(cab.get("COD_DOC")), _safe_str(cab.get("COD_DOC")))
                ),
                "Establecimiento": _texto_o_no_disponible(cab.get("ESTAB")),
                "Punto de Emisión": _texto_o_no_disponible(cab.get("PTO_EMI")),
                "Secuencial": _texto_o_no_disponible(cab.get("SECUENCIAL")),
                "Dirección Matriz": _texto_o_no_disponible(cab.get("DIR_MATRIZ")),
                "Contribuyente RIMPE": _texto_o_no_disponible(cab.get("CONTRIBUYENTE_RIMPE")),
                "RUC Emisor": _texto_o_no_disponible(cab.get("RUC_EMISOR")),
                "Clave de Acceso": _texto_o_no_disponible(cab.get("CLAVE_ACCESO")),
                "Fecha de Emisión": _formato_fecha(cab.get("FECHA_EMISION")),
                "Razón Social Comprador": _texto_o_no_disponible(cab.get("RAZON_SOCIAL_COMPRADOR")),
                "Dirección Comprador": _texto_o_no_disponible(cab.get("DIRECCION_COMPRADOR")),
                "Moneda": _texto_o_no_disponible(cab.get("MONEDA")),
                "Plazo Pago": plazo_pago,
                "Unidad Tiempo Pago": unidad_tiempo,
                "Descripciones": descripcion_items,
                "Forma Pago": forma_pago,
                "Total Sin Impuestos": _safe_float(cab.get("TOTAL_SIN_IMPUESTOS")),
                "Base Gravada": base_gravada,
                "Base No Gravada": base_no_gravada,
                "Tarifas IVA": tarifas_iva,
                "Monto IVA": monto_iva,
                "Total Descuento": _safe_float(cab.get("TOTAL_DESCUENTO")),
                "Propina": _safe_float(cab.get("PROPINA")),
                "Importe Total": _safe_float(cab.get("IMPORTE_TOTAL")),
                "Total Pago": total_pago,
                "Código Documento Modificado": _texto_o_no_disponible(cab.get("COD_DOC_MODIFICADO")),
                "Número Documento Modificado": _texto_o_no_disponible(cab.get("NUM_DOC_MODIFICADO")),
                "Fecha Emisión Doc. Sustento": _texto_o_no_disponible(cab.get("FECHA_EMISION_DOC_SUSTENTO")),
                "Motivo": _texto_o_no_disponible(cab.get("MOTIVO_MODIFICACION")),
                "Valor Modificación": _safe_float(valor_modificacion),
                "Campos Adicionales": campos_adicionales,
                "Base Gravada 12%": base_gravada_12,
                "Monto IVA 12%": monto_iva_12,
            }
        )

    df_resumen_xml05 = pd.DataFrame(resumen_xml05_rows, columns=resumen_xml05_columns)

    with pd.ExcelWriter(excel_salida, engine="openpyxl") as writer:
        df_resumen.to_excel(writer, index=False, sheet_name="Resumen")
        df_cabecera.to_excel(writer, index=False, sheet_name="Cabecera")
        df_detalle.to_excel(writer, index=False, sheet_name="Detalle")
        df_impuestos.to_excel(writer, index=False, sheet_name="Impuestos")
        df_pagos.to_excel(writer, index=False, sheet_name="Pagos")
        if not df_retenciones.empty:
            df_retenciones.to_excel(writer, index=False, sheet_name="Retenciones")
        if not df_errores.empty:
            df_errores.to_excel(writer, index=False, sheet_name="Errores_Parsing")
        if not df_resumen_xml05.empty:
            df_resumen_xml05.to_excel(writer, index=False, sheet_name="05")

        wb = writer.book
        if "Resumen" in wb.sheetnames:
            wb["Resumen"].freeze_panes = "A2"
        for nombre in [
            "Cabecera",
            "Detalle",
            "Impuestos",
            "Pagos",
            "Retenciones",
            "Errores_Parsing",
            "05",
        ]:
            if nombre in wb.sheetnames:
                wb[nombre].freeze_panes = "A2"

    _ajustar_columnas_excel(excel_salida)
    print(f"Reporte XML generado: {excel_salida.name}")


def construir_reporte_emitidos(df_emitidos: pd.DataFrame, excel_salida: Path):
    if df_emitidos.empty:
        print("No hay datos de emitidos para generar reporte.")
        return

    numeric_cols = ["Subtotal", "IVA", "Total"]
    for col in numeric_cols:
        if col in df_emitidos.columns:
            df_emitidos[col] = pd.to_numeric(df_emitidos[col], errors="coerce").fillna(0)

    piv = (
        df_emitidos.groupby("Cliente", dropna=False)[numeric_cols]
        .sum()
        .reset_index()
    )

    with pd.ExcelWriter(excel_salida, engine="openpyxl") as xls:
        df_emitidos.to_excel(xls, index=False, sheet_name="Detalle Emitidos")
        piv.to_excel(xls, index=False, sheet_name="Totales por Cliente")

    _ajustar_columnas_excel(excel_salida)
    _insertar_grafico_corporativo(excel_salida)
    print(f"Reporte Emitidos generado: {excel_salida.name}")


def _ajustar_columnas_excel(archivo_excel: Path):
    wb = load_workbook(archivo_excel)
    for ws in wb.worksheets:
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[column].width = max(10, min(max_length + 2, 50))
    wb.save(archivo_excel)


def _insertar_grafico_corporativo(archivo_excel: Path):
    wb = load_workbook(archivo_excel)
    target_sheet = None
    for nombre in ["Totales por Emisor", "Totales por Cliente"]:
        if nombre in wb.sheetnames:
            target_sheet = nombre
            break
    if not target_sheet:
        wb.save(archivo_excel)
        return

    ws = wb[target_sheet]
    if ws.max_row < 2:
        wb.save(archivo_excel)
        return

    chart = BarChart()
    chart.title = "Totales por Emisor/Cliente"
    chart.x_axis.title = "Entidad"
    chart.y_axis.title = "Total ($)"
    chart.height = 10
    chart.width = 20

    data = Reference(ws, min_col=4, min_row=1, max_row=ws.max_row)
    cats = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    azul_sri = "1E4AA8"
    gris_suave = "A0A0A0"
    chart.graphicalProperties = GraphicalProperties(ln=ColorChoice(prstClr=gris_suave))
    chart.graphicalProperties.solidFill = azul_sri
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True

    ws.add_chart(chart, "H2")
    wb.save(archivo_excel)