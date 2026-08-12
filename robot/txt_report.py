"""Modo rapido: reporte armado con el TXT que ofrece el portal del SRI.

El portal de "Comprobantes electronicos recibidos" expone un enlace
**"Descargar reporte"** (`a#frmPrincipal:lnkTxtlistado`) que entrega en TXT
las mismas columnas que muestra la tabla en pantalla. Ese archivo ya trae
clave de acceso, RUC/razon social del emisor, fechas, numero de comprobante
e importe total — o sea, todo lo que el reporte necesita.

El modo rapido aprovecha eso: en vez de abrir cada fila para bajar su XML y
su PDF (decenas de round-trips por pagina), descarga el TXT de cada hoja de
la consulta y arma el Excel con esa informacion. Una consulta de 300
comprobantes pasa de minutos a segundos.

Sobre el alcance del TXT: no esta documentado si el enlace exporta solo la
pagina visible (50 filas) o el resultado completo. Este modulo no asume
ninguna de las dos: descarga el TXT en **cada** pagina y despues deduplica
por clave de acceso. Si el portal exporta todo, las hojas 2..N llegan
repetidas y el dedupe las colapsa; si exporta solo lo visible, cada hoja
aporta sus 50 filas. El resultado es el mismo en ambos casos.

El Excel se construye respetando el encabezado real del TXT — no un esquema
fijo inventado aca. Si el SRI agrega o renombra una columna, aparece igual
en el reporte; el mapeo de alias solo se usa para dar formato (texto forzado
en claves, formato monetario en importes) y para deduplicar.
"""
from __future__ import annotations

import csv
import re
import unicodedata
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from robot._logging import get_logger
from robot.file_utils import _detectar_delimitador, _es_clave


logger = get_logger(__name__)


# Enlace "Descargar reporte" en la pantalla de Comprobantes recibidos.
TXT_LINK_SELECTOR = "a#frmPrincipal\\:lnkTxtlistado"

# Alias de encabezado -> rol logico. Se comparan normalizados (sin acentos,
# mayusculas, espacios colapsados). Solo sirven para formato y dedupe.
_ALIAS_CLAVE = {
    "CLAVE DE ACCESO", "CLAVE ACCESO", "CLAVE_ACCESO", "NUMERO DE AUTORIZACION",
    "NUMERO AUTORIZACION", "AUTORIZACION",
}
_ALIAS_TEXTO = {
    "CLAVE DE ACCESO", "CLAVE ACCESO", "CLAVE_ACCESO", "NUMERO DE AUTORIZACION",
    "NUMERO AUTORIZACION", "AUTORIZACION", "RUC EMISOR", "RUC", "IDENTIFICACION RECEPTOR",
    "IDENTIFICACION", "SERIE COMPROBANTE", "SERIE", "NUMERO COMPROBANTE",
    "NUMERO DE COMPROBANTE", "ESTABLECIMIENTO", "PUNTO DE EMISION", "SECUENCIAL",
}
_ALIAS_NUMERICO = {
    "IMPORTE TOTAL", "VALOR TOTAL", "VALOR SIN IMPUESTOS", "SUBTOTAL", "IVA",
    "MONTO", "TOTAL", "VALOR", "IMPORTE",
}
_ALIAS_FECHA = {
    "FECHA EMISION", "FECHA DE EMISION", "FECHA AUTORIZACION",
    "FECHA DE AUTORIZACION", "FECHA",
}


def _normalizar_encabezado(texto: str) -> str:
    """Normaliza un nombre de columna para compararlo con los alias."""
    base = unicodedata.normalize("NFKD", texto or "").encode("ascii", "ignore").decode("ascii")
    base = re.sub(r"[^A-Za-z0-9 ]+", " ", base).upper()
    return re.sub(r"\s+", " ", base).strip()


def _leer_texto(path: Path) -> str:
    """Lee un TXT del SRI tolerando UTF-8 y Latin-1."""
    datos = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return datos.decode(encoding)
        except UnicodeDecodeError:
            continue
    return datos.decode("utf-8", errors="ignore")


def _parsear_numero(texto: str):
    """Convierte un importe del TXT a float. Devuelve None si no es numerico.

    Tolera los dos formatos que aparecen segun la configuracion regional del
    portal: `1234.56` y `1.234,56`. El ultimo separador presente manda.
    """
    valor = (texto or "").strip()
    if not valor:
        return None
    if not re.fullmatch(r"[-+]?[\d.,]+", valor):
        return None
    if "," in valor and "." in valor:
        if valor.rfind(",") > valor.rfind("."):
            valor = valor.replace(".", "").replace(",", ".")
        else:
            valor = valor.replace(",", "")
    elif "," in valor:
        valor = valor.replace(",", ".")
    try:
        return float(valor)
    except ValueError:
        return None


def descargar_txt_listado(page, destino_dir: Path, nombre: str, timeout: int = 60000):
    """Hace clic en "Descargar reporte" y guarda el TXT como `destino_dir/nombre`.

    Devuelve el `Path` guardado, o None si el enlace no existe o la descarga
    falla (se registra en el log y el flujo continua con las demas paginas).
    """
    link = page.locator(TXT_LINK_SELECTOR)
    try:
        if not link.count():
            logger.warning("Modo rapido: no se encontro el enlace 'Descargar reporte'.")
            return None
    except Exception as err:
        logger.warning(f"Modo rapido: no se pudo consultar el enlace 'Descargar reporte': {err}")
        return None

    try:
        with page.expect_download(timeout=timeout) as descarga_info:
            link.first.click(no_wait_after=True)
        descarga = descarga_info.value
        destino_dir.mkdir(parents=True, exist_ok=True)
        destino = destino_dir / nombre
        descarga.save_as(str(destino))
        return destino
    except Exception as err:
        logger.warning(f"Modo rapido: fallo la descarga del TXT '{nombre}': {err}")
        return None


def parsear_txts(paths) -> tuple[list[str], list[dict]]:
    """Lee uno o varios TXT del SRI y devuelve `(columnas, filas)`.

    `columnas` es el encabezado real del primer TXT que traiga uno. Las filas
    se deduplican por clave de acceso — necesario porque no sabemos si el
    portal exporta la pagina visible o el listado completo (ver docstring del
    modulo). Cuando una fila no tiene clave, se deduplica por su contenido.
    """
    columnas: list[str] = []
    filas: list[dict] = []
    vistos: set = set()

    for path in paths or []:
        ruta = Path(path)
        if not ruta.exists():
            continue
        try:
            contenido = _leer_texto(ruta)
        except Exception as err:
            logger.warning(f"Modo rapido: no se pudo leer '{ruta.name}': {err}")
            continue
        if not contenido.strip():
            continue

        sep = _detectar_delimitador(contenido[:4096])
        lector = csv.reader(contenido.splitlines(), delimiter=sep)
        registros = [fila for fila in lector if any((celda or "").strip() for celda in fila)]
        if not registros:
            continue

        # La primera linea es encabezado salvo que ya traiga una clave de
        # acceso (caso improbable de un TXT sin cabecera).
        primera = registros[0]
        tiene_encabezado = not any(_es_clave(celda) for celda in primera)
        if tiene_encabezado:
            encabezado = [(celda or "").strip() for celda in primera]
            cuerpo = registros[1:]
        else:
            encabezado = [f"Columna {i + 1}" for i in range(len(primera))]
            cuerpo = registros

        # Nombres unicos: si el SRI repite un titulo, se desambigua con sufijo.
        encabezado_unico: list[str] = []
        for idx, nombre in enumerate(encabezado):
            base = nombre or f"Columna {idx + 1}"
            candidato = base
            repeticion = 2
            while candidato in encabezado_unico:
                candidato = f"{base} ({repeticion})"
                repeticion += 1
            encabezado_unico.append(candidato)

        if not columnas:
            columnas = encabezado_unico
        elif encabezado_unico != columnas:
            logger.warning(
                f"Modo rapido: '{ruta.name}' trae un encabezado distinto al de la "
                f"primera hoja; sus columnas se alinean por posicion."
            )

        for registro in cuerpo:
            fila = {}
            for idx, nombre in enumerate(columnas):
                fila[nombre] = (registro[idx] or "").strip() if idx < len(registro) else ""
            # Columnas extra que el encabezado no contemplaba: no se pierden.
            for idx in range(len(columnas), len(registro)):
                extra = f"Columna {idx + 1}"
                if extra not in columnas:
                    columnas.append(extra)
                fila[extra] = (registro[idx] or "").strip()

            clave = next(
                (valor for valor in fila.values() if _es_clave(valor)),
                "",
            )
            huella = clave or "|".join(fila.get(col, "") for col in columnas)
            if huella in vistos:
                continue
            vistos.add(huella)
            filas.append(fila)

    return columnas, filas


def guardar_reporte_txt_excel(columnas: list[str], filas: list[dict], path: Path) -> bool:
    """Escribe el Excel del modo rapido. Devuelve False si no hay filas."""
    if not filas or not columnas:
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "Comprobantes"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    roles = {}
    for nombre in columnas:
        normalizado = _normalizar_encabezado(nombre)
        if normalizado in _ALIAS_TEXTO or normalizado in _ALIAS_CLAVE:
            roles[nombre] = "texto"
        elif normalizado in _ALIAS_NUMERICO:
            roles[nombre] = "numero"
        elif normalizado in _ALIAS_FECHA:
            roles[nombre] = "fecha"
        else:
            roles[nombre] = ""

    for col_idx, nombre in enumerate(columnas, start=1):
        celda = ws.cell(row=1, column=col_idx, value=nombre)
        celda.font = header_font
        celda.fill = header_fill
        celda.alignment = header_align
        rol = roles[nombre]
        if rol == "texto":
            ancho = 52 if _normalizar_encabezado(nombre) in _ALIAS_CLAVE else 20
        elif rol == "numero":
            ancho = 16
        elif rol == "fecha":
            ancho = 20
        else:
            ancho = max(14, min(45, len(nombre) + 4))
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho

    for row_idx, fila in enumerate(filas, start=2):
        for col_idx, nombre in enumerate(columnas, start=1):
            bruto = fila.get(nombre, "")
            rol = roles[nombre]
            if rol == "numero":
                numero = _parsear_numero(bruto)
                if numero is None:
                    celda = ws.cell(row=row_idx, column=col_idx, value=bruto)
                else:
                    celda = ws.cell(row=row_idx, column=col_idx, value=numero)
                    celda.number_format = "#,##0.00"
            else:
                celda = ws.cell(row=row_idx, column=col_idx, value=bruto)
                # Claves de 49 digitos y RUCs deben quedar como texto o Excel
                # los convierte a notacion cientifica y se pierden digitos.
                if rol == "texto" and bruto:
                    celda.number_format = "@"
            celda.alignment = Alignment(vertical="center")

    ws.freeze_panes = "A2"
    wb.save(str(path))
    return True


def construir_reporte_txt(txt_paths, destino: Path) -> tuple[Path | None, int]:
    """Parsea los TXT y escribe el Excel. Devuelve `(path_o_None, n_filas)`."""
    columnas, filas = parsear_txts(txt_paths)
    if not filas:
        return None, 0
    destino.parent.mkdir(parents=True, exist_ok=True)
    if destino.exists():
        try:
            destino.unlink()
        except PermissionError:
            sufijo = 1
            while True:
                candidato = destino.with_name(f"{destino.stem}_{sufijo}{destino.suffix}")
                if not candidato.exists():
                    destino = candidato
                    break
                sufijo += 1
    if not guardar_reporte_txt_excel(columnas, filas, destino):
        return None, 0
    return destino, len(filas)
