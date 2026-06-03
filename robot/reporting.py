"""Generación de reportes Excel a partir de las descargas del SRI.

Cada función `_guardar_reporte_*` toma una colección de filas (list[dict] o
DataFrame) y produce un `.xlsx` con el formato apropiado para el tipo de
comprobante: factura, retención, nota de crédito / débito, recibidos.

`_consolidar_reportes_excel` une varios reportes mensuales en un único
archivo consolidado por tipo.

Todas las funciones devuelven `bool` (True si el archivo se generó OK)
o `Path | None` (en el caso del consolidador). Capturan excepciones de
I/O y de openpyxl y las loguean como warning.

Originalmente vivían en `robot/downloader.py`; extraídas en la
Sub-fase 2c-ii-b del refactor.
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from robot._logging import get_logger
from robot.data_formatters import (
    _emitidos_retencion_default_row,
    _factura_emitidos_default_row,
    _nota_credito_emitidos_default_row,
    _nota_debito_emitidos_default_row,
    _parse_decimal,
    _parse_datetime_local,
)
from robot.report_columns import (
    EMITIDOS_FACTURA_NUMERIC_COLUMNS,
    EMITIDOS_FACTURA_REPORT_COLUMNS,
    EMITIDOS_FACTURA_TEXT_FORCE_COLUMNS,
    EMITIDOS_NOTA_CREDITO_NUMERIC_COLUMNS,
    EMITIDOS_NOTA_CREDITO_REPORT_COLUMNS,
    EMITIDOS_NOTA_CREDITO_TEXT_FORCE_COLUMNS,
    EMITIDOS_NOTA_DEBITO_NUMERIC_COLUMNS,
    EMITIDOS_NOTA_DEBITO_REPORT_COLUMNS,
    EMITIDOS_NOTA_DEBITO_TEXT_FORCE_COLUMNS,
    EMITIDOS_RETENCION_NUMERIC_COLUMNS,
    EMITIDOS_RETENCION_REPORT_COLUMNS,
    EMITIDOS_RETENCION_TEXT_FORCE_COLUMNS,
    PDF_REPORT_COLUMNS,
    RETENCION_REPORT_COLUMNS,
)


logger = get_logger(__name__)


# --------------------------------------------------------------------------- #
# Recibidos — Retención
# --------------------------------------------------------------------------- #
def _guardar_reporte_pdf_retencion_excel(rows: list[dict], excel_path: Path) -> bool:
    if not rows:
        return False
    df = pd.DataFrame(rows)
    for col in RETENCION_REPORT_COLUMNS:
        if col not in df.columns:
            df[col] = ""
    df = df[RETENCION_REPORT_COLUMNS]

    numeric_cols = [
        "Base_Imponible_Ret_IVA",
        "Porcentaje_Ret_IVA",
        "Valor_Retenido_IVA",
        "Base_Imponible_Ret_IR",
        "Porcentaje_Ret_IR",
        "Valor_Retenido_IR",
        "Base_Imponible_Ret_IR_1",
        "Porcentaje_Ret_IR_1",
        "Valor_Retenido_IR_1",
        "Base_Imponible_Ret_IVA_1",
        "Porcentaje_Ret_IVA_1",
        "Valor_Retenido_IVA_1",
    ]

    def _to_text(val):
        if val is None:
            return ""
        if isinstance(val, float):
            if pd.isna(val):
                return ""
            if val.is_integer():
                return str(int(val))
        return str(val).strip()

    def _to_number(val):
        if val is None:
            return ""
        if isinstance(val, float):
            if pd.isna(val):
                return ""
            return val
        if isinstance(val, int):
            return val
        if isinstance(val, str):
            if not val.strip():
                return ""
            parsed = _parse_decimal(val)
            return parsed if parsed is not None else val
        return val

    for col in df.columns:
        if col in numeric_cols:
            df[col] = df[col].map(_to_number)
        else:
            df[col] = df[col].map(_to_text)
    try:
        df.to_excel(excel_path, index=False)
    except Exception:
        return False
    return True


# --------------------------------------------------------------------------- #
# Emitidos — Retención
# --------------------------------------------------------------------------- #
def _guardar_reporte_pdf_retencion_emitidos_excel(rows: list[dict], excel_path: Path) -> bool:
    if not rows:
        return False
    df = pd.DataFrame(rows)
    for col in EMITIDOS_RETENCION_REPORT_COLUMNS:
        if col not in df.columns:
            df[col] = _emitidos_retencion_default_row().get(col, "")
    df = df[EMITIDOS_RETENCION_REPORT_COLUMNS].copy()

    def _to_text(val):
        if val is None:
            return ""
        if isinstance(val, float) and pd.isna(val):
            return ""
        return str(val).strip()

    def _to_number(val):
        if val is None:
            return 0
        if isinstance(val, float):
            if pd.isna(val):
                return 0
            return val
        if isinstance(val, int):
            return val
        parsed = _parse_decimal(str(val))
        return parsed if parsed is not None else 0

    for col in df.columns:
        if col in EMITIDOS_RETENCION_NUMERIC_COLUMNS:
            df[col] = df[col].map(_to_number)
        else:
            df[col] = df[col].map(_to_text)

    try:
        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Retencion")
            ws = writer.sheets["Retencion"]
            for idx, column in enumerate(EMITIDOS_RETENCION_REPORT_COLUMNS, start=1):
                if column in EMITIDOS_RETENCION_TEXT_FORCE_COLUMNS:
                    for row_idx in range(2, ws.max_row + 1):
                        ws.cell(row=row_idx, column=idx).number_format = "@"
                else:
                    for row_idx in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row_idx, column=idx)
                        if cell.value == "0":
                            cell.value = 0
                max_len = len(column)
                for row_idx in range(1, ws.max_row + 1):
                    value = ws.cell(row=row_idx, column=idx).value
                    if value is None:
                        continue
                    max_len = max(max_len, len(str(value)))
                ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 12), 52)
        return True
    except Exception:
        return False


# --------------------------------------------------------------------------- #
# Emitidos — Nota de Crédito
# --------------------------------------------------------------------------- #
def _guardar_reporte_pdf_nota_credito_emitidos_excel(rows: list[dict], excel_path: Path) -> bool:
    if not rows:
        return False
    df = pd.DataFrame(rows)
    for col in EMITIDOS_NOTA_CREDITO_REPORT_COLUMNS:
        if col not in df.columns:
            df[col] = _nota_credito_emitidos_default_row().get(col, "")
    df = df[EMITIDOS_NOTA_CREDITO_REPORT_COLUMNS].copy()

    def _to_text(val):
        if val is None:
            return ""
        if isinstance(val, float) and pd.isna(val):
            return ""
        return str(val).strip()

    def _to_number(val):
        if val is None:
            return 0
        if isinstance(val, float):
            if pd.isna(val):
                return 0
            return val
        if isinstance(val, int):
            return val
        parsed = _parse_decimal(str(val))
        return parsed if parsed is not None else 0

    for col in df.columns:
        if col in EMITIDOS_NOTA_CREDITO_NUMERIC_COLUMNS:
            df[col] = df[col].map(_to_number)
        else:
            df[col] = df[col].map(_to_text)

    try:
        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="04")
            ws = writer.sheets["04"]
            for idx, column in enumerate(EMITIDOS_NOTA_CREDITO_REPORT_COLUMNS, start=1):
                if column in EMITIDOS_NOTA_CREDITO_TEXT_FORCE_COLUMNS:
                    for row_idx in range(2, ws.max_row + 1):
                        ws.cell(row=row_idx, column=idx).number_format = "@"
                max_len = len(column)
                for row_idx in range(1, ws.max_row + 1):
                    value = ws.cell(row=row_idx, column=idx).value
                    if value is None:
                        continue
                    max_len = max(max_len, len(str(value)))
                ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 12), 52)
        return True
    except Exception:
        return False


# --------------------------------------------------------------------------- #
# Emitidos — Nota de Débito
# --------------------------------------------------------------------------- #
def _guardar_reporte_pdf_nota_debito_emitidos_excel(rows: list[dict], excel_path: Path) -> bool:
    if not rows:
        return False
    df = pd.DataFrame(rows)
    for col in EMITIDOS_NOTA_DEBITO_REPORT_COLUMNS:
        if col not in df.columns:
            df[col] = _nota_debito_emitidos_default_row().get(col, "")
    df = df[EMITIDOS_NOTA_DEBITO_REPORT_COLUMNS].copy()

    def _to_text(val):
        if val is None:
            return ""
        if isinstance(val, float) and pd.isna(val):
            return ""
        return str(val).strip()

    def _to_number(val):
        if val is None:
            return 0
        if isinstance(val, float):
            if pd.isna(val):
                return 0
            return val
        if isinstance(val, int):
            return val
        parsed = _parse_decimal(str(val))
        return parsed if parsed is not None else 0

    for col in df.columns:
        if col in EMITIDOS_NOTA_DEBITO_NUMERIC_COLUMNS:
            df[col] = df[col].map(_to_number)
        else:
            df[col] = df[col].map(_to_text)

    try:
        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="05")
            ws = writer.sheets["05"]
            for idx, column in enumerate(EMITIDOS_NOTA_DEBITO_REPORT_COLUMNS, start=1):
                if column in EMITIDOS_NOTA_DEBITO_TEXT_FORCE_COLUMNS:
                    for row_idx in range(2, ws.max_row + 1):
                        ws.cell(row=row_idx, column=idx).number_format = "@"
                max_len = len(column)
                for row_idx in range(1, ws.max_row + 1):
                    value = ws.cell(row=row_idx, column=idx).value
                    if value is None:
                        continue
                    max_len = max(max_len, len(str(value)))
                ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 12), 52)
        return True
    except Exception:
        return False


# --------------------------------------------------------------------------- #
# Emitidos — Factura
# --------------------------------------------------------------------------- #
def _guardar_reporte_pdf_factura_emitidos_excel(rows: list[dict], excel_path: Path) -> bool:
    if not rows:
        return False
    df = pd.DataFrame(rows)
    for col in EMITIDOS_FACTURA_REPORT_COLUMNS:
        if col not in df.columns:
            df[col] = _factura_emitidos_default_row().get(col, "")
    df = df[EMITIDOS_FACTURA_REPORT_COLUMNS].copy()

    def _to_text(val):
        if val is None:
            return ""
        if isinstance(val, float) and pd.isna(val):
            return ""
        return str(val).strip()

    def _to_number(val):
        if val is None:
            return 0
        if isinstance(val, float):
            if pd.isna(val):
                return 0
            return val
        if isinstance(val, int):
            return val
        parsed = _parse_decimal(str(val))
        return parsed if parsed is not None else 0

    for col in df.columns:
        if col in EMITIDOS_FACTURA_NUMERIC_COLUMNS:
            df[col] = df[col].map(_to_number)
        else:
            df[col] = df[col].map(_to_text)

    try:
        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="01")
            ws = writer.sheets["01"]
            for idx, column in enumerate(EMITIDOS_FACTURA_REPORT_COLUMNS, start=1):
                if column in EMITIDOS_FACTURA_TEXT_FORCE_COLUMNS:
                    for row_idx in range(2, ws.max_row + 1):
                        ws.cell(row=row_idx, column=idx).number_format = "@"
                max_len = len(column)
                for row_idx in range(1, ws.max_row + 1):
                    value = ws.cell(row=row_idx, column=idx).value
                    if value is None:
                        continue
                    max_len = max(max_len, len(str(value)))
                ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 12), 52)
        return True
    except Exception:
        return False


# --------------------------------------------------------------------------- #
# Recibidos — Detalle PDF
# --------------------------------------------------------------------------- #
def _guardar_reporte_pdf_excel(rows: list[dict], excel_path: Path) -> bool:
    if not rows:
        return False
    df = pd.DataFrame(rows)
    for col in PDF_REPORT_COLUMNS:
        if col not in df.columns:
            df[col] = ""
    df = df[PDF_REPORT_COLUMNS]
    text_cols = [
        "tipoDocumento",
        "rucEmisor",
        "razonSocialEmisor",
        "nombreComercial",
        "direccionMatrizEmisor",
        "direccionSucursalEmisor",
        "contribuyenteEspecial",
        "agenteRetencion",
        "obligadoContabilidad",
        "tipoContribuyenteRIMPE",
        "numeroComprobante",
        "establecimiento",
        "puntoEmision",
        "secuencial",
        "fechaEmision",
        "fechaAutorizacion",
        "razonSocialComprador",
        "identificacionComprador",
        "direccionComprador",
        "placa",
        "guia",
        "comprobanteModificado",
        "fechaEmisionModificado",
        "razonModificacion",
        "valorModificacion",
        "descripcionesProductos",
        "formaPago",
        "ambiente",
        "emision",
        "claveAcceso",
        "informacionAdicional",
    ]

    def _to_text(val):
        if val is None:
            return ""
        if isinstance(val, float):
            if pd.isna(val):
                return ""
            if val.is_integer():
                return str(int(val))
        return str(val).strip()

    for col in text_cols:
        if col in df.columns:
            df[col] = df[col].map(_to_text)

    numeric_cols = [
        "subtotalTarifaEspecial",
        "subtotal15",
        "subtotal12",
        "subtotal8",
        "subtotal5",
        "subtotal0",
        "subtotalNoObjetoIVA",
        "subtotalExentoIVA",
        "subtotalSinImpuestos",
        "totalDescuento",
        "ivaTarifaEspecial",
        "iva15",
        "iva12",
        "iva8",
        "iva5",
        "ice",
        "irbpnr",
        "propina",
        "valorTotal",
        "valorTotalSinSubsidio",
        "formaPagoMonto",
    ]

    def _to_number(val):
        if val is None:
            return ""
        if isinstance(val, float):
            if pd.isna(val):
                return ""
            return val
        if isinstance(val, int):
            return val
        if isinstance(val, str):
            if not val.strip():
                return ""
            parsed = _parse_decimal(val)
            return parsed if parsed is not None else val
        return val

    for col in numeric_cols:
        if col in df.columns:
            df[col] = df[col].map(_to_number)
    try:
        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Detalle PDF", index=False)
            ws = writer.sheets["Detalle PDF"]
            col_index = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}
            for columna in ("rucEmisor", "numeroComprobante", "establecimiento", "puntoEmision", "secuencial", "claveAcceso"):
                idx = col_index.get(columna)
                if not idx:
                    continue
                col_letter = get_column_letter(idx)
                for row_idx in range(2, ws.max_row + 1):
                    celda = ws[f"{col_letter}{row_idx}"]
                    celda.number_format = "@"
                    if celda.value is not None:
                        celda.value = str(celda.value)
                    celda.alignment = Alignment(horizontal="left")
    except Exception:
        return False
    return True


# --------------------------------------------------------------------------- #
# Consolidador (varios reportes → uno)
# --------------------------------------------------------------------------- #
def _consolidar_reportes_excel(reportes: list[str], destino: Path) -> Path | None:
    rutas = [Path(p) for p in reportes if p and Path(p).exists()]
    if not rutas:
        return None
    dataframes: list[pd.DataFrame] = []
    columnas: list[str] | None = None
    for ruta in rutas:
        try:
            df = pd.read_excel(ruta)
        except Exception as err:
            logger.warning(f"No se pudo leer reporte para consolidar: {ruta} ({err})")
            continue
        if df is None or df.empty:
            continue
        if columnas is None:
            columnas = list(df.columns)
        else:
            for col in df.columns:
                if col not in columnas:
                    columnas.append(col)
        dataframes.append(df)
    if not dataframes or not columnas:
        return None
    for idx, df in enumerate(dataframes):
        for col in columnas:
            if col not in df.columns:
                df[col] = ""
        dataframes[idx] = df[columnas]
    try:
        destino.parent.mkdir(parents=True, exist_ok=True)
        combinado = pd.concat(dataframes, ignore_index=True)
        combinado.to_excel(destino, index=False)
        return destino
    except Exception as err:
        logger.warning(f"No se pudo escribir reporte consolidado: {destino} ({err})")
        return None


# --------------------------------------------------------------------------- #
# Reporte de listado Emitidos (DataFrame de pandas, con estilos)
# --------------------------------------------------------------------------- #
def _guardar_reporte_emitidos_excel(df_emitidos: pd.DataFrame, excel_path: Path, titulo: str = "EMITIDOS") -> bool:
    if df_emitidos.empty:
        return False

    columns_order = [
        "COMPROBANTE",
        "SERIE_COMPROBANTE",
        "CLAVE_ACCESO",
        "FECHA_AUTORIZACION",
        "FECHA_EMISION",
        "VALOR_SIN_IMPUESTOS",
        "IVA",
        "IMPORTE_TOTAL",
    ]
    for col in columns_order:
        if col not in df_emitidos.columns:
            df_emitidos[col] = ""
    df_emitidos = df_emitidos[columns_order].copy()

    def _coerce_decimal_value(val):
        if isinstance(val, (int, float)):
            return float(val)
        parsed = _parse_decimal(val) if isinstance(val, str) else None
        return parsed if parsed is not None else val

    def _coerce_datetime_value(val):
        if isinstance(val, datetime):
            return val
        parsed = _parse_datetime_local(val) if isinstance(val, str) else None
        return parsed if parsed is not None else val

    for columna in ["VALOR_SIN_IMPUESTOS", "IVA", "IMPORTE_TOTAL"]:
        df_emitidos[columna] = df_emitidos[columna].apply(_coerce_decimal_value)
    for columna in ["FECHA_AUTORIZACION"]:
        df_emitidos[columna] = df_emitidos[columna].apply(_coerce_datetime_value)

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        sheet_name = "Emitidos"
        df_emitidos.to_excel(writer, index=False, sheet_name=sheet_name, startrow=1)
        ws = writer.sheets[sheet_name]
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns_order))
        titulo_cell = ws.cell(row=1, column=1, value=titulo)
        titulo_cell.font = Font(bold=True, size=14)
        titulo_cell.alignment = Alignment(horizontal="center", vertical="center")

        header_fill = PatternFill("solid", fgColor="305496")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[2]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.freeze_panes = "A3"

        text_columns = {"COMPROBANTE", "SERIE_COMPROBANTE", "CLAVE_ACCESO", "FECHA_EMISION"}
        numeric_columns = {"VALOR_SIN_IMPUESTOS", "IVA", "IMPORTE_TOTAL"}
        date_columns = {"FECHA_AUTORIZACION"}

        for idx, column in enumerate(columns_order, start=1):
            max_len = len(column)
            for cell_tuple in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=idx, max_col=idx):
                cell = cell_tuple[0]
                valor = cell.value
                if valor is None:
                    continue
                if isinstance(valor, datetime):
                    texto_len = len(valor.strftime("%d/%m/%Y %H:%M"))
                else:
                    texto_len = len(str(valor))
                if texto_len > max_len:
                    max_len = texto_len
            ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 45)

        for columna in text_columns:
            if columna not in columns_order:
                continue
            col_idx = columns_order.index(columna) + 1
            for cell_tuple in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                celda = cell_tuple[0]
                celda.number_format = "@"
                celda.alignment = Alignment(horizontal="left", vertical="center")

        for columna in numeric_columns:
            if columna not in columns_order:
                continue
            col_idx = columns_order.index(columna) + 1
            for cell_tuple in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                celda = cell_tuple[0]
                if isinstance(celda.value, (int, float)):
                    celda.number_format = "#,##0.00"
                    celda.alignment = Alignment(horizontal="right", vertical="center")

        for columna in date_columns:
            if columna not in columns_order:
                continue
            col_idx = columns_order.index(columna) + 1
            for cell_tuple in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                celda = cell_tuple[0]
                if isinstance(celda.value, datetime):
                    celda.number_format = "dd/mm/yyyy hh:mm"
                    celda.alignment = Alignment(horizontal="center", vertical="center")

    return True

